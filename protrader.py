
import os
from kiteconnect import KiteTicker
from kiteconnect import KiteConnect
import datetime
from datetime import timedelta
import pandas as pd
pd.options.mode.chained_assignment = None 
import math,glob
import time,csv,os
import numpy as np
from os import path
import numpy as np
import random
import kitelogin
import xlwings as xw
import gc
import re
from pathlib import Path
import sys 
from py_vollib.black_scholes.implied_volatility import *
from py_lets_be_rational.exceptions import BelowIntrinsicException
from py_vollib.black_scholes.implied_volatility import implied_volatility
from py_vollib.black_scholes.greeks.analytical import delta, gamma, rho, theta, vega
import warnings
def main():
    import os
    from kiteconnect import KiteTicker
    from kiteconnect import KiteConnect
    import datetime
    from datetime import timedelta
    import pandas as pd
    pd.options.mode.chained_assignment = None 
    import math,glob
    import time,csv,os
    import numpy as np
    from os import path
    import numpy as np
    import random
    import kitelogin
    import xlwings as xw
    import gc
    import re
    from pathlib import Path
    import sys 

    warnings.filterwarnings("ignore", category=DeprecationWarning)
    warnings.filterwarnings("ignore", category=FutureWarning)
    global kite,kws;kite = None;kws = None
    global mkt_open_time,mkt_close_time
    mkt_open_time = datetime.datetime(datetime.datetime.now().year, datetime.datetime.now().month, datetime.datetime.now().day,9,14)
    mkt_close_time = datetime.datetime(datetime.datetime.now().year, datetime.datetime.now().month, datetime.datetime.now().day,23,31)
    # QUEUE FOR PUTTING TICKS INFINITELY
    #global EventQ
    #EventQ = queue.Queue()
    Number_OF_OTM_ITM = 20
    Number_OF_Strikes_OTM = 2
    Number_OF_Strikes_ITM = 2
    Data_retention_minute = 5
    #print('CONFIGURATION: # of ITM strikes:',Number_OF_Strikes_ITM,'# of ITM strikes:',Number_OF_Strikes_OTM,'Data_retention_minute:',Data_retention_minute)
    # FOLDER WHERE ALL THE FILES ARE STORED
    global data_folder
    print("Starting pro trader.. All the Best..")
    # Function to apply the new aggregation logicimport os
    import shutil


    def get_column_letter(col_idx_1_based):
        
        col_idx = col_idx_1_based - 1  
        if col_idx < 26:
            return chr(col_idx + 65)  
        else:
        
            first_letter = chr((col_idx // 26) + 64)  
            second_letter = chr((col_idx % 26) + 65)
            return first_letter + second_lettern






    def aggregate_based_on_atm(df):
        df['ATM'] = df['name'].map(atm_values)  # Map ATM values based on index name

        # Aggregation conditions based on CE and PE
        conditions = {
            'buy_itm': (df['CE_PE'] == 'CE') & (df['strike'] <= df['ATM']) | (df['CE_PE'] == 'PE') & (df['strike'] >= df['ATM']),
            'sell_itm': (df['CE_PE'] == 'CE') & (df['strike'] <= df['ATM']) | (df['CE_PE'] == 'PE') & (df['strike'] >= df['ATM']),
            'buy_otm': (df['CE_PE'] == 'CE') & (df['strike'] >= df['ATM']) | (df['CE_PE'] == 'PE') & (df['strike'] <= df['ATM']),
            'sell_otm': (df['CE_PE'] == 'CE') & (df['strike'] >= df['ATM']) | (df['CE_PE'] == 'PE') & (df['strike'] <= df['ATM'])
        }

        # Creating the columns based on conditions
        for key in conditions:
            df[key] = df.loc[conditions[key], 'Totl Bid Qty'] if 'buy' in key else df.loc[conditions[key], 'Totl Ask Qty']

        # Group by and sum the quantities for each condition
        grouped = df.groupby(['name', 'CE_PE', 'expiry']).agg({
            'buy_itm': 'sum',
            'sell_itm': 'sum',
            'buy_otm': 'sum',
            'sell_otm': 'sum'
        }).reset_index()

        # Rename columns to match the output format
        grouped.columns = ['Name', 'CE_PE', 'Expiry', 'Buy_ITM', 'Sell_ITM', 'Buy_OTM', 'Sell_OTM']
        return grouped



    # wb = xw.Book('credentials.xlsx')
    # user_name = wb.sheets['Login and Configuration'].range('B1').value
    # passwd = wb.sheets['Login and Configuration'].range('B2').value
    # login_pin = wb.sheets['Login and Configuration'].range('B3').value
    # client_secret = wb.sheets['Login and Configuration'].range('B4').value
    # api_key = wb.sheets['Login and Configuration'].range('B5').value
    # att,ap = kitelogin.get_access_token(user_name,passwd,login_pin,client_secret,api_key,"pro")
    # kite = KiteConnect(api_key=ap)
    # kite.set_access_token(att)
    # kws = KiteTicker(api_key, att,reconnect=True, reconnect_max_tries=300, reconnect_max_delay=30, connect_timeout=30)
    # profile = kite.profile();print('profile:',profile)
    # print("login over")

    file_path_2 = 'credentials.xlsx'
    df_login = pd.read_excel(file_path_2, sheet_name='Login and Configuration', engine='openpyxl',header=None)

    user_name = df_login.iloc[0, 1]  # B1
    passwd = df_login.iloc[1, 1]     # B2
    login_pin = df_login.iloc[2, 1]  # B3
    client_secret = df_login.iloc[3, 1]  # B4
    api_key = df_login.iloc[4, 1] 

    print("user_name is",user_name)
    print("passwd is",passwd)
    print("login_pin is",login_pin)
    print("client_secret is",client_secret)
    print("api_key is",api_key)
    att, ap = kitelogin.get_access_token(user_name, passwd, login_pin, client_secret, api_key, "pro")
    kite = KiteConnect(api_key=ap)
    kite.set_access_token(att)
    kws = KiteTicker(api_key, att, reconnect=True, reconnect_max_tries=300, reconnect_max_delay=30, connect_timeout=30)
    profile = kite.profile()
    print('profile:', profile)


    def read_last_line_from_file(filename):
        try:
            with open(filename, 'r') as file:
                lines = file.readlines()
                if len(lines) > 1:  # Check if there are at least two lines
                    penultimate_line = lines[-2].strip()  # Get the penultimate line
                    return penultimate_line
                elif lines:  # If the file has exactly one line
                    return "The file has only one line."
                else:
                    return "The file is empty."
        except FileNotFoundError:
            return "The file does not exist."



    curr_dt_rank = time.strftime("%Y%m%d", time.localtime())
    txt_file= "dates_pro.txt"

    def read_line_from_file(filename):
        try:
            with open(filename, 'r') as file:
                lines = file.readlines()  
                if lines:  
                    last_line = lines[-1].strip()  
                    return last_line
                else:
                    return "The file is empty."
        except FileNotFoundError:
            return "The file does not exist."

    last_data=read_line_from_file(txt_file)


    if curr_dt_rank!= last_data:
        print("current date rank is",curr_dt_rank)
        print("last line data is",last_data)
        print("Its a new day and we are starting for the first time, Going to start Algo")
        

        new_day=True
    else:
        print("We are starting the algo again in the same day")
        new_day=False
        


    def is_date_in_last_line(file_path, date):
        try:
            with open(file_path, 'r') as file:
                lines = file.readlines()
                if lines:
                    
                    last_line = lines[-1].strip()
                    return last_line.endswith(date)
                else:
                    return False
        except FileNotFoundError:
            return False


    now = datetime.datetime.now()
    cutoff = now.replace(hour=7, minute=59, second=0, microsecond=0)
    if now < cutoff:
        print("Before 08:00 AM, not executing my_function.")
    else:
        if not is_date_in_last_line(txt_file, curr_dt_rank):
            with open(txt_file, 'a') as file: 
                if os.path.exists(txt_file) and os.path.getsize(txt_file) > 0:
                    
                    file.write(f"\n{curr_dt_rank}")
                else:
                    
                    file.write(curr_dt_rank)
            print(f"Today's date ({curr_dt_rank}) has been added to '{txt_file}'.")
            
        else:
            print("Today's date is already the last line in the file. No changes made.")
            


    last_line_value = read_last_line_from_file(txt_file)


    def aggregate_based_on_atm(df,atm_values):
        df['ATM'] = df['name'].map(atm_values)  # Map ATM values based on index name

        # Aggregation conditions based on CE and PE
        conditions = {
            'buy_itm': (df['CE_PE'] == 'CE') & (df['strike'] <= df['ATM']) | (df['CE_PE'] == 'PE') & (df['strike'] >= df['ATM']),
            'sell_itm': (df['CE_PE'] == 'CE') & (df['strike'] <= df['ATM']) | (df['CE_PE'] == 'PE') & (df['strike'] >= df['ATM']),
            'buy_otm': (df['CE_PE'] == 'CE') & (df['strike'] >= df['ATM']) | (df['CE_PE'] == 'PE') & (df['strike'] <= df['ATM']),
            'sell_otm': (df['CE_PE'] == 'CE') & (df['strike'] >= df['ATM']) | (df['CE_PE'] == 'PE') & (df['strike'] <= df['ATM'])
        }

        # Creating the columns based on conditions
        for key in conditions:
            df[key] = df.loc[conditions[key], 'Totl Bid Qty'] if 'buy' in key else df.loc[conditions[key], 'Totl Ask Qty']

        # Group by and sum the quantities for each condition
        grouped = df.groupby(['name', 'CE_PE', 'expiry']).agg({
            'buy_itm': 'sum',
            'sell_itm': 'sum',
            'buy_otm': 'sum',
            'sell_otm': 'sum'
        }).reset_index()

        # Rename columns to match the output format
        grouped.columns = ['Name', 'CE_PE', 'Expiry', 'Buy_ITM', 'Sell_ITM', 'Buy_OTM', 'Sell_OTM']
        return grouped


    def get_historic_data_new(name, timeframe, delta):
      print('Hist Extracting for:',name)
      for i in range(0,1000):
        while True:
         try:
            
            if name.startswith('SENSEX') or name.startswith('BANKEX'):
                Instrument_type = "BFO"
            else:
                Instrument_type = "NFO"


            to_date = datetime.datetime.now().date()
            from_date = to_date - datetime.timedelta(days=int(delta))
            token = kite.ltp([Instrument_type+':' + name])[Instrument_type+':' + name]['instrument_token']

            data = kite.historical_data(token, from_date, to_date, timeframe,oi=True)
            data = pd.DataFrame(data)

            if len(data.index) >=0:
                data['symbol'] = name
                return data
            else:
                return pd.DataFrame()
         except Exception as e:
            print("exception occured get_historic_data:" + str(e),i);
            #ogfile.flush(); os.fsync(#ogfile.fileno())
            time.sleep(5);err_msg = 'single positional indexer is out-of-bounds'
            if (err_msg in str(e) and i == 2):
                data = pd.DataFrame()
                return data
            elif ( i == 3):
                data = pd.DataFrame()
                return data
            #pass
         break 


    def update_option_chain(df,index_name, sheet_name,expiry_today):
        print("index name is",index_name)
        print("exp today is",expiry_today)
        if index_name in expiry_today:
            file_map = {
                'NIFTY': 'NF_Pro_Trading.xlsx',
                'BANKNIFTY': 'BNF_Pro_Trading.xlsx',
                'FINNIFTY': 'FIN_Pro_Trading.xlsx',
                'MIDCPNIFTY': 'Midcp_Pro_Trading.xlsx',
                'SENSEX':'Sen_Pro_Trading.xlsx',
                'BANKEX':'BAN_Pro_Trading.xlsx',
                    
            }
            if index_name not in file_map:
                file_name = file_map['BANKNIFTY']
            else:
                file_name = file_map[index_name]
            full_path = os.path.join(os.getcwd(), file_name)

            if xw.apps.count == 0:
                app = xw.App(visible=True)
            else:
                app = xw.apps.active

            try:
                wb = app.books[file_name]
            except:
                try:
                    wb = app.books.open(full_path)
                except FileNotFoundError:
                    wb = app.books.add()
                    wb.save(full_path)
            if index_name not in file_map:
                sheet = wb.sheets['BNF_ProTrading_CW']
                df = df.loc[:, ~df.columns.str.startswith('Unnamed')]  
                if index_name=='HDFC':                 
                    sheet.range('A47').options(index=False, header=False).value = df
                elif index_name=='ICICI':                 
                    sheet.range('N47').options(index=False, header=False).value = df
                elif index_name=='SBI':                 
                    sheet.range('A54').options(index=False, header=False).value = df
                elif index_name=='KOTAK':                 
                    sheet.range('N54').options(index=False, header=False).value = df
                elif index_name=='AXIS':                 
                    sheet.range('A61').options(index=False, header=False).value = df
                elif index_name=='INDUSIND':                 
                    sheet.range('N61').options(index=False, header=False).value = df

            else:

                if sheet_name not in [sheet.name for sheet in wb.sheets]:
                    wb.sheets.add(sheet_name)
                sheet = wb.sheets[sheet_name]

                df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
                # Start from A6 and do not include the DataFrame header
                sheet.range('A6').options(index=False, header=False).value = df

            wb.save()
        else:
            print("Non expiry index")


    def calculate_averages(historical_data):

        if len(historical_data) > 3:
            relevant_data = historical_data.iloc[-4:-1]  # Get the three rows before the last row
            avg_close = relevant_data['close'].mean()
            
        else:
            avg_close = historical_data['close'].mean()  # Fallback to using all data if <4 rows
        
        
        return avg_close

    def process_dataframe(results_dict):

        processed_results = {}
        for key, odf in results_dict.items():
            odf['3min_hist_LTP'] = pd.NA
            
            
            for index, row in odf.iterrows():
                historical_data = get_historic_data_new(row['Stock'],'minute',10)
                print("historical data over")
                avg_close = calculate_averages(historical_data)
                
                # Update the DataFrame with the calculated averages
                odf.at[index, '3min_hist_LTP'] = avg_close
            
            
            processed_results[key] = odf
        
        return processed_results

    def remove_duplicates_from_results(results):
        cleaned_results = {}
        for key, df in results.items():
            cleaned_df = df.drop_duplicates()
            cleaned_results[key] = cleaned_df
        return cleaned_results

    def rearrange_dfs(dfs_dict, time_value):
        transformed_dfs = {}
        for key, df in dfs_dict.items():

            df['LTP_Chng'] = df['LTP']-df['3min_hist_LTP']  
            df['Time'] = time_value

            df_transformed = df[['Time', 'strike', 'CE_PE','type','vega', 'vega_change', 'theta', 'theta_change', 
                                'delta', 'delta_change', '3min_hist_LTP', 'LTP', 'LTP_Chng', 
                                'Act_Vol']].copy()

            df_transformed.columns = ['Time', 'Strike','CE_PE','type','Vega_CE/PE', 'Vega_Chng', 'Theta_CE/PE', 'Theta_Chng', 
                                    'Delta_CE/PE', 'Delta_Chng', '3min_hist_LTP', 'LTP', 'LTP_Chng', 
                                    'Act_Vol']
            
            # Update the transformed DataFrame in the new dictionary
            transformed_dfs[key] = df_transformed
        
        return transformed_dfs

    def calculate_strikes_two(atm_value, strike_diff, option_type,exp):
        if exp:
            strikes = []
            if option_type.upper() == 'CE':
                start_strike = atm_value - (strike_diff * 1)  
            else:  
                start_strike = atm_value + (strike_diff * 1)  

            for i in range(3):  
                if option_type.upper() == 'CE':
                    strikes.append(start_strike + (i * strike_diff))
                else: 
                    strikes.append(start_strike - (i * strike_diff))
            
            return strikes
        else:
            strikes = []
            if option_type.upper() == 'CE':
                start_strike = atm_value - (strike_diff * 2)  
            else:  
                start_strike = atm_value + (strike_diff * 2)  

            for i in range(5):  
                if option_type.upper() == 'CE':
                    strikes.append(start_strike + (i * strike_diff))
                else: 
                    strikes.append(start_strike - (i * strike_diff))
            
            return strikes

    def classify_strikes(strikes, atm, option_type):
        
        strike_map = {}
        for strike in strikes:
            if strike == atm:
                strike_map[strike] = 'ATM'
            elif strike < atm:
                strike_map[strike] = 'ITM' if 'CE' in option_type else 'OTM'
            else:
                strike_map[strike] = 'OTM' if 'CE' in option_type else 'ITM'
        return strike_map



    def process_conditions_separately(current_df, previous_df,tm,atm_values,today_expiries):
        indices = ["BANKNIFTY", "NIFTY", "FINNIFTY", "MIDCPNIFTY", "SENSEX","BANKEX"]
        results = {}
        
        diff={'BANKNIFTY': 100, 'NIFTY': 50, 'FINNIFTY': 50, 'MIDCPNIFTY': 25, 'SENSEX': 100, 'BANKEX': 100}

        for index in indices:
            expiries = sorted(current_df[current_df['name'] == index]['expiry'].unique())
            is_expiry_day = index in today_expiries

            for i, expiry in enumerate(expiries):
                print("expiry is",expiry)
                expiry_key = 'current' if i == 0 else 'next'
                combined_results = {f"{expiry_key}_increase": [], f"{expiry_key}_decrease": []}

                for option_type in ["CE", "PE"]:
                    atm=atm_values[index]
                    print("index is",index)
                    print("atm is",atm)
                    strike_diff=diff[index]
                    print("strike diff is",strike_diff)
                    if is_expiry_day:
                        strikes=calculate_strikes_two(atm, strike_diff, option_type,True)
                    else:
                        strikes=calculate_strikes_two(atm, strike_diff, option_type,False)

                    strike_classifications = classify_strikes(strikes, atm, option_type)
                    print("strikes are",strikes)
                    print("strike dict",strike_classifications)
                    print(current_df.columns)
                    current_filtered = current_df[(current_df['name'] == index) & (current_df['CE_PE'] == option_type) & (current_df['expiry'] == expiry)& (current_df['strike'].isin(strikes))]
                    print("current filtered over")
                    print(previous_df.columns)

                    previous_filtered = previous_df[(previous_df['name'] == index) & (previous_df['CE_PE'] == option_type) & (previous_df['expiry'] == expiry)& (previous_df['strike'].isin(strikes))]
                    print("pre filtered over")
                    
                    current_filtered['type']=current_filtered['strike'].map(strike_classifications)
                    previous_filtered['type']=previous_filtered['strike'].map(strike_classifications)
                    current_filtered.to_csv("current_filtered.csv")
                    previous_filtered.to_csv("previous_filtered.csv")
                    print("previous filtered over")
                    
                    if not current_filtered.empty and not previous_filtered.empty:
                        combined_df = pd.merge(current_filtered, previous_filtered, on=["Stock", "type"], suffixes=('_curr', '_prev'))
                        combined_df['vega_change'] = combined_df['vega_curr'] - combined_df['vega_prev']
                        combined_df['theta_change'] = combined_df['theta_curr'] - combined_df['theta_prev']
                        combined_df['delta_change'] = combined_df['delta_curr'] - combined_df['delta_prev']
                        
                        vega_percent = 0.05 if is_expiry_day else 0.10
                        theta_delta_points = 20 if is_expiry_day else 10

                        if (index == "SENSEX" or index =="BANKEX")and is_expiry_day:
                            theta_delta_points = 30 

                    
                        for suffix in ['increase', 'decrease']:
                            conditions = [
                                ('vega', combined_df['vega_change'] / combined_df['vega_prev'] > vega_percent, suffix == 'increase'),
                                ('theta', combined_df['theta_change'] < -theta_delta_points, suffix == 'increase'),
                                ('delta', combined_df['delta_change'] > theta_delta_points, suffix == 'increase'),
                                ('vega', combined_df['vega_change'] / combined_df['vega_prev'] <= -vega_percent, suffix == 'decrease'),
                                ('theta', combined_df['theta_change'] >= theta_delta_points, suffix == 'decrease'),
                                ('delta', combined_df['delta_change'] <= -theta_delta_points, suffix == 'decrease'),
                            ]

                            for column, condition, applicable in conditions:
                                if applicable:
                                    temp_df = combined_df[condition].copy()
                                    # Initialize all columns to NA, then selectively assign values for the current condition
                                    for col in ['delta_curr', 'theta_curr', 'vega_curr', 'delta_change', 'theta_change', 'vega_change']:
                                        if not (col.startswith(column) or col == f'{column}_change'):
                                            temp_df[col] = pd.NA
                                    combined_results[f"{expiry_key}_{suffix}"].append(temp_df)

                # Combine results for each condition across CE and PE and finalize DataFrames
                for key, condition_result in combined_results.items():
                    if condition_result:
                        final_df = pd.concat(condition_result).drop_duplicates(subset=['Stock'])
                        final_columns = ['Stock', 'name_curr', 'strike_curr','type', 'CE_PE_curr', 'LTP_curr',
                                        'delta_curr', 'theta_curr', 'vega_curr', 'vega_change',
                                        'theta_change', 'delta_change', 'Vol_curr']
                        final_df = final_df[final_columns].rename(columns={
                            'name_curr': 'name',
                            'strike_curr': 'strike',
                            'CE_PE_curr': 'CE_PE',
                            'LTP_curr': 'LTP',
                            'delta_curr': 'delta',
                            'theta_curr': 'theta',
                            'vega_curr': 'vega',
                            'Vol_curr': 'Act_Vol'
                        })
                        results[f"{index.lower()}_{key}"] = final_df

                # For indices with only one expiry, break after processing
                if index in ["FINNIFTY", "MIDCPNIFTY", "SENSEX","BANKEX"] and i == 0:
                    break

        results=remove_duplicates_from_results(results)
        results=process_dataframe(results)
        results=rearrange_dfs(results,tm)

        return results


    # def #ogfileInit():
        
    #     dir=fr'.\\output logs\\'
    
    #     if not os.path.exists(dir):
    #                 print('creating #ogfile directory')
    #                 os.makedirs(dir)
                    
    #     nowtime=time.strftime('%d-%m-%Y__%H_%M', time.localtime())
    #     #ogfilename=dir+'Greek_ALGO_'+nowtime+'.log'
    #     global #ogfile
        
    #     #ogfile = open(#ogfilename, "a+")
    #     sys.stdout = #ogfile
    #     print('log file created')
    #     print(#ogfilename,'\n')
    #     return #ogfile


    print("Check the #ogfiles for further logs of the day")
    # #ogfile=#ogfileInit()
    curr_dt_rank = time.strftime("%Y%m%d", time.localtime())
    txt_file= "dates.txt"





    def read_last_line_from_file(filename):
        try:
            with open(filename, 'r') as file:
                lines = file.readlines()
                if len(lines) > 1:  # Check if there are at least two lines
                    penultimate_line = lines[-2].strip()  # Get the penultimate line
                    return penultimate_line
                elif lines:  # If the file has exactly one line
                    return "The file has only one line."
                else:
                    return "The file is empty."
        except FileNotFoundError:
            return "The file does not exist."

    def standardize_date_format(df, date_column):
        
        try:
            
            df[date_column] = pd.to_datetime(df[date_column], dayfirst=True).dt.strftime('%Y-%m-%d')
        except:
            
            pass
        return df



    def read_line_from_file(filename):
        try:
            with open(filename, 'r') as file:
                lines = file.readlines()  
                if lines:  
                    last_line = lines[-1].strip()  
                    return last_line
                else:
                    return "The file is empty."
        except FileNotFoundError:
            return "The file does not exist."

    last_data=read_line_from_file(txt_file)




    def is_date_in_last_line(file_path, date):
        try:
            with open(file_path, 'r') as file:
                lines = file.readlines()
                if lines:
                    
                    last_line = lines[-1].strip()
                    return last_line.endswith(date)
                else:
                    return False
        except FileNotFoundError:
            return False








    def check_timeframes(tm):
        # Define start times and increments for each timeframe
        timeframes = {
            '2_min': ('0917', 2),
            '3_min': ('0918', 3),
            '5_min': ('0920', 5),
            '10_min': ('0925', 10),
            '15_min': ('0930', 15),
            '30_min': ('0945', 30),
        }
        
        # Function to check if tm is within the interval
        def is_within_interval(tm, start_time, increment):
            tm_minutes = int(tm[:2]) * 60 + int(tm[2:])
            start_minutes = int(start_time[:2]) * 60 + int(start_time[2:])
            return (tm_minutes >= start_minutes) and ((tm_minutes - start_minutes) % increment == 0)
        
        # Check tm against all timeframes
        results = {tf: is_within_interval(tm, start, inc) for tf, (start, inc) in timeframes.items()}
        
        return results
    def on_close(ws, code, reason):
        # CLOSES THE CONNECTION FOR THIS SESSION
        ws.stop()
    def on_ticks(ws, ticks):
        global mkt_open_time, mkt_close_time,instruments_list,instruments_list_Full #,EventQ
        # PUTTING TICKS IN QUEUE DURING MARKET HOURS
        New_instruments = list(set(instruments_list).difference(instruments_list_Full))
        if(len(New_instruments) > 0 ):
            print("New subscription found ... ",(New_instruments),datetime.datetime.now())
            #ogfile.flush(); os.fsync(#ogfile.fileno())
            try:
                ws.subscribe(New_instruments)
                ws.set_mode(ws.MODE_FULL, New_instruments)
                instruments_list_Full = instruments_list_Full + New_instruments
            except Exception as e:
                print (e)

        if True: 
            global df,df_tick_dump
            if True:
                    df_tick1 = pd.DataFrame(ticks)
                    df_tick = df_tick1[(df_tick1.tradable == True)]
                    if len(df_tick.index) >0:
                        if 'tradable' in df_tick and 'mode' in df_tick and 'oi' in df_tick and 'oi_day_high' in df_tick and 'oi_day_low' in df_tick:
                            df_tick = df_tick.drop(['tradable','mode','oi_day_high','oi_day_low'], axis=1)
                            df_tick = df_tick.drop(['depth'], axis=1)
                            df_tick = df_tick.dropna() #drop null calue rows
                            #print('df_tick:',df_tick);
                            #time.sleep(1)
                            lendf = len(df_tick_dump.index)
                            if not lendf < 50000:
                                df_tick_dump = pd.DataFrame()
                            df_tick_dump = pd.concat([df_tick_dump, df_tick], ignore_index=True, axis=0)
                    
    # FUNCTION TO SUBSCRIBE THE SCRIPTS FOR LIVE FEED
    def on_connect(ws, response):
        global token_dict,instruments_list
        #instruments = list(token_dict.keys()); #print('instruments:',instruments);exit()
        #ws.subscribe(instruments)
        #print('Subscribing inst list:',instruments_list)
        ws.subscribe(instruments_list)
        ws.set_mode(ws.MODE_FULL, instruments_list)
        # with open("ins.txt", "w") as file:
        #     for instrument in instruments_list:
        #         file.write("%s\n" % instrument)
        # exit()
    # Callback when reconnect is on progress
    def on_reconnect(ws, attempts_count):
        print("Reconnecting: {}".format(attempts_count))
        #ogfile.flush(); os.fsync(#ogfile.fileno())
        #print("Reconnecting: {}".format(attempts_count))
    # Callback when connection closed with error.
    def on_error(ws, code, reason):
        print("Connection error: {code} - {reason}".format(code=code, reason=reason))
        #ogfile.flush(); os.fsync(#ogfile.fileno())
        #print("Connection error: {code} - {reason}".format(code=code, reason=reason))

    def greeks(premium, expiry, asset_price, strike_price, intrest_rate, instrument_type):
        for i in range(0,100):
            while True:
                try:
                    t = ((datetime.datetime(expiry.year, expiry.month, expiry.day, 15,
                                            30) - datetime.datetime.now()) / datetime.timedelta(days=1)) / 365
                    S = asset_price
                    K = float(strike_price)
                    r = intrest_rate
                    if premium == 0 and False:
                        print('1')
                        if t <= 0 and False:
                            print('2')
                        if S <= 0 and False:
                            print('3')
                        if K <= 0 and False:
                            print('4')
                        if r <= 0 and False:
                            print('5')
                    if premium == 0 or t <= 0 or S <= 0 or K <= 0 or r <= 0:
                        raise Exception
                    flag = instrument_type[0].lower()
                    imp_v = implied_volatility(premium, S, K, t, r, flag)
                    #print('Conditoin satisfied',type(premium) ,type(t) ,type(S) ,type(K) , type(r))
                    #print('imp_v:',imp_v,flag,premium, expiry, asset_price, strike_price, intrest_rate, instrument_type,'premium ,t ,S ,K , r:',premium ,t ,S ,K , r);
                    #return [imp_v,delta(flag, S, K, t, r, imp_v),gamma(flag, S, K, t, r, imp_v),rho(flag, S, K, t, r, imp_v),theta(flag, S, K, t, r, imp_v),vega(flag, S, K, t, r, imp_v)]
                    #print([imp_v,delta(flag, S, K, t, r, imp_v),gamma(flag, S, K, t, r, imp_v),rho(flag, S, K, t, r, imp_v),theta(flag, S, K, t, r, imp_v),vega(flag, S, K, t, r, imp_v)]);exit()
                    a=imp_v*100;b=delta(flag, S, K, t, r, imp_v)*100;c=gamma(flag, S, K, t, r, imp_v);d=rho(flag, S, K, t, r, imp_v);e=theta(flag, S, K, t, r, imp_v);f=vega(flag, S, K, t, r, imp_v)
                    if b <5:
                        h=1
                    #ogfile.flush(); os.fsync(#ogfile.fileno())
                    

                    
                    val = [a,b,c,d,e,f]
                    if len(val) !=6:
                        print('Returned less values.!!')
                        #ogfile.flush(); os.fsync(#ogfile.fileno())
                        print(val);exit()
                        return a,b,c,d,e,f
                except BelowIntrinsicException:
                    return 0,0,0,0,0,0
                except Exception as e:
                    #print("exception occured greeks:" + str(e))
                    return 0,0,0,0,0,0


    def filter_and_prepare_data_fixed(df, index_names, expiry_dates):
    
        expiry_dates_converted = [pd.to_datetime(date).strftime('%d-%m-%Y') for date in expiry_dates]
        
        df_selected = df[['Index', 'expiry', 'call_put', 'Greek_type', 'Greek_1529_y']].rename(columns={'Greek_1529_y': 'Greek_1529'})

        filtered_df = df_selected[(df_selected['Index'].isin(index_names)) & (df_selected['expiry'].isin(expiry_dates_converted))]

        for index_name in index_names:
            for expiry_date in expiry_dates_converted:
                if not ((filtered_df['Index'] == index_name) & (filtered_df['expiry'] == expiry_date)).any():
                    
                    placeholder_data = []
                    for greek_type in ['delta', 'theta', 'iv', 'vega']:
                        placeholder_data.append([index_name, expiry_date, 'CE', greek_type, 0])
                        placeholder_data.append([index_name, expiry_date, 'PE', greek_type, 0])

                    placeholder_rows = pd.DataFrame(placeholder_data, columns=['Index', 'expiry', 'call_put', 'Greek_type', 'Greek_1529'])
                    filtered_df = pd.concat([filtered_df, placeholder_rows], ignore_index=True)
        
        return filtered_df




    def fill_missing_values_with_adjustment(df):
        
        def adjust_value(value):
            percentage = random.uniform(0.1, 0.3)  
            adjust = value * percentage
            return value + adjust if random.choice([True, False]) else value - adjust


        last_col = df.columns[-1]
        second_last_col = df.columns[-2]

        for index, row in df.iterrows():
            if pd.isna(row[last_col]):
                df.at[index, last_col] = adjust_value(row[second_last_col])

        return df


    def get_greeks(option,option_ltp,strike,Index_LTP_N,Index_LTP_BN,Index_LTP_mid,Index_LTP_fin,Index_LTP_sen,Index_LTP_ban,ltp_hdfc,ltp_icici,ltp_sbi,ltp_axis,ltp_kotak,ltp_indus,CE_PE,expiry):
        #print('In get_greeks:',option,option_ltp,strike,Index_LTP_N,Index_LTP_BN,CE_PE,expiry)
        for i in range(0,100):
            while True:
                try:
                    #return [0,0,0,0,0,0]
                    if 'FUT' in option:
                        return 0,0,0,0,0,0
                    if ( option[:5] == 'BANKN' ):
                        Index_LTP = Index_LTP_BN
                    elif ( option[:5] == 'NIFTY' ):
                        Index_LTP = Index_LTP_N
                    elif ( option[:5] == 'MIDCP' ):
                     Index_LTP = Index_LTP_mid
                    elif ( option[:5] == 'FINNI' ):
                        Index_LTP = Index_LTP_fin
                    elif ( option[:5] == 'SENSE' ):
                        Index_LTP = Index_LTP_sen
                    elif ( option[:5] == 'BANKE' ):
                        Index_LTP = Index_LTP_ban
                    elif ( option[:5] == 'HDFCB' ):
                        Index_LTP = ltp_hdfc
                    elif ( option[:5] == 'ICICI' ):
                        Index_LTP = ltp_icici
                    elif ( option[:4] == 'SBIN' ):
                        Index_LTP = ltp_sbi
                    elif ( option[:5] == 'AXISB' ):
                        Index_LTP = ltp_axis
                    elif ( option[:5] == 'KOTAK' ):
                        Index_LTP = ltp_kotak
                    elif ( option[:5] == 'INDUS' ):
                        Index_LTP = ltp_indus

                    if True:
                                                                current_min = datetime.datetime.strptime(expiry, '%Y-%m-%d')
                                                                #   print(f"{current_min}-expiry, subol-{option}")
                                                                return greeks(premium=option_ltp,
                                                                expiry=current_min,
                                                                asset_price=Index_LTP,
                                                                strike_price=strike,
                                                                intrest_rate=0.1,
                                                                instrument_type=CE_PE)
                except Exception as e:
                    print("exception occured on get_greeks:" + str(e))
                    #ogfile.flush(); os.fsync(#ogfile.fileno())
                    time.sleep(1)
                    if i == 9:
                        return 0
                break


    def subtract_one_minute(tm: str) -> str:
        time_obj = datetime.datetime.strptime(tm, '%H%M')
        new_time_obj = time_obj - datetime.timedelta(minutes=1)
        new_tm = new_time_obj.strftime('%H%M')
        
        return new_tm

    def get_N_BN_spot():
        for i in range(0,100):
            while True:
                try:
                    ltp_nifty = 0;ltp_Banknifty = 0;
                    ohlc_nifty=kite.ohlc('NSE:{}'.format('NIFTY 50'))
                    ltp_nifty = ohlc_nifty['NSE:{}'.format('NIFTY 50')]['last_price']  
                    ohlc_Banknifty=kite.ohlc('NSE:{}'.format('NIFTY BANK'))
                    ltp_Banknifty = ohlc_Banknifty['NSE:{}'.format('NIFTY BANK')]['last_price']
                    ohlc_finnifty=kite.ohlc('NSE:{}'.format('NIFTY FIN SERVICE'))
                    ltp_finnifty = ohlc_finnifty['NSE:{}'.format('NIFTY FIN SERVICE')]['last_price']
                    ohlc_sensex=kite.ohlc('BSE:{}'.format('SENSEX'))
                    ltp_sensex = ohlc_sensex['BSE:{}'.format('SENSEX')]['last_price']
                    ohlc_bankex=kite.ohlc('BSE:{}'.format('BANKEX'))
                    ltp_bankex = ohlc_bankex['BSE:{}'.format('BANKEX')]['last_price']
                    ohlc_midcpnif=kite.ohlc('NSE:{}'.format('NIFTY MID SELECT'))
                    ltp_midcpnif = ohlc_midcpnif['NSE:{}'.format('NIFTY MID SELECT')]['last_price']

                    return ltp_nifty,ltp_Banknifty,ltp_finnifty,ltp_sensex,ltp_bankex,ltp_midcpnif
                except Exception as e:
                    print("exception occured get_N_BN_spot:" + str(e),i)
                    #ogfile.flush(); os.fsync(#ogfile.fileno())
                    time.sleep(2)
                    if i == 99 :
                        return ltp_nifty,ltp_Banknifty,ltp_finnifty,ltp_sensex,ltp_bankex,ltp_midcpnif

                    break
    def get_N_BN_spot_with_stocks():
        for i in range(0,100):
            while True:
                try:
                    ltp_nifty = 0;ltp_Banknifty = 0;
                    ohlc_nifty=kite.ohlc('NSE:{}'.format('NIFTY 50'))
                    ltp_nifty = ohlc_nifty['NSE:{}'.format('NIFTY 50')]['last_price']  
                    ohlc_Banknifty=kite.ohlc('NSE:{}'.format('NIFTY BANK'))
                    ltp_Banknifty = ohlc_Banknifty['NSE:{}'.format('NIFTY BANK')]['last_price']
                    ohlc_finnifty=kite.ohlc('NSE:{}'.format('NIFTY FIN SERVICE'))
                    ltp_finnifty = ohlc_finnifty['NSE:{}'.format('NIFTY FIN SERVICE')]['last_price']
                    ohlc_sensex=kite.ohlc('BSE:{}'.format('SENSEX'))
                    ltp_sensex = ohlc_sensex['BSE:{}'.format('SENSEX')]['last_price']
                    ohlc_bankex=kite.ohlc('BSE:{}'.format('BANKEX'))
                    ltp_bankex = ohlc_bankex['BSE:{}'.format('BANKEX')]['last_price']
                    ohlc_midcpnif=kite.ohlc('NSE:{}'.format('NIFTY MID SELECT'))
                    ltp_midcpnif = ohlc_midcpnif['NSE:{}'.format('NIFTY MID SELECT')]['last_price']
                    ohlc_hdfcbank=kite.ohlc('NSE:{}'.format('HDFCBANK'))
                    ltp_hdfcbank = ohlc_hdfcbank['NSE:{}'.format('HDFCBANK')]['last_price']
                    ohlc_icici=kite.ohlc('NSE:{}'.format('ICICIBANK'))
                    ltp_icici = ohlc_icici['NSE:{}'.format('ICICIBANK')]['last_price']
                    ohlc_sbin=kite.ohlc('NSE:{}'.format('SBIN'))
                    ltp_sbin = ohlc_sbin['NSE:{}'.format('SBIN')]['last_price']
                    ohlc_axis=kite.ohlc('NSE:{}'.format('AXISBANK'))
                    ltp_axis = ohlc_axis['NSE:{}'.format('AXISBANK')]['last_price']
                    ohlc_kotak=kite.ohlc('NSE:{}'.format('KOTAKBANK'))
                    ltp_kotak = ohlc_kotak['NSE:{}'.format('KOTAKBANK')]['last_price']
                    ohlc_indusind=kite.ohlc('NSE:{}'.format('INDUSINDBK'))
                    ltp_indusind = ohlc_indusind['NSE:{}'.format('INDUSINDBK')]['last_price']

                    return ltp_nifty,ltp_Banknifty,ltp_finnifty,ltp_sensex,ltp_bankex,ltp_midcpnif,ltp_hdfcbank,ltp_icici,ltp_sbin,ltp_axis,ltp_kotak,ltp_indusind
                except Exception as e:
                    print("exception occured get_N_BN_spot:" + str(e),i)
                    #ogfile.flush(); os.fsync(#ogfile.fileno())
                    time.sleep(2)
                    if i == 99 :
                        return ltp_nifty,ltp_Banknifty,ltp_finnifty,ltp_sensex,ltp_bankex,ltp_midcpnif

                    break


    def process_df_dump(df_tick):
        global instruments_list,df_tick_master,batcher
        for m in range(1,10):
            while True:
                try:
                    if True:
                                global df,rotation
                                df_tick = df_tick.dropna() #drop null calue rows
                                if len(df_tick.index) >0:
                                    df_tick.to_csv("first.csv")                
                                    ohlc_list = df_tick.ohlc.values.tolist()
                                    
                                    df_ohlc = pd.DataFrame.from_dict(ohlc_list);
                                    
                                    
                                    ohlc_close = df_ohlc.close.values;df_tick['close'] = ohlc_close;#del ohlc_close;#print(ohlc_high);
                                    
                                    
                                    df_tick['volume_traded'] = df_tick['volume_traded'].fillna(0)
                                    

                                    df_tick['Net Change'] = df_tick['last_price'] - df_tick['close']
                                    df_tick['Net Change'] = df_tick['Net Change'].apply(lambda x: round(x, 2))
                                    
                                    df_tick['change'] = df_tick['change'].apply(lambda x: round(x, 2))
                                    
                                    
                                    df_tick.rename(columns = {'instrument_token':'Stock','last_price':'LTP','change':'%Chg','volume_traded':'Vol','total_buy_quantity':'Totl Bid Qty','total_sell_quantity':'Totl Ask Qty'}, inplace = True);
                                    
                                    df = pd.concat([df, df_tick], ignore_index=True, axis=0)
                                    df.to_csv("second.csv")
                                    
                                    #print('after concat:',len(df.index))
                                    if True:
                                                df.exchange_timestamp = pd.to_datetime(df.exchange_timestamp, format="%Y-%m-%d %H:%M:%S")
                                                df.sort_values(by=['Stock', 'exchange_timestamp','Vol'], ascending=True, inplace = True)
                                                df.drop_duplicates(subset =['Stock'], keep = 'last', inplace = True)
                                                #df_Stock_list = df['Stock'].tolist();print('df_Stock_list len:',len(df_Stock_list))
                                    #print('after remove duplicate:',len(df.index))
                                    #df_z_Token_Symbol1 = df_z_Token_Symbol[df_z_Token_Symbol['Stock'].isin(df_Stock_list)]
                                    #print('df lenghts:',len(df.index),len(df_z_Token_Symbol.index))
                                    #df.to_csv('1.csv');df_z_Token_Symbol1.to_csv('2.csv');exit()
                                    #print('checkpoint 0')
                                    
                                    df2 = pd.merge(df,df_z_Token_Symbol,left_on='Stock',right_on='Stock'); #on='Stock',
                                    print("df2 merge over")
                                    #print('df2 lenghts:',len(df2.index))
                                    #   df2.to_csv('temp123.csv');exit()
                                    df2.rename(columns = {'Stock':'instrument_token'}, inplace = True);
                                    df2.rename(columns = {'tradingsymbol':'Stock'}, inplace = True);
                                    
                                    
                                    #   ltp_nifty,ltp_Banknifty,ltp_finnifty,ltp_sensex,ltp_bankex,ltp_midcpnif = get_N_BN_spot()
                                    ltp_nifty,ltp_Banknifty,ltp_finnifty,ltp_sensex,ltp_bankex,ltp_midcpnif,ltp_hdfcbank,ltp_icici,ltp_sbin,ltp_axis,ltp_kotak,ltp_indusind=get_N_BN_spot_with_stocks()
                                    
                                    
                                    
                                    df2 = df2[['Stock','name','LTP','Net Change','%Chg','Vol','oi','Totl Bid Qty','Totl Ask Qty','exchange_timestamp','strike','CE_PE','expiry']]
                                    
                                    
                                    #print('checkpoint 2.11')
                                    lister = []
                                    df2['greeks'] = df2.apply(lambda row: get_greeks(row['Stock'],row['LTP'],row['strike'],ltp_nifty,ltp_Banknifty,ltp_midcpnif,ltp_finnifty,ltp_sensex,ltp_bankex,ltp_hdfcbank,ltp_icici,ltp_sbin,ltp_axis,ltp_kotak,ltp_indusind,row['CE_PE'],row['expiry']),axis = 1)
                                    #   df_0915_day_Data['greeks'] = df_0915_day_Data.apply(lambda row: get_greeks(row['symbol'],row['close'],row['strike'],ltp_nifty,ltp_Banknifty,ltp_midcpnif,ltp_finnifty,ltp_sensex,ltp_Bankex,row['CE_PE'],row['expiry']),axis = 1)
                                    
                                    print("df2 greeeka over")
                                    
                                    #df2.to_csv('temp123.csv')
                                    greeks_list = df2.greeks.values.tolist()#;print('greeks_list:',greeks_list)
                                    Stock_list = df2.Stock.values.tolist()
                                    df91 = pd.DataFrame(greeks_list, columns = ['iv','delta','gamma','rho','theta','vega']) #, dtype = float
                                    df91['Stock'] = Stock_list
                                    df2 = pd.merge(df2,df91,on='Stock');
                                    
                                    #print(df2);exit()
                                    #df91 = pd.DataFrame(greeks_list)
                                    #print('checkpoint 2.12')
                                    #df2.to_csv('temp123.csv');exit()
                                    #print('df2',df2);exit()
                                    df2.rename(columns = {'exchange_timestamp':'Date'}, inplace = True);
                                    df2['strike'] = df2['strike'].astype(int)
                                    
                                    df2 = df2[['Date','Stock','name','strike','CE_PE','LTP','Net Change','%Chg','Vol','oi','Totl Bid Qty','Totl Ask Qty','iv','delta','gamma','rho','theta','vega','expiry']]
                                    df_vega = df2.groupby(['name','expiry','CE_PE'], as_index=False)['vega'].sum()
                                    df_iv = df2.groupby(['name','expiry','CE_PE'], as_index=False)['iv'].sum()
                                    
                                    df2 = pd.merge(df2,df_vega,on=['name','expiry','CE_PE'])
                                    
                                    df2.rename(columns = {'vega_x':'vega','vega_y':'sum_vega'}, inplace = True); #,'theta_x':'theta','theta_y':'sum_theta','delta_x':'delta','delta_y':'sum_delta'
                                    
                                    df2 = pd.merge(df2,df_iv,on=['name','expiry','CE_PE']);
                                    
                                    df2.rename(columns = {'iv_x':'iv','iv_y':'sum_iv'}, inplace = True);
                                    
                                    df2['Batch'] = batcher
                                    if not len(df_tick_master.index) < 100000:
                                        df_tick_master = pd.DataFrame()
                                        rotation+=1
                                
                                    df_tick_master = pd.concat([df_tick_master, df2], ignore_index=True, axis=0)
                                    
                                    
                                    #time.sleep(0.1)
                                    counter = 0
                                    batcher+=1
                                    return df_tick_master
                                else:
                                    return pd.DataFrame()
                except Exception as e:
                    print("exception occured process_df_dump:" + str(e))
                    time.sleep(1);
                    break



    ## EXPIRY Identification logic.
    def Expiry_selection(Expiry_to_use,Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M, Index_to_TAKE_TRADE,curr_next):
    
        list2 = [];list11 = []

 
        curr_dt = time.strftime("%Y-%m-%d", time.localtime())
        df = pd.read_csv('instruments.csv',dtype={"segment":"category","name":"category","tradingsymbol":"str","exchange_token":"int32","instrument_token":"int32","lot_size":"int8","exchange":"category","instrument_type":"category","expiry":"category"})
        dfw = df[['name','expiry','segment']]
        #dfw['expiry'] = pd.to_datetime(dfw['expiry'])
        dfw = dfw[(dfw.name == Index_to_TAKE_TRADE)]
        dfw = dfw[(dfw.segment == 'NFO-OPT') | (dfw.segment == 'BFO-OPT')]
        list11=dfw['expiry'].unique()

        
        for date1 in list11:
            if Expiry_Day_use_Expiry_to_USE_W == 'NO':
                if (date1 >= curr_dt):
                    list2.append(date1)
            elif Expiry_Day_use_Expiry_to_USE_W == 'YES':
                if (date1 > curr_dt):
                    list2.append(date1)
        if (len(list2) == 0):
            list2.append('')
        list2.sort()
        Weekly_expiry = list2[0]
        Next_Weekly_Expiry = list2[1]
        
        ##Monthly Expiry Selection
        list2 = []
        #curr_dt = time.strftime("%Y-%m-%d", time.localtime())
        dfw = df[['name','expiry','segment']]
        dfw = dfw[(dfw.name == Index_to_TAKE_TRADE)]
        dfw = dfw[(dfw.segment == 'NFO-FUT') | (dfw.segment == 'BFO-FUT')]
        list1=dfw['expiry'].unique()
        print(f"monthly expiry list of {Index_to_TAKE_TRADE} is {list1}")
        for date1 in list1:
            if Expiry_Day_use_Expiry_to_USE_M == 'NO':
                if (date1 >= curr_dt) and date1>=Next_Weekly_Expiry:
                    list2.append(date1)
            elif Expiry_Day_use_Expiry_to_USE_M == 'YES':
                if (date1 > curr_dt) and date1>=Next_Weekly_Expiry:
                    list2.append(date1)
        if (len(list2) == 0):
            list2.append('')
        
        #list2 = list(set(list2).difference(curr_next))
        list2.sort()
        MonthlyExpiry = list2[0]
        Next_Month_Expiry = list2[1];
        
        #print('Identified Monthly Expiry :',MonthlyExpiry)
        #print('Identified Weekly Expiry :',Weekly_expiry)
        #monthly_exp = MonthlyExpiry
        
        if Expiry_to_use == 'WEEKLY' :
            Expiry_dt = Weekly_expiry
        elif Expiry_to_use == 'NEXT_WEEK' :
            Expiry_dt = Next_Weekly_Expiry
        elif Expiry_to_use == 'MONTHLY' :
            Expiry_dt = MonthlyExpiry
        elif Expiry_to_use == 'NEXT_MONTH' :
            Expiry_dt = Next_Month_Expiry
        elif Expiry_to_use == 'MANUAL' :
            Expiry_dt = Weekly_expiry
            if Manual_expiry in list11:
                print('The Manually orivided Expiry date Validation is passed!!. ALGO is continuing')
                #ogfile.flush(); os.fsync(#ogfile.fileno())
                Expiry_dt = Manual_expiry
            else:
                print('The Manually provided Expiry date',Manual_expiry,'is not correct, Please check it, ALGO is exiting!!')
                #ogfile.flush(); os.fsync(#ogfile.fileno())
                exit()
        else:
            print('Expiry is not valid',Expiry_to_use);exit()
        #print('ALGO is Going to use this Expiry day:',Expiry_dt,'Expiry_to_use:',Expiry_to_use);
        MonthlyExpiry = list1[0]
        return Expiry_dt,MonthlyExpiry


    def get_column_letter(col_idx_1_based):
        
        col_idx = col_idx_1_based - 1  
        if col_idx < 26:
            return chr(col_idx + 65)  
        else:
        
            first_letter = chr((col_idx // 26) + 64)  
            second_letter = chr((col_idx % 26) + 65)
            return first_letter + second_letter

    class TimeChecker:
        def __init__(self):
            self.first_time = True
            self.new_tm = None

        def update_time(self, tm):
            if self.new_tm is None or self.new_tm != tm:
                self.new_tm = tm
                self.first_time = False
                print(f"Time updated to {tm}, first_time is set to True.")
            else:
                if self.first_time:
                    self.first_time = True
                    print(f"Time remains {tm}, first_time now set to False.")
                else:
                    print(f"Time still {tm}, first_time remains False.")

        def reset(self):
            self.first_time = True
            self.new_tm = None
            print("TimeChecker has been reset.")



    time_checker = TimeChecker()
    def modify_df(df):
    
        columns_with_min = [col for col in df.columns if col.endswith('_Min')]
        
        for col in columns_with_min:
            
            for i in range(1, len(df[col]), 2):
                df.at[i, col] = 'sentiment'
        
        return df

    def processing(Ranked_dff_csv_past0,tm):
            tm = str(tm)[11:][:5].replace(':','')
            Greek_nm = 'Greek_' + tm
            df_combined = pd.DataFrame()
            Ranked_dff_csv_past0 = Ranked_dff_csv_past0[['Date','Stock','name','strike','expiry','CE_PE','LTP','Net Change','%Chg','Vol','iv','delta','gamma','rho','theta','vega']]
            df_vega = Ranked_dff_csv_past0.groupby(['name','CE_PE','expiry'], as_index=False)['vega'].sum()
            df_vega['Greek_type'] = 'vega'
            df_vega.rename(columns = {'vega':Greek_nm}, inplace = True);
            #print(df_vega.columns.values)
            df_theta = Ranked_dff_csv_past0.groupby(['name','CE_PE','expiry'], as_index=False)['theta'].sum()
            df_theta['Greek_type'] = 'theta'
            df_theta.rename(columns = {'theta':Greek_nm}, inplace = True);
            #print(df_theta.columns.values)
            df_delta = Ranked_dff_csv_past0.groupby(['name','CE_PE','expiry'], as_index=False)['delta'].sum()
            df_delta['Greek_type'] = 'delta'
            df_delta.rename(columns = {'delta':Greek_nm}, inplace = True);
            #print(df_delta.columns.values)
            df_iv = Ranked_dff_csv_past0.groupby(['name','CE_PE','expiry'], as_index=False)['iv'].sum()
            df_iv['Greek_type'] = 'iv'
            df_iv.rename(columns = {'iv':Greek_nm}, inplace = True);
            #print('df_vega:');print(df_vega);print(df_theta);print(df_delta)
            #df_combined1 = pd.merge(df_vega,df_theta,on=['name','CE_PE','expiry'])
            #df_combined = pd.merge(df_combined1,df_delta,on=['name','CE_PE','expiry'])
            df_combined = pd.concat([df_combined, df_vega], ignore_index=True, axis=0)#;print('After Vega:',len(df_combined.index))
            df_combined = pd.concat([df_combined, df_theta], ignore_index=True, axis=0)#;print('After Theta:',len(df_combined.index))
            df_combined = pd.concat([df_combined, df_delta], ignore_index=True, axis=0)#;print('After Delta:',len(df_combined.index))
            df_combined = pd.concat([df_combined, df_iv], ignore_index=True, axis=0)#;print('After iv:',len(df_combined.index))
            #df_combined = df_combined.round({"iv":4, "delta":4, "gamma":10, "rho":10, "theta":10, "vega":10})
            df_combined = df_combined[['name','expiry','CE_PE','Greek_type',Greek_nm]]
            #print('df_combined:');print(df_combined)
            return df_combined

    def get_required_vega_data_df():
        while True:
            #df_v = pd.read_csv('All_Data_Master_20230324_090926_new.csv', index_col=0)
            globbed_files = glob.glob("D:\pythonProject\pythonProject_HP\Strategies_Source Code\Pradeep_Jainam code for analysis\All_Data_Master_20230403_*.csv")
            df_v = pd.DataFrame()
            ii=0
            for csv in globbed_files:
                if True:
                        ii+=1
                        df_file = pd.read_csv(csv) #, header=None#, sep='|' add it if separater is differrnt
                        #print('Extracted:',csv, ii,'DF len:',len(df_file.index))
                        df_v = pd.concat([df_v,df_file],ignore_index=True);
            #df_v['expiry'] = '2023-03-29'
            #df_v.to_csv('All_Data_Master_20230324_090926_new.csv',index=False);exit()
            #print(df_v)
            if len(df_v.index) >0:
                return df_v
    def determine_theta_sentiment(value):
        if -5 < value < 5:
            return "Sideways"
        elif 5 <= value < 10:
            return "Moderate Bullish"
        elif value >= 10:
            return "Bullish"
        elif -10 < value <= -5:
            return "Moderate Bearish"
        else:  # value <= -10
            return "Bearish"
    def get_historic_data(name, timeframe, delta,Instrument_type):
        print('Hist Extracting for:',name)
        for i in range(0,1000):
            while True:
                    try:
                            to_date = datetime.datetime.now().date()
                            from_date = to_date - datetime.timedelta(days=int(delta))
                            token = kite.ltp([Instrument_type+':' + name])[Instrument_type+':' + name]['instrument_token']
                            data = kite.historical_data(token, from_date, to_date, timeframe,oi=True)
                            data = pd.DataFrame(data)
                            if len(data.index) >=0:
                                data['symbol'] = name
                                return data
                            else:
                                return pd.DataFrame()
                    except Exception as e:
                            print("exception occured get_historic_data:" + str(e),i);
                            #ogfile.flush(); os.fsync(#ogfile.fileno())
                            time.sleep(5);err_msg = 'single positional indexer is out-of-bounds'
                            if (err_msg in str(e) and i == 2):
                                data = pd.DataFrame()
                                return data
                            elif ( i == 3):
                                data = pd.DataFrame()
                                return data
            #pass
                    break

    def get_historic_data(name, timeframe, delta,Instrument_type):
        print('Hist Extracting for:',name)
        for i in range(0,1000):
            while True:
                    try:
                            to_date = datetime.datetime.now().date()
                            from_date = to_date - datetime.timedelta(days=int(delta))
                            token = kite.ltp([Instrument_type+':' + name])[Instrument_type+':' + name]['instrument_token']
                            data = kite.historical_data(token, from_date, to_date, timeframe,oi=True)
                            data = pd.DataFrame(data)
                            if len(data.index) >=0:
                                data['symbol'] = name
                                return data
                            else:
                                return pd.DataFrame()
                    except Exception as e:
                            print("exception occured get_historic_data:" + str(e),i);
                            #ogfile.flush(); os.fsync(#ogfile.fileno())
                            time.sleep(5);err_msg = 'single positional indexer is out-of-bounds'
                            if (err_msg in str(e) and i == 2):
                                data = pd.DataFrame()
                                return data
                            elif ( i == 3):
                                data = pd.DataFrame()
                                return data
            #pass
                    break

    def manage_columns(df):

        max_columns = 9
        if len(df.columns) > max_columns:        
            greek_0915_index = df.columns.get_loc("Greek_0915")
            column_to_remove = df.columns[greek_0915_index + 1]
            df.drop(columns=[column_to_remove], inplace=True)
            
        return df

    def process_df_dump_prev_Day_Data(df_tick):
        global instruments_list,df_tick_master,batcher
        for m in range(1,10):
            while True:
                try:
                    if True:
                                global df,rotation
                                
                                

                                df_tick = df_tick.dropna() #drop null calue rows
                                
                                if len(df_tick.index) >0:
                                    #print(df_tick)
                                    #df_tick.to_csv('temp.csv');exit()
                                    ohlc_list = df_tick.ohlc.values.tolist()
                                    df_ohlc = pd.DataFrame.from_dict(ohlc_list);
                                    ohlc_close = df_ohlc.close.values;df_tick['close'] = ohlc_close;#del ohlc_close;#print(ohlc_high);
                                    df_tick['volume_traded'] = df_tick['volume_traded'].fillna(0)
                                    df_tick['Net Change'] = df_tick['last_price'] - df_tick['close']
                                    df_tick['Net Change'] = df_tick['Net Change'].apply(lambda x: round(x, 2))
                                    df_tick['change'] = df_tick['change'].apply(lambda x: round(x, 2))
                                    df_tick.rename(columns = {'instrument_token':'Stock','last_price':'LTP','change':'%Chg','volume_traded':'Vol','total_buy_quantity':'Totl Bid Qty','total_sell_quantity':'Totl Ask Qty'}, inplace = True);
                                    #print('before concat:',len(df.index))
                                    df = pd.concat([df, df_tick], ignore_index=True, axis=0)
                                
                                #print('after concat:',len(df.index))
                                    if True:
                                                df.exchange_timestamp = pd.to_datetime(df.exchange_timestamp, format="%Y-%m-%d %H:%M:%S")
                                                df.sort_values(by=['Stock', 'exchange_timestamp','Vol'], ascending=True, inplace = True)
                                                df.drop_duplicates(subset =['Stock'], keep = 'last', inplace = True)
                                                #df_Stock_list = df['Stock'].tolist();print('df_Stock_list len:',len(df_Stock_list))
                                    
                                    df2 = pd.merge(df,df_z_Token_Symbol,left_on='Stock',right_on='Stock'); #on='Stock',
                                    
                                    df2.rename(columns = {'Stock':'instrument_token'}, inplace = True);
                                    df2.rename(columns = {'tradingsymbol':'Stock'}, inplace = True);
                                    
                                    ltp_nifty,ltp_Banknifty,ltp_finnifty,ltp_sensex,ltp_bankex,ltp_midcpnif = get_N_BN_spot()
                                    
                                    
                                    df2 = df2[['Stock','name','LTP','Net Change','%Chg','Vol','Totl Bid Qty','Totl Ask Qty','exchange_timestamp','strike','CE_PE','expiry']]
                                    #print('checkpoint 2.11')
                                    
                                    lister = []
                                    df2['greeks'] = df2.apply(lambda row: get_greeks(row['Stock'],row['LTP'],row['strike'],ltp_nifty,ltp_Banknifty,row['CE_PE'],row['expiry']),axis = 1)
                                    #df2.to_csv('temp123.csv')
                                    greeks_list = df2.greeks.values.tolist()#;print('greeks_list:',greeks_list)
                                    Stock_list = df2.Stock.values.tolist()
                                    df91 = pd.DataFrame(greeks_list, columns = ['iv','delta','gamma','rho','theta','vega']) #, dtype = float
                                    df91['Stock'] = Stock_list
                                    df2 = pd.merge(df2,df91,on='Stock');
                                    
                                    #print(df2);exit()
                                    #df91 = pd.DataFrame(greeks_list)
                                    #print('checkpoint 2.12')
                                    #df2.to_csv('temp123.csv');exit()
                                    #print('df2',df2);exit()
                                    df2.rename(columns = {'exchange_timestamp':'Date'}, inplace = True);
                                    df2['strike'] = df2['strike'].astype(int)
                                    df2 = df2[['Date','Stock','name','strike','CE_PE','LTP','Net Change','%Chg','Vol','iv','delta','gamma','rho','theta','vega','expiry']]
                                    df_vega = df2.groupby(['name','expiry','CE_PE'], as_index=False)['vega'].sum()
                                    df_iv = df2.groupby(['name','expiry','CE_PE'], as_index=False)['iv'].sum()
                                    #df_theta = df2.groupby(['name','CE_PE'], as_index=False)['theta'].sum()
                                    #df_delta = df2.groupby(['name','CE_PE'], as_index=False)['delta'].sum()
                                    #print('df_vega:',df_vega);exit()
                                    #df_vega = df2.groupby(['name','CE_PE'], as_index=False)['vega'].sum()
                                    #print('df_vega:',df_vega)
                                    df2 = pd.merge(df2,df_vega,on=['name','expiry','CE_PE']);
                                    #df2 = pd.merge(df2,df_theta,on=['name','CE_PE']);
                                    #df2 = pd.merge(df2,df_delta,on=['name','CE_PE']);
                                    df2.rename(columns = {'vega_x':'vega','vega_y':'sum_vega'}, inplace = True); #,'theta_x':'theta','theta_y':'sum_theta','delta_x':'delta','delta_y':'sum_delta'
                                    df2 = pd.merge(df2,df_iv,on=['name','expiry','CE_PE']);
                                    df2.rename(columns = {'iv_x':'iv','iv_y':'sum_iv'}, inplace = True);
                                    
                                    #df2['LTP'] = df2['LTP'].apply(lambda x: round(x, 2))
                                    #df2['Net Change'] = df2['Net Change'].apply(lambda x: round(x, 2))
                                    #df2['%Chg'] = df2['%Chg'].apply(lambda x: round(x, 2))
                                    #df2.to_csv('All_Data_test1.csv', index=False)
                                    #print('df2 Final:',df2);df2.to_csv('temp1234.csv');exit()
                                    #print('checkpoint 3')
                                    df2['Batch'] = batcher
                                    if not len(df_tick_master.index) < 100000:
                                        df_tick_master = pd.DataFrame()
                                        rotation+=1
                                    df_tick_master = pd.concat([df_tick_master, df2], ignore_index=True, axis=0)
                                    
                                    #df_tick_master.to_csv(master_file_nm+str(rotation)+'.csv', index=False);
                                    #time.sleep(0.1)
                                    counter = 0
                                    batcher+=1
                                    return df_tick_master
                                else:
                                    return pd.DataFrame()
                except Exception as e:
                    print("exception occured process_df_dump:" + str(e))
                    #ogfile.flush(); os.fsync(#ogfile.fileno())
                    time.sleep(1);
                break


    def wait_for_next_minute(min_start):
        current_time = datetime.datetime.now()
        print("Current time is:", current_time)
        #ogfile.flush(); os.fsync(#ogfile.fileno())
        next_minute_15_sec = min_start + datetime.timedelta(minutes=1, seconds=15)
        print("Next minute + 15 seconds is:", next_minute_15_sec) 
        #ogfile.flush(); os.fsync(#ogfile.fileno())   
        if current_time < next_minute_15_sec:
            time_to_next_minute_15_sec = (next_minute_15_sec - current_time).total_seconds()
            print(f"Sleeping for {time_to_next_minute_15_sec} seconds...")
            #ogfile.flush(); os.fsync(#ogfile.fileno())
            time.sleep(time_to_next_minute_15_sec)
        else:
            print("Next minute + 15 seconds already reached or passed.")
            #ogfile.flush(); os.fsync(#ogfile.fileno())

    import os
    import pandas as pd
    from openpyxl.utils import get_column_letter, column_index_from_string
    import xlsxwriter
    import os
    import pandas as pd
    from openpyxl.utils import get_column_letter, column_index_from_string
    import os
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl import Workbook

    import os
    import pandas as pd
    from openpyxl.utils import get_column_letter, column_index_from_string

    def filter_and_export(starting_cells, f_time, df, index_name, mode, sheet_name, expiry_today):
        print("Inside filter and export")
        print("index_name is", index_name)
        print("expiry_today is", expiry_today)
        print("f time is",f_time)

        df.to_csv("incomingdf.csv")

        if index_name not in expiry_today:
            print(f"Index name {index_name} not in expiry_today")
            return pd.DataFrame(), starting_cells, f_time

        print("index_name is", index_name)

        file_name_map = {
            'NIFTY': 'NF_theta_delta.xlsx',
            'BANKNIFTY': 'BNF_theta_delta.xlsx',
            'FINNIFTY': 'FIN_theta_delta.xlsx',
            'MIDCPNIFTY': 'Midcp_theta_delta.xlsx',
            'SENSEX': 'Sen_theta_delta.xlsx',
            'BANKEX': 'BAN_theta_delta.xlsx'
        }
        file_name = file_name_map[index_name]
        print("file_name:", file_name)

        full_path = os.path.join(os.getcwd(), file_name)
        print("full_path:", full_path)

        if f_time:
            # Read existing data if any
            if os.path.exists(full_path):
                existing_df = pd.read_excel(full_path, sheet_name=sheet_name, engine='openpyxl')
                print(f"Sheet data read successfully: {sheet_name}")
            else:
                existing_df = pd.DataFrame()
                print(f"No existing file found. Creating new file: {full_path}")
            
            # Concatenate existing data with new data if it exists
            if not existing_df.empty:
                combined_df = pd.concat([existing_df, df], ignore_index=True)
            else:
                combined_df = df
            
            # Write the updated DataFrame back to the Excel file starting from A1
            combined_df.to_csv("incoming2.csv")
            with pd.ExcelWriter(full_path, engine='xlsxwriter') as writer:
                combined_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
            
            print(f"Data written and workbook saved: {full_path}")
            return combined_df, starting_cells, f_time
        else:
            # Write the new DataFrame back to the Excel file starting from A1
            df.to_csv("incoming3.csv")
            with pd.ExcelWriter(full_path, engine='xlsxwriter') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
            
            print(f"Data written and workbook saved: {full_path}")
            return df, starting_cells, f_time

    



    def filter_and_export_to_excel(starting_cell,f_time,df, index_name, mode,sheet_name,expiry_today):
     if index_name in expiry_today:
        if not df.empty:

            file_name_map = {
                'NIFTY': 'NF_Pro_Trading.xlsx',
                'BANKNIFTY': 'BNF_Pro_Trading.xlsx',
                'FINNIFTY': 'FIN_Pro_Trading.xlsx',
                'MIDCPNIFTY': 'Midcp_Pro_Trading.xlsx',
                'SENSEX':'Sen_Pro_Trading.xlsx',
                'BANKEX':'BAN_Pro_Trading.xlsx'
            }
            file_name = file_name_map[index_name]
            full_path = os.path.join(os.getcwd(), file_name)

            if xw.apps.count == 0:
                app = xw.App(visible=True)
                app.books.add()
                if len(app.books) > 1:
                    for book in list(app.books):
                        if book.name.startswith('Book') and 'Sheet1' in [s.name for s in book.sheets]:
                            book.close()
            else:
                app = xw.apps.active

            try:
                wb = app.books[file_name]
            except:
                try:
                    wb = app.books.open(full_path)
                except FileNotFoundError:
                    wb = app.books.add()
                    wb.save(full_path)

            
            
        

            sheet = wb.sheets[sheet_name]
            
            # df = df.loc[:, ~df.columns.str.startswith('Unnamed')]  
            try:

                sheet.range(starting_cell).options(index=False, header=False).value = df 
            except Exception as e:
                print("error is",e)       

            wb.save()
            time.sleep(0.1)
            




    def pro_export_minute(df, index_name, mode, sheet_name,expiry_today):

     if index_name in expiry_today:
    
        file_name_map = {
            'NIFTY': 'NF_Pro_Trading.xlsx',
            'BANKNIFTY': 'BNF_Pro_Trading.xlsx',
            'FINNIFTY': 'FIN_Pro_Trading.xlsx',
            'MIDCPNIFTY': 'Midcp_Pro_Trading.xlsx',
            'SENSEX': 'Sen_Pro_Trading.xlsx',
            'BANKEX': 'BAN_Pro_Trading.xlsx'

        }
        file_name = file_name_map[index_name]
        full_path = os.path.join(os.getcwd(), file_name)
        if xw.apps.count == 0:
            app = xw.App(visible=True)
        else:
            app = xw.apps.active

        try:
            wb = app.books[file_name]
        except:
            try:
                wb = app.books.open(full_path)
            except FileNotFoundError:
                wb = app.books.add()
                wb.save(full_path)

        sheet = wb.sheets[sheet_name]
        starting_cell = starting_cells.get(mode)

        df_filtered = df[df['name'] == index_name]
        df_filtered['expiry'] = pd.to_datetime(df_filtered['expiry'], format='%Y-%m-%d')

        if mode == "current":
            expiry_date = df_filtered['expiry'].min()
        elif mode == "next":
            next_expiry = df_filtered[df_filtered['expiry'] > df_filtered['expiry'].min()]['expiry'].min()
            expiry_date = next_expiry

        df_filtered = df_filtered[df_filtered['expiry'] == expiry_date]

        if not df_filtered.empty:

            data = df_filtered.iloc[0]
            sheet.range('C2').value = data['protrader_senti']
            sheet.range('E2').value = data['protrader_data']
            sheet.range('C3').value = data['retailtrader_senti']
            sheet.range('E3').value = data['retailtrader_data']

        wb.save(full_path)
        
    
    def pro_export(starting_cells, f_time, df, index_name, mode, sheet_name, expiry_today):
        if index_name in expiry_today:
            file_name_map = {
                'NIFTY': 'NF_Fii_Pro.xlsx',
                'BANKNIFTY': 'BNF_Fii_Pro.xlsx',
                'FINNIFTY': 'FIN_Fii_Pro.xlsx',
                'MIDCPNIFTY': 'Mid_Fii_Pro.xlsx',
                'SENSEX': 'Sen_Fii_Pro.xlsx',
                'BANKEX': 'BAN_Fii_Pro.xlsx',
            }
            file_name = file_name_map[index_name]
            full_path = os.path.join(os.getcwd(), file_name)

            if f_time:
                # Read existing data if any
                if os.path.exists(full_path):
                    # existing_df = pd.read_excel(full_path, sheet_name=sheet_name, engine='openpyxl')
                    existing_df  = pd.read_excel(full_path, sheet_name=sheet_name, engine='openpyxl')
                    print(f"Sheet data read successfully: {sheet_name}")
                else:
                    existing_df = pd.DataFrame()
                    print(f"No existing file found. Creating new file: {full_path}")
                
                # Concatenate existing data with new data if it exists
                if not existing_df.empty:
                    combined_df = pd.concat([existing_df, df], ignore_index=True)
                else:
                    combined_df = df
                
                # Write the updated DataFrame back to the Excel file starting from A1
                with pd.ExcelWriter(full_path, engine='xlsxwriter') as writer:
                    combined_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
                
                print(f"Data written and workbook saved: {full_path}")
            
                return combined_df,starting_cells, f_time
            else:
                # Write the new DataFrame back to the Excel file starting from A1
                with pd.ExcelWriter(full_path, engine='xlsxwriter') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
                
                print(f"Data written and workbook saved: {full_path}")
                return df,starting_cells, f_time
        else:
            return df,starting_cells, f_time


    def spot_update(df, index_name):
        file_name_map = {
            'NIFTY': 'NF_Pro_Trading.xlsx',
            'BANKNIFTY': 'BNF_Pro_Trading.xlsx',
            'FINNIFTY': 'FIN_Pro_Trading.xlsx',
            'MIDCPNIFTY': 'Midcp_Pro_Trading.xlsx',
            'SENSEX':'Sen_Pro_Trading.xlsx'
        }
        file_name = file_name_map['BANKNIFTY']
        full_path = os.path.join(os.getcwd(), file_name)

        if xw.apps.count == 0:
            app = xw.App(visible=True)
            app.books.add()
            if len(app.books) > 1:
                for book in list(app.books):
                    if book.name.startswith('Book') and 'Sheet1' in [s.name for s in book.sheets]:
                        book.close()
        else:
            app = xw.apps.active

        try:
            wb = app.books[file_name]
        except:
            try:
                wb = app.books.open(full_path)
            except FileNotFoundError:
                wb = app.books.add()
                wb.save(full_path)

        
        df_filtered = df[df['Name'] == index_name]

        column_to_start_from = df.columns.get_loc('Open')  # Find the index of 'Time' column
        df_to_export = df_filtered.iloc[:, column_to_start_from:]  # Select from 'Time' column onwards
        sheet = wb.sheets['BNF_ProTrading_CW']
        
        if index_name=='NIFTY BANK':
            sheet.range('B3').options(index=False, header=False).value = df_to_export
        elif index_name=='NIFTY 50':
            sheet.range('F3').options(index=False, header=False).value = df_to_export
        elif index_name=='NIFTY FIN SERVICE':
            sheet.range('J3').options(index=False, header=False).value = df_to_export
        elif index_name=='NIFTY MID SELECT':
            sheet.range('N3').options(index=False, header=False).value = df_to_export
        elif index_name=='SENSEX':
            sheet.range('R3').options(index=False, header=False).value = df_to_export
        elif index_name=='INDIA VIX':
            sheet.range('V3').options(index=False, header=False).value = df_to_export
        
        
        wb.save(full_path)

    

    def update_excel_sheet(df, sheet_name):
        
        file_name = 'Zerodha Live Quotes NFO Greeks_Project.xlsx'  
        full_path = os.path.join(os.getcwd(), file_name) 

        if xw.apps.count == 0:
            app = xw.App(visible=True)
        else:
            app = xw.apps.active

        try:
            wb = app.books[file_name]
        except:
            try:
                wb = app.books.open(full_path)
            except FileNotFoundError:
                wb = app.books.add()
                wb.save(full_path)


        if sheet_name not in [sheet.name for sheet in wb.sheets]:
            wb.sheets.add(sheet_name)
        sheet = wb.sheets[sheet_name]

        df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
        sheet.range('A1').options(index=False).value = df

        wb.save()

    def compare_options(current_df, previous_df, name="BANKNIFTY", ce_pe="CE"):
        # Filter both DataFrames for the specific index and Call/Put type
        current_filtered = current_df[(current_df['name'] == name) & (current_df['CE_PE'] == ce_pe)]
        previous_filtered = previous_df[(previous_df['name'] == name) & (previous_df['CE_PE'] == ce_pe)]

        # Join the DataFrames on 'Stock' to compare the same options
        combined_df = pd.merge(current_filtered, previous_filtered, on="Stock", suffixes=('_curr', '_prev'))

        # Calculate the changes for vega, theta, and delta
        combined_df['vega_change'] = combined_df['vega_curr'] - combined_df['vega_prev']
        combined_df['theta_change'] = combined_df['theta_curr'] - combined_df['theta_prev']
        combined_df['delta_change'] = combined_df['delta_curr'] - combined_df['delta_prev']

        # Filter based on the criteria provided:
        # - Vega increase of more than 5%
        # - Theta decrease of more than 10 points
        # - Delta increase of more than 10 points
        filtered_df = combined_df[
            (combined_df['vega_change'] / combined_df['vega_prev'] > 0.05) |
            (combined_df['theta_change'] < -10) |
            (combined_df['delta_change'] > 10)
        ]

        # Selecting required columns for the new DataFrame
        final_columns = ['Stock', 'name_curr', 'strike_curr', 'CE_PE_curr', 'LTP_curr', 'delta_curr', 'theta_curr', 'vega_curr', 'vega_change', 'theta_change', 'delta_change']

        result_df = filtered_df[final_columns].rename(columns={
            'name_curr': 'name',
            'strike_curr': 'strike',
            'CE_PE_curr': 'CE_PE',
            'LTP_curr': 'LTP',
            'delta_curr': 'delta',
            'theta_curr': 'theta',
            'vega_curr': 'vega'
        })

        return result_df



    def process_combined_dfs(current_df, previous_df):
        indices = ["BANKNIFTY", "NIFTY", "FINNIFTY", "MIDCPNIFTY", "SENSEX","BANKEX"]
        option_types = ["CE", "PE"]
        results = {}

        for index in indices:
            # Determine unique expiries and sort them to identify current and next expiries
            expiries = sorted(current_df[current_df['name'] == index]['expiry'].unique())
            
            for i, expiry in enumerate(expiries):
                for condition_suffix in ['increase', 'decrease']:
                    combined_results = []  # This will hold the combined results for CE and PE

                    for option_type in option_types:
                        # Filter for each index, option type, and expiry
                        current_filtered = current_df[(current_df['name'] == index) & (current_df['CE_PE'] == option_type) & (current_df['expiry'] == expiry)]
                        previous_filtered = previous_df[(previous_df['name'] == index) & (previous_df['CE_PE'] == option_type) & (previous_df['expiry'] == expiry)]

                        # Merge and calculate changes
                        if not current_filtered.empty and not previous_filtered.empty:
                            combined_df = pd.merge(current_filtered, previous_filtered, on="Stock", suffixes=('_curr', '_prev'))
                            combined_df['vega_change'] = combined_df['vega_curr'] - combined_df['vega_prev']
                            combined_df['theta_change'] = combined_df['theta_curr'] - combined_df['theta_prev']
                            combined_df['delta_change'] = combined_df['delta_curr'] - combined_df['delta_prev']

                            # Filter based on condition
                            if condition_suffix == 'increase':
                                condition = (combined_df['vega_change'] / combined_df['vega_prev'] > 0.05) | \
                                            (combined_df['theta_change'] < -10) | \
                                            (combined_df['delta_change'] > 10)
                            else:  # decrease condition
                                condition = (combined_df['vega_change'] / combined_df['vega_prev'] <= 0.05) | \
                                            (combined_df['theta_change'] >= 10) | \
                                            (combined_df['delta_change'] <= -10)

                            filtered_df = combined_df[condition]
                            final_columns = ['Stock', 'name_curr', 'strike_curr', 'CE_PE_curr', 'LTP_curr',
                                            'delta_curr', 'theta_curr', 'vega_curr', 'vega_change',
                                            'theta_change', 'delta_change']
                            result_df = filtered_df[final_columns].rename(columns={
                                'name_curr': 'name',
                                'strike_curr': 'strike',
                                'CE_PE_curr': 'CE_PE',
                                'LTP_curr': 'LTP',
                                'delta_curr': 'delta',
                                'theta_curr': 'theta',
                                'vega_curr': 'vega'
                            })
                            combined_results.append(result_df)

                    # Combine CE and PE results
                    combined_df = pd.concat(combined_results, ignore_index=True)
                    key = f"{index.lower()}_{('current' if i == 0 else 'next')}_expiry_{condition_suffix}"
                    results[key] = combined_df

                    # Break the loop for indices with only one expiry
                    if index in ["FINNIFTY", "MIDCPNIFTY", "SENSEX","BANKEX"]:
                        break

        return results




    def determine_delta_sentiment_pe(value):
            if -5 < value < 5:
                return "Sideways"
            elif 5 <= value < 10:
                return "Moderate Bearish"
            elif value >= 10:
                return "Bearish"
            elif -10 < value <= -5:
                return "Moderate Bullish"
            else:  # value <= -10
                return "Bullish"






    def determine_delta_sentiment_ce(value):
    
            if -5 < value < 5:
                return "Sideways"
            elif -10 <= value < -5:
                return "Moderate Bearish"
            elif value < -10:
                return "Bearish"
            elif 5 <= value < 10:
                return "Moderate Bullish"
            else:  # value >= 10
                return "Bullish"


    def determine_sentiment(value):
        if -5 < value < 5:
            return "Sideways"
        elif 5 <= value < 10:
            return "Moderate Bullish"
        elif value >= 10:
            return "Bullish"
        elif -10 < value <= -5:
            return "Moderate Bearish"
        else:  # value <= -10
            return "Bearish"

    def determine_vega_sentiment(value):
        if -5 < value < 5:
            return "Sideways"
        elif 5 <= value < 10:
            return "Moderate Bullish"
        elif value >= 10:
            return "Bullish"
        elif -10 < value <= -5:
            return "Moderate Bearish"
        else:  # value <= -10
            return "Bearish"
                                

    def create_index_1min(index, call_put):
        if index == 'theta':
            return 'Theta_' + call_put
        elif index == 'vega':
            return 'Vega_' + call_put
        elif index == 'delta':
            return 'Delta_' + call_put


    def merge_with_delta(newhope_df, probably_latest_delta_df,tf):

        newhope_df['expiry'] = pd.to_datetime(newhope_df['expiry'], format='%d-%m-%Y').dt.strftime('%d-%m-%Y')
        probably_latest_delta_df['expiry'] = pd.to_datetime(probably_latest_delta_df['expiry'], format='%d-%m-%Y').dt.strftime('%d-%m-%Y')
        
        # Initialize new columns for Delta_CE and Delta_PE in newhope_df
        newhope_df['Delta_CE'] = None
        newhope_df['Delta_PE'] = None

        # Iterate through newhope_df to populate Delta_CE and Delta_PE
        for index, row in newhope_df.iterrows():
            # Filter probably_latest_delta_df for matching Index, expiry, and 1_Min/time
            matching_rows = probably_latest_delta_df[
                (probably_latest_delta_df['Index'] == row['Index']) & 
                (probably_latest_delta_df['expiry'] == row['expiry']) &
                (probably_latest_delta_df['time'] == row[tf])
            ]

            # Populate Delta_CE and Delta_PE
            for _, delta_row in matching_rows.iterrows():
                if delta_row['delta'] == 'Delta_CE':
                    newhope_df.at[index, 'Delta_CE'] = delta_row['value'] if index % 2 == 0 else delta_row['Sentiment(CE/PE)']
                elif delta_row['delta'] == 'Delta_PE':
                    newhope_df.at[index, 'Delta_PE'] = delta_row['value'] if index % 2 == 0 else delta_row['Sentiment(CE/PE)']

        # Drop unnecessary columns from probably_latest_delta_df in the merge
        final_df = newhope_df.drop(columns=['time', 'Greek_type', 'value', 'Sentiment(CE/PE)', 'delta'], errors='ignore')
        
        # Ensure the DataFrame is in the correct order as requested
        final_df = final_df[['Index', 'expiry', tf, 'Delta_CE', 'Delta_PE', 'Theta_CE', 'Theta_PE', 'Vega_CE', 'Vega_PE', 'Sentiment_1', 'Sentiment_2']]
        
        return final_df


    def update_dashboard(indexName, timeframe, dataframe):
        
        file_map = {
            'NIFTY': 'Dashboard_N_BN.xlsx',
            'BANKNIFTY': 'Dashboard_N_BN.xlsx',
            'FINNIFTY': 'Dashboard_FIN_MIDCP.xlsx',
            'MIDCPNIFTY': 'Dashboard_FIN_MIDCP.xlsx',
            'SENSEX':'Dashboard_SEN_BAN.xlsx',
            "BANKEX":'Dashboard_SEN_BAN.xlsx'      
        }

        if indexName not in file_map:
            raise ValueError(f"No Excel file mapped for indexName '{indexName}'")

        file_name = file_map[indexName]
        full_path = os.path.join(os.getcwd(), file_name)

        if xw.apps.count == 0:
            app = xw.App(visible=True)
        else:
            app = xw.apps.active

        try:

            wb = app.books[file_name]
        except KeyError:
            try:
                wb = app.books.open(full_path)
            except FileNotFoundError:

                wb = app.books.add()
                wb.save(full_path)


        filtered_df = dataframe[dataframe['Index'] == indexName]
        
        
        sheet_names = ['Sentiment_Dashboard_CW', 'Sentiment_Dashboard_NW', 'Sentiment_Dashboard_CM']

        starting_cells = {
            ('BANKNIFTY', '1min'): 'A9', ('BANKNIFTY', '3min'): 'A14', ('BANKNIFTY', '5min'): 'A19',('BANKNIFTY', '10min'): 'A24',('BANKNIFTY', '15min'): 'A29',('BANKNIFTY', '30min'): 'A34',('BANKNIFTY', 'live'): 'A4',
            ('NIFTY', '1min'): 'K9', ('NIFTY', '3min'): 'K14', ('NIFTY', '5min'): 'K19',('NIFTY', '10min'): 'K24',('NIFTY', '15min'): 'K29',('NIFTY', '30min'): 'K34',('NIFTY', 'live'): 'K4',
            ('FINNIFTY', '1min'): 'A9', ('FINNIFTY', '3min'): 'A14', ('FINNIFTY', '5min'): 'A19',('FINNIFTY', '10min'): 'A24',('FINNIFTY', '15min'): 'A29',('FINNIFTY', '30min'): 'A34',('FINNIFTY', 'live'): 'A4',
            ('MIDCPNIFTY', '1min'): 'K9', ('MIDCPNIFTY', '3min'): 'K14', ('MIDCPNIFTY', '5min'): 'K19',('MIDCPNIFTY', '10min'): 'K24',('MIDCPNIFTY', '15min'): 'K29',('MIDCPNIFTY', '30min'): 'K34',('MIDCPNIFTY', 'live'): 'K4',
            ('SENSEX', '1min'): 'A9', ('SENSEX', '3min'): 'A14', ('SENSEX', '5min'): 'A19',('SENSEX', '10min'): 'A24',('SENSEX', '15min'): 'A29',('SENSEX', '30min'): 'A34',('SENSEX', 'live'): 'A4',
            ('BANKEX', '1min'): 'K9', ('BANKEX', '3min'): 'K14', ('BANKEX', '5min'): 'K19',('BANKEX', '10min'): 'K24',('BANKEX', '15min'): 'K29',('BANKEX', '30min'): 'K34', ('BANKEX', 'live'): 'K4',    
        }

        starting_cell = starting_cells.get((indexName, timeframe), None)
        if not starting_cell:
            raise ValueError(f"Invalid combination of index name '{indexName}' and timeframe '{timeframe}' provided.")
        

        if timeframe == '1min':
            column_name="1_Min"
        
        elif timeframe == '3min':
            column_name="3_Min"

        elif timeframe== "5min":
            column_name="5_Min"

        elif timeframe== "10min":
            column_name="10_Min"

        elif timeframe== "15min":
            column_name="15_Min"
        
        elif timeframe== "30min":
            column_name="30_Min"
        
        filtered_df['expiry'] = pd.to_datetime(filtered_df['expiry'], format='%d-%m-%Y')
        expiry_dates = sorted(filtered_df['expiry'].unique())

        for i, expiry_date in enumerate(expiry_dates[:3]):
            if i < len(sheet_names):
                sheet_name = sheet_names[i]
                if sheet_name not in [sheet.name for sheet in wb.sheets]:
                    sheet = wb.sheets.add(name=sheet_name, after=len(wb.sheets))
                else:
                    sheet = wb.sheets[sheet_name]
                    # Optional: Clear contents before updating
                    # sheet.clear_contents()
                
                data_to_export = filtered_df[filtered_df['expiry'] == expiry_date]
                data_to_export = data_to_export.drop(["expiry", "Index"], axis=1)
                data_to_export.iat[1, 0] = "sentiment"
                
                data_to_export.to_csv("datatoexport.csv")

                data_to_export.reset_index(drop=True, inplace=True)
                
                print("printing.....")
                #ogfile.flush(); os.fsync(#ogfile.fileno())
                sheet.range(starting_cell).value = data_to_export.values
                
                # sheet.range(starting_cell).value = data_to_export.columns.tolist()  # Update column headers
                # sheet.range(starting_cell).offset(1,0).value = data_to_export.values  # Update data

        wb.save()
        

    def save_option_chains_to_csv(merged_option_chains_dict):

        for key, ldf in merged_option_chains_dict.items():
            file_path = f"{key}.csv"
            ldf.to_csv(file_path, index=False)
            print(f"Saved: {file_path}")
    


    def calculate_strikes(atm_value, strike_diff, option_type,method):
        
        strikes = []

        if method =='normal':

            if option_type.upper() == 'CE':
                start_strike = atm_value - (strike_diff * 2)  
            else:  
                start_strike = atm_value + (strike_diff * 2)  

            for i in range(35):  
                if option_type.upper() == 'CE':
                    strikes.append(start_strike + (i * strike_diff))
                else: 
                    strikes.append(start_strike - (i * strike_diff))
        else:
            if option_type.upper() == 'CE':
                start_strike = atm_value - (strike_diff * 2)  
            else:  
                start_strike = atm_value + (strike_diff * 2)  

            for i in range(5):  
                if option_type.upper() == 'CE':
                    strikes.append(start_strike + (i * strike_diff))
                else: 
                    strikes.append(start_strike - (i * strike_diff))

        
        return strikes


    def process_first(vega_df,theta_df,tf):   
        def prepare_df(df):
    
            df_values = df[['Index', 'expiry', 'time', 'call_put', 'value']].copy()
            df_sentiments = df[['Index', 'expiry', 'time', 'call_put', 'Sentiment(CE/PE)']].copy()
            df_values_pivot = df_values.pivot_table(index=['Index', 'expiry', 'time'], columns='call_put', values='value', aggfunc='first').reset_index()
            df_values_pivot.columns = ['Index', 'expiry', tf , 'Theta_CE', 'Theta_PE']
            df_sentiments_pivot = df_sentiments.pivot_table(index=['Index', 'expiry', 'time'], columns='call_put', values='Sentiment(CE/PE)', aggfunc='first').reset_index()
            df_sentiments_pivot.columns = ['Index', 'expiry', tf, 'Theta_CE', 'Theta_PE']
            combined_df = pd.concat([df_values_pivot, df_sentiments_pivot]).sort_values(by=['Index', 'expiry', tf], kind='mergesort').reset_index(drop=True)
            return combined_df

        theta_prepared = prepare_df(theta_df)
        vega_prepared = prepare_df(vega_df)
        final_df = pd.merge(theta_prepared, vega_prepared, on=['Index', 'expiry', tf], how='outer')
        filtered_df = final_df.groupby(['Index', 'expiry', tf]).apply(lambda x: x.iloc[[0, -1]]).reset_index(drop=True)
        filtered_df.columns = ['Index', 'expiry', tf, 'Theta_CE', 'Theta_PE', 'Vega_CE', 'Vega_PE']
        return filtered_df


    def process_second(filtered_df, new_df, new_df2, tf):
        try:
            # Ensure 'expiry' columns in all DataFrames are datetime objects
            filtered_df['expiry'] = pd.to_datetime(filtered_df['expiry'], format='%d-%m-%Y', errors='raise')
            new_df['expiry'] = pd.to_datetime(new_df['expiry'], format='%d-%m-%Y', errors='raise')
            new_df2['expiry'] = pd.to_datetime(new_df2['expiry'], format='%d-%m-%Y', errors='raise')


        except Exception as e:
            print(f"Error converting dates: {e}")
            #ogfile.flush(); os.fsync(#ogfile.fileno())
            return None

        try:
            # Merging with sentiment_1_df
            merged_df = pd.merge(filtered_df, new_df, on=['Index', 'expiry'], how='left')
            merged_df['Sentiment_2'] = None
            merged_df.loc[merged_df.index % 2 == 0, 'Sentiment_2'] = merged_df['vega_sentiment']
            merged_df.loc[merged_df.index % 2 != 0, 'Sentiment_2'] = merged_df['sentiment_2']
            # Drop unnecessary columns from merge
            merged_df.drop(['time', 'Greek_type', 'vega_sentiment', 'sentiment_2'], axis=1, inplace=True)

        except Exception as e:
            print(f"Error during merge operation: {e}")
            #ogfile.flush(); os.fsync(#ogfile.fileno())
            return None

        try:
            # Merging with sentiment_2_df
            merged_2_df = pd.merge(merged_df, new_df2, on=['Index', 'expiry'], how='left')
            merged_2_df['Sentiment_1'] = None
            merged_2_df.loc[merged_2_df.index % 2 == 0, 'Sentiment_1'] = merged_2_df['vega_sentiment']
            merged_2_df.loc[merged_2_df.index % 2 != 0, 'Sentiment_1'] = merged_2_df['sentiment_1']
            # Drop unnecessary columns from merge
            final_df = merged_2_df.drop(['time', 'Greek_type', 'vega_sentiment', 'sentiment_1'], axis=1)

        except Exception as e:
            print(f"Error during second merge operation: {e}")
            #ogfile.flush(); os.fsync(#ogfile.fileno())
            return None

        # Cleaning up the DataFrame
        final_df = final_df.loc[:, ~final_df.columns.str.contains('^Unnamed')]
        final_df.reset_index(drop=True, inplace=True)

        # Final checks and saving
        if final_df.isnull().any().any():
            print("Warning: NULL values detected in final DataFrame. Review data sources and transformations.")
            #ogfile.flush(); os.fsync(#ogfile.fileno())

    

        return final_df




    def compute_difference(sub_df, time_value,i,sentiment):

        try:
            if sentiment=="sentiment1":  
                diff = sub_df[sub_df['call_put'] == 'PE']['final_value'].values[0] - sub_df[sub_df['call_put'] == 'CE']['final_value'].values[0]   
            elif sentiment=="sentiment2":
                diff = sub_df[sub_df['call_put'] == 'CE']['final_value'].values[0] - sub_df[sub_df['call_put'] == 'PE']['final_value'].values[0]
            print("finding diff is over")
            #ogfile.flush(); os.fsync(#ogfile.fileno())
        except Exception as e:
            print("exception while finding diff",e)  
        #ogfile.flush(); os.fsync(#ogfile.fileno())  
        
        if -5 < diff < 5:
            sentiment = "Sideways"
        elif 5 <= diff < 10:
            sentiment = "Moderate Bullish"
        elif diff >= 10:
            sentiment = "Bullish"
        elif -10 < diff <= -5:
            sentiment = "Moderate Bearish"
        else:  # diff <= -10
            sentiment = "Bearish"
        
        # Return a series including the time value, calculated difference, and determined sentiment
        return pd.Series({
            'time': time_value,
            'Index': sub_df['Index'].iloc[0],
            'expiry': sub_df['expiry'].iloc[0],
            'Greek_type': sub_df['Greek_type'].iloc[0],
            'vega_sentiment': diff,
            f'sentiment_{i}': sentiment  # New sentiment_1 column based on vega_sentiment
        })
        

    def convert_date_format(date_list):
        new_format_list = []
        for date_str in date_list:
            # Parse the date string to a datetime object
            date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d')
            # Format the datetime object to the new string format
            new_date_str = datetime.datetime.strftime(date_obj, '%d-%m-%Y')
            new_format_list.append(new_date_str)
        return new_format_list

    def processing_prev_Day_Data(Ranked_dff_csv_past0,tm):
            tm = str(tm)[11:][:5].replace(':','')
            Greek_nm = 'Greek_' + tm
            df_combined = pd.DataFrame()
            Ranked_dff_csv_past0 = Ranked_dff_csv_past0[['date','name','strike','expiry','CE_PE','close','volume','iv','delta','gamma','rho','theta','vega']]
            #date  open  high  low  close  volume     oi            symbol    name     expiry  strike CE_PE  token         iv     delta        gamma       rho     theta     vega
            df_vega = Ranked_dff_csv_past0.groupby(['name','CE_PE','expiry'], as_index=False)['vega'].sum()
            df_vega = round(df_vega, 2)

            df_vega['Greek_type'] = 'vega'
            df_vega.rename(columns = {'vega':Greek_nm}, inplace = True);
            #print(df_vega.columns.values)
            df_theta = Ranked_dff_csv_past0.groupby(['name','CE_PE','expiry'], as_index=False)['theta'].sum()
            df_theta = round(df_theta, 2)
            
            df_theta['Greek_type'] = 'theta'
            df_theta.rename(columns = {'theta':Greek_nm}, inplace = True);
            #print(df_theta.columns.values)
            df_delta = Ranked_dff_csv_past0.groupby(['name','CE_PE','expiry'], as_index=False)['delta'].sum()
            df_delta = round(df_delta, 2)
            
            df_delta['Greek_type'] = 'delta'
            df_delta.rename(columns = {'delta':Greek_nm}, inplace = True);
            #print(df_delta.columns.values)
            df_iv = Ranked_dff_csv_past0.groupby(['name','CE_PE','expiry'], as_index=False)['iv'].sum()
            df_iv = round(df_iv, 2)

            df_iv['Greek_type'] = 'iv'
            df_iv.rename(columns = {'iv':Greek_nm}, inplace = True);
            #print('df_vega:');print(df_vega);print(df_theta);print(df_delta)
            #df_combined1 = pd.merge(df_vega,df_theta,on=['name','CE_PE','expiry'])
            #df_combined = pd.merge(df_combined1,df_delta,on=['name','CE_PE','expiry'])
            df_combined = pd.concat([df_combined, df_vega], ignore_index=True, axis=0)#;print('After Vega:',len(df_combined.index))
            df_combined = pd.concat([df_combined, df_theta], ignore_index=True, axis=0)#;print('After Theta:',len(df_combined.index))
            df_combined = pd.concat([df_combined, df_delta], ignore_index=True, axis=0)#;print('After Delta:',len(df_combined.index))
            df_combined = pd.concat([df_combined, df_iv], ignore_index=True, axis=0)#;print('After iv:',len(df_combined.index))
            #df_combined = df_combined.round({"iv":4, "delta":4, "gamma":10, "rho":10, "theta":10, "vega":10})
            df_combined = df_combined[['name','expiry','CE_PE','Greek_type',Greek_nm]]
            #print('df_combined:');print(df_combined)
            return df_combined


    def update_and_calculate_changes(grouped_df, empty_df):
        if empty_df.empty:
            
            empty_df = grouped_df.copy()
            empty_df['CE_change'] = pd.NA
            empty_df['PE_change'] = pd.NA
            empty_df['Fut_change'] = pd.NA
        else:
        
            for index, row in grouped_df.iterrows():
                # Extract last known values for the same 'name' and 'expiry'
                last_row = empty_df[(empty_df['name'] == row['name']) & (empty_df['expiry'] == row['expiry'])].iloc[-1]
                ce_change = ((row['ATV_CE'] - last_row['ATV_CE']) / abs(last_row['ATV_CE'])) * 100 if last_row['ATV_CE'] != 0 else 0
                pe_change = ((row['ATV_PE'] - last_row['ATV_PE']) / abs(last_row['ATV_PE'])) * 100 if last_row['ATV_PE'] != 0 else 0
                fut_change=((row['Fut_Atv'] - last_row['Fut_Atv']) / abs(last_row['Fut_Atv'])) * 100 if last_row['Fut_Atv'] != 0 else 0
                
                # Set new change values in grouped_df before appending
                grouped_df.at[index, 'CE_change'] = ce_change
                grouped_df.at[index, 'PE_change'] = pe_change
                grouped_df.at[index, 'Fut_change'] = fut_change

            # Append new rows from grouped_df to empty_df
            empty_df = pd.concat([empty_df, grouped_df], ignore_index=True)

        return empty_df

    def determine_delta_sentiment(value, option_type):
        if option_type == "PE":
            if -5 < value < 5:
                return "Sideways"
            elif 5 <= value < 10:
                return "Moderate Bearish"
            elif value >= 10:
                return "Bearish"
            elif -10 < value <= -5:
                return "Moderate Bullish"
            else:  # value <= -10
                return "Bullish"
        elif option_type == "CE":
            if -5 < value < 5:
                return "Sideways"
            elif -10 <= value < -5:
                return "Moderate Bearish"
            elif value < -10:
                return "Bearish"
            elif 5 <= value < 10:
                return "Moderate Bullish"
            else:  # value >= 10
                return "Bullish"

    def process_option_chains(dfs_dict):
        for key, df in dfs_dict.items():
            # Calculate the totals for the specified columns
            totals = {
                'change_ce_oi':df['change_ce_oi'].sum(),
                'oi_ce': df['oi_ce'].sum(),
                'Vol_ce': df['Vol_ce'].sum(),
                'delta_ce':df['delta_ce'].sum(),
                'delta_change_ce':df['delta_change_ce'].sum(),
                'theta_ce':df['theta_ce'].sum(),
                'theta_change_ce':df['theta_change_ce'].sum(),
                'vega_ce': df['vega_ce'].sum(),
                'vega_change_ce':df['vega_change_ce'].sum(),
                'oi_pe': df['oi_pe'].sum(),
                'Vol_pe': df['Vol_pe'].sum(),
                'delta_pe':df['delta_pe'].sum(),
                'delta_change_pe':df['delta_change_pe'].sum(),
                'theta_pe':df['theta_pe'].sum(),
                'theta_change_pe':df['theta_change_pe'].sum(),
                'vega_pe': df['vega_pe'].sum(),
                'vega_change_pe':df['vega_change_pe'].sum(),
                'change_pe_oi':df['change_pe_oi'].sum(),
                
            }
            # Create a DataFrame with the totals to concatenate with the original DataFrame
            totals_df = pd.DataFrame([totals])
            # Concatenate the totals DataFrame to the original DataFrame
            df = pd.concat([df, totals_df], ignore_index=True)
            # Calculate the percentage for the specified columns (except the last row)
            for column in ['change_ce_oi','oi_ce', 'Vol_ce', 'oi_pe', 'Vol_pe','change_pe_oi']:
                # df.loc[:df.index[-2], column] = (df[column] / totals[column]) * 100
                if totals[column]==0:
                    totals[column]=1
                df.loc[:df.index[-2], column] = (df[column] / totals[column]) * 100

                
            # Reassign the total values to the last row for the specified columns
            for column in totals.keys():
                df.at[df.index[-1], column] = totals[column]
            
            # Update the dictionary with the modified DataFrame
            dfs_dict[key] = df

        return dfs_dict





    def merge_option_chains(ce_df, pe_df):

        ce_df_renamed = ce_df.drop(columns=["Unnamed: 0"], errors='ignore')
        pe_df_renamed = pe_df.drop(columns=["Unnamed: 0"], errors='ignore')

        # ce_df_renamed.columns = ['oi_ce', 'Vol_ce', 'delta_ce', 'theta_ce', 'vega_ce', 'LTP_ce', 'strike']
        ce_df_renamed.columns = ['change_ce_oi','oi_ce', 'Vol_ce', 'delta_ce','delta_change_ce', 'theta_ce', 'theta_change_ce','vega_ce', 'vega_change_ce','noted_ltp_ce','LTP_ce','ltp_change_ce', 'strike']

        # pe_df_renamed.columns = ['strike', 'LTP_pe', 'vega_pe', 'theta_pe', 'delta_pe', 'Vol_pe', 'oi_pe']
        pe_df_renamed.columns = ['strike','noted_ltp_pe', 'LTP_pe', 'ltp_change_pe','vega_pe', 'vega_change_pe','theta_pe', 'theta_change_pe','delta_pe','delta_change_pe', 'Vol_pe', 'oi_pe','change_pe_oi']


        ce_df_sorted = ce_df_renamed.sort_values(by='strike', ascending=False).reset_index(drop=True)
        extra_strikes = set(pe_df_renamed['strike']) - set(ce_df_sorted['strike'])
        extra_rows = pd.DataFrame({'strike': list(extra_strikes)})
        for col in ce_df_sorted.columns:
            if col not in extra_rows.columns:
                extra_rows[col] = None

        ce_df_augmented = pd.concat([ce_df_sorted, extra_rows], ignore_index=True).sort_values(by='strike', ascending=False).reset_index(drop=True)
        merged_df = pd.merge(ce_df_augmented, pe_df_renamed, on='strike', how='outer').sort_values(by='strike', ascending=False).reset_index(drop=True)

        return merged_df


    def merge_all_option_chains(dfs_dict):
        
        ce_columns_to_keep = ['change_ce_oi','oi', 'Vol', 'delta','delta_change_ce', 'theta', 'theta_change_ce','vega', 'vega_change_ce','first_ltp','LTP','ltp_change_ce', 'strike']
        pe_columns_to_keep = ['strike','first_ltp', 'LTP', 'ltp_change_pe','vega', 'vega_change_pe','theta', 'theta_change_pe','delta','delta_change_pe', 'Vol', 'oi','change_pe_oi']
        
        merged_option_chains_dict = {}
        try:
            for key, df in dfs_dict.items():
                if '_ce_' in key:
                    df = df[ce_columns_to_keep]
                else:
                    df = df[pe_columns_to_keep]
                
                dfs_dict[key] = df
        except Exception as e:
            print("error in first step",e)
            

        index_prefixes = ['banknifty', 'nifty', 'finnifty', 'midcpnifty', 'sensex','bankex']
        expirations = ['currentexp', 'nextexp']
        try:
            for index in index_prefixes:
                for exp in expirations:
                    ce_key = f'{index}_ce_{exp}'
                    pe_key = f'{index}_pe_{exp}'

                    if ce_key in dfs_dict and pe_key in dfs_dict:
                        # Ensure the columns specified exist in the DataFrame
                        assert set(ce_columns_to_keep).issubset(dfs_dict[ce_key].columns), "Missing CE columns"
                        assert set(pe_columns_to_keep).issubset(dfs_dict[pe_key].columns), "Missing PE columns"
                        
                        # Check if the number of columns to keep is the same for CE and PE
                        assert len(ce_columns_to_keep) == len(pe_columns_to_keep), "CE and PE column lists are of different lengths"
                        

                        merged_key = f'{index}_{exp}_opt_chain'
                        merged_option_chains_dict[merged_key] = merge_option_chains(dfs_dict[ce_key], dfs_dict[pe_key])
        except AssertionError as e:
            print("Assertion Error:", e)
        except Exception as e:
            print("Error in second step:", e)

        # try:
        #     for index in index_prefixes:
        #         for exp in expirations:
        #             ce_key = f'{index}_ce_{exp}'
        #             pe_key = f'{index}_pe_{exp}'
                    
        #             if ce_key in dfs_dict and pe_key in dfs_dict:
        #                 merged_key = f'{index}_{exp}_opt_chain'
        #                 merged_option_chains_dict[merged_key] = merge_option_chains(dfs_dict[ce_key], dfs_dict[pe_key])
        # except Exception as e:
        #     print("error in fsecond step",e)

        

        return merged_option_chains_dict


    def trim_dataframe(df, total_rows=30):

        rows_to_keep = total_rows - 1  # Since we need to keep the last row intact
        rows_to_remove_each_side = (len(df) - rows_to_keep) // 2

        if rows_to_remove_each_side > 0:
            df_trimmed = df[rows_to_remove_each_side:-rows_to_remove_each_side].copy()
        else:
            df_trimmed = df.copy()

        if not df_trimmed.iloc[-1].equals(df.iloc[-1]):
            df_trimmed = pd.concat([df_trimmed, df.iloc[[-1]]])

        return df_trimmed

    def process_df_to_dfs(df_input):    
        df_input['expiry'] = pd.to_datetime(df_input['expiry'], format='%Y-%m-%d')
        bn_expiry_dates = np.sort(df_input[df_input['name'] == 'BANKNIFTY']['expiry'].unique())
        n_expiry_dates = np.sort(df_input[df_input['name'] == 'NIFTY']['expiry'].unique())
        bn_current_week_expiry, bn_next_week_expiry = bn_expiry_dates[:2] if len(bn_expiry_dates) >= 2 else (bn_expiry_dates[0], None)
        n_current_week_expiry, n_next_week_expiry = n_expiry_dates[:2] if len(n_expiry_dates) >= 2 else (n_expiry_dates[0], None)    
        required_dfs = {}
        
        def filter_and_assign(index_name, option_type, expiry_date, key_suffix):
            df_filtered = df_input[(df_input['name'] == index_name) & (df_input['CE_PE'] == option_type) & (df_input['expiry'] == expiry_date)]
            required_dfs[f"{index_name.lower()}_{option_type.lower()}_{key_suffix}"] = df_filtered
        
        for index_name in ['BANKNIFTY', 'NIFTY']:
            for option_type in ['CE', 'PE']:
                current_week_expiry = bn_current_week_expiry if index_name == 'BANKNIFTY' else n_current_week_expiry
                next_week_expiry = bn_next_week_expiry if index_name == 'BANKNIFTY' else n_next_week_expiry            
                filter_and_assign(index_name, option_type, current_week_expiry, 'currentexp')
                if next_week_expiry is not None:
                    filter_and_assign(index_name, option_type, next_week_expiry, 'nextexp')
        
        for index_name in ['FINNIFTY', 'MIDCPNIFTY', 'SENSEX','BANKEX']:
            for option_type in ['CE', 'PE']:
                current_week_expiry = df_input[df_input['name'] == index_name]['expiry'].min()
                filter_and_assign(index_name, option_type, current_week_expiry, 'currentexp')
        
        return required_dfs

    def add_change_columns(required_dfs):
        for key, df in required_dfs.items():
            
            option_type = 'ce' if 'CE' in df['CE_PE'].iloc[0] else 'pe'
            # df[f'change_{option_type}_oi'] = df[f'change_{option_type}_oi'] = ((df['oi'] - df['first_oi']) / df['first_oi']) * 100  
            df[f'change_{option_type}_oi'] = df['oi'] - df['first_oi']
            df[f'delta_change_{option_type}'] = df['delta'] - df['first_delta']
            df[f'theta_change_{option_type}'] = df['first_theta'] - df['theta']
            if option_type == 'ce':
                df['vega_change_ce'] = df['vega'] - df['first_vega']
            else:
                df['vega_change_pe'] = df['vega'] - df['first_vega']
                
            df[f'ltp_change_{option_type}'] = df['LTP'] - df['first_ltp']

        return required_dfs

    def update_strike_counts(kdf):
        
        if 'count' not in kdf.columns:
            kdf['count'] = 1  
        else:
            
            count_series = kdf.groupby(['Strike', 'CE_PE']).cumcount() + 1
            kdf['count'] = count_series



        return kdf



    def add_greeks_to_dfs(required_dfs, df_greeks):
        try:
            for key, df in required_dfs.items():
                df['first_oi'] = None
                df['first_delta'] = None
                df['first_theta'] = None
                df['first_vega'] = None
                df['first_ltp'] = None

                for index, row in df.iterrows():
                    match = df_greeks[(df_greeks['name'] == row['name']) & 
                                    (df_greeks['CE_PE'] == row['CE_PE']) & 
                                    (df_greeks['strike'] == row['strike']) & 
                                    (df_greeks['Stock'] == row['Stock'])]
                    
                    # Debugging output
                    if match.empty:
                        print(f"No match found for row index {index} in dataframe {key}")
                        print(row)
                    else:
                        df.loc[index, 'first_oi'] = match.iloc[0]['oi']
                        df.loc[index, 'first_delta'] = match.iloc[0]['delta']
                        df.loc[index, 'first_theta'] = match.iloc[0]['theta']
                        df.loc[index, 'first_vega'] = match.iloc[0]['vega']
                        df.loc[index, 'first_ltp'] = match.iloc[0]['LTP']

            return required_dfs

        except Exception as e:
            print("error in new is", e)


    def add_decay_and_percentage_changes(df):
        def add_columns(df, prefix, compare_greater=False):
            decay_column = f'{prefix}_decay'
            pct_change_column = f'{prefix}_%_change'
            
            ce_column = f'{prefix}_change_ce'
            pe_column = f'{prefix}_change_pe'
            
            if compare_greater:
                df[pct_change_column] = df.apply(lambda x: (x[ce_column] + x[pe_column]), axis=1)
                df[decay_column] = df.apply(
                            lambda x: 'Put Side' if (x[ce_column] + x[pe_column]) > 0 else 'Call Side' if (x[ce_column] + x[pe_column]) < 0 else 'Neutral', axis=1)
    
                
                # This line determines which side has the greater absolute change for 'delta'
                # df[decay_column] = df.apply(lambda x: 'Call Side' if (x[ce_column] > x[pe_column]) else 'Put Side', axis=1)
            else:
                # This line determines which side has the lesser absolute change for 'vega' and 'theta'
                df[decay_column] = df.apply(lambda x: 'Call Side' if (x[ce_column] < x[pe_column]) else 'Put Side', axis=1)
            
            # This line calculates the absolute difference between the changes of CE and PE
                df[pct_change_column] = df.apply(lambda x: abs(x[ce_column] - x[pe_column]), axis=1)

        # Applying the function with the appropriate comparison type
        for prefix in ['vega', 'theta']:
            add_columns(df, prefix, compare_greater=False)
        
        # For 'delta', compare which is greater
        add_columns(df, 'delta', compare_greater=True)

        return df



    def clear_or_fill_specific_ranges(file_names, special_files_info):
        
        if xw.apps.count == 0:
            app = xw.App(visible=False)  
        else:
            app = xw.apps.active  

        for file_name in file_names:
            full_path = os.path.join(os.getcwd(), file_name)

            try:
                wb = app.books.open(full_path)
            except FileNotFoundError:
                wb = app.books.add()
                wb.save(full_path)

            
            if file_name in special_files_info:
                for sheet in wb.sheets:
                    starting_cells = special_files_info[file_name]
                    for start_cell in starting_cells:
                        end_cell = sheet.range(start_cell).offset(row_offset=1, column_offset=8)
                        range_to_clear = sheet.range(start_cell, end_cell)
                        range_to_clear.clear_contents()
            else:
                
                for i, sheet in enumerate(wb.sheets, start=1):
                    if not (file_name == 'Zerodha Live Quotes NFO Greeks_Project.xlsx' and i == 1):
                        sheet.used_range.clear_contents()

            wb.save()
            wb.close()

        if len(app.books) == 0:
            app.quit()


    def process_df(df, tm):
        df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
        if 'Greek_1529' in df.columns:
            greek_1529_index = df.columns.get_loc('Greek_1529')
            df = df.iloc[:, list(range(greek_1529_index)) + [-1]]
        df.rename(columns={'final_value': f'Greek_{tm}'}, inplace=True)
        return df

    def create_bank_dfs(volumes, prev_vol_df, noted_ce_oi_corrected, noted_pe_oi_corrected, result_df, tm):
        bank_dfs = {}
        symbols = ["HDFCBANK", "ICICIBANK", "SBIN", "AXISBANK", "KOTAKBANK", "INDUSINDBK"]
        nse_symbols = ["NSE:HDFCBANK", "NSE:ICICIBANK", "NSE:SBIN", "NSE:AXISBANK", "NSE:KOTAKBANK", "NSE:INDUSINDBK"]

        for symbol, nse_symbol in zip(symbols, nse_symbols):
            prev_day_volume = prev_vol_df.loc[prev_vol_df['symbol'] == symbol, 'volume'].iloc[0]
            
            current_day_volume = volumes[nse_symbol]
            
            change_volume = current_day_volume-prev_day_volume
            change_volume_percent = (change_volume / prev_day_volume) * 100
            change_volume_percent= round(change_volume_percent,2)
            noted_ce_oi = noted_ce_oi_corrected[symbol]
            noted_pe_oi = noted_pe_oi_corrected[symbol]
            total_ce_oi_new = result_df[(result_df['name'] == symbol) & (result_df['CE_PE'] == 'CE')]['oi'].sum()
            
            total_pe_oi_new = result_df[(result_df['name'] == symbol) & (result_df['CE_PE'] == 'PE')]['oi'].sum()
            

            total_oi = total_ce_oi_new + total_pe_oi_new
            call_oi_percent = (total_ce_oi_new / total_oi) * 100 if total_oi != 0 else 0
            call_oi_percent=round(call_oi_percent,2)
            put_oi_percent = (total_pe_oi_new / total_oi) * 100 if total_oi != 0 else 0
            put_oi_percent=round(put_oi_percent,2)
            change_in_call_oi_percent = ((total_ce_oi_new-noted_ce_oi) / noted_ce_oi) * 100 if noted_ce_oi != 0 else 0
            change_in_call_oi_percent=round(change_in_call_oi_percent,2)
            change_in_put_oi_percent = ((total_pe_oi_new-noted_pe_oi) / noted_pe_oi) * 100 if noted_pe_oi != 0 else 0
            change_in_put_oi_percent=round(change_in_put_oi_percent,2)
            row_data = {
                "Live_Data": tm,
                "Prev_D_Volume": round((prev_day_volume / 1e5),2),  
                "Curr_D_Volume": round((current_day_volume / 1e5),2),  
                "Change_Volume": round((change_volume / 1e5),2), 
                "Change_Volume_%": change_volume_percent,  
                "Noted_CE_OI": noted_ce_oi / 1e5,  
                "Call_OI_%": call_oi_percent,  
                "Change in Call OI_%": change_in_call_oi_percent,  
                "Noted_PE_OI": noted_pe_oi / 1e5,  
                "Put_OI_%": put_oi_percent, 
                "Change in Put OI_%": change_in_put_oi_percent  
            }

            bank_dfs[symbol] = pd.DataFrame([row_data])
        
        return bank_dfs


    def clear_excel_sheets(file_names, sheet_info):
        # Ensure Excel is not visible to speed up operations
        if xw.apps.count == 0:
            app = xw.App(visible=False)
        else:
            app = xw.apps.active

        for file_name, sheets in sheet_info.items():
            try:
                full_path = os.path.join(os.getcwd(), file_name)
                wb = app.books.open(full_path)
                for sheet_name in sheets:
                    sheet = wb.sheets[sheet_name]
                    # Determine the last cell in the sheet dynamically
                    last_cell = sheet.cells.last_cell.get_address()
                    # Construct the range from the 3rd row to the last cell
                    range_to_clear = f'A3:{last_cell}'
                    # Clear contents from the third row downwards
                    sheet.range(range_to_clear).clear_contents()
                wb.save()
                wb.close()
            except Exception as e:
                print(f"Failed to clear {sheet_name} in {file_name} due to: {e}")

        # Close the Excel application if it was opened by this script
        if len(app.books) == 0:
            app.quit()


    def fill_excel_sheets_with_empty_df(file_name, sheet_names, rows=1000, cols=50, filler_value=None):
        # Ensure Excel is not visible to speed up operations
        if xw.apps.count == 0:
            app = xw.App(visible=False)
        else:
            app = xw.apps.active

        full_path = os.path.join(os.getcwd(), file_name)
        wb = app.books.open(full_path)
        
        # Create an empty DataFrame with specified dimensions and filler value
        empty_df = pd.DataFrame(filler_value, index=range(rows), columns=range(cols))

        for sheet_name in sheet_names:
            sheet = wb.sheets[sheet_name]
            # Write the empty DataFrame to Excel starting from cell A3
            sheet.range('A3').options(index=False, header=False).value = empty_df
        
        wb.save()
        wb.close()

        # Close the Excel application if it was opened by this script
        if len(app.books) == 0:
            app.quit()


    # Mapping of file names to their respective sheets
    sheet_info = {
        'NF_Pro_Trading.xlsx': ['VEGA & Theta_Chng_CW', 'VEGA & Theta_Chng_NW', 'Fii & PRO vs Retailers'],
        'BNF_Pro_Trading.xlsx': ['VEGA & Theta_Chng_CW', 'VEGA & Theta_Chng_NW', 'Fii & PRO vs Retailers'],
        'FIN_Pro_Trading.xlsx': ['VEGA & Theta_Chng_CW', 'Fii & PRO vs Retailers'],
        'Midcp_Pro_Trading.xlsx': ['VEGA & Theta_Chng_CW', 'Fii & PRO vs Retailers'],
        'Sen_Pro_Trading.xlsx': ['VEGA & Theta_Chng_CW', 'Fii & PRO vs Retailers'],
        # "Sen_theta_delta.xlsx": ['main'],
        # "NF_theta_delta.xlsx": ['main'],
        # "Midcp_theta_delta.xlsx": ['main'],
        # "FIN_theta_delta.xlsx": ['main'],
        # "BNF_theta_delta.xlsx": ['main'],
    }
    def clear_sheet_contents(file_path, sheet_name):
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.value = None

            workbook.save(file_path)
            print(f"All contents of the sheet '{sheet_name}' in '{file_path}' have been cleared.")
        except Exception as e:
            print(f"An error occurred: {e}")

    print("entered bef main func")

    if True:

        print("entered into main func")
        if True:
            if new_day:
                # clear_excel_sheets(['NF_Pro_Trading.xlsx', 'BNF_Pro_Trading.xlsx', 'FIN_Pro_Trading.xlsx', 'Midcp_Pro_Trading.xlsx', 'Sen_Pro_Trading.xlsx'], sheet_info)
                clear_sheet_contents("NF_theta_delta.xlsx", "main")
                clear_sheet_contents("BNF_theta_delta.xlsx", "main")
                clear_sheet_contents("FIN_theta_delta.xlsx", "main")
                clear_sheet_contents("Sen_theta_delta.xlsx", "main")
                clear_sheet_contents("Midcp_theta_delta.xlsx", "main")
                clear_sheet_contents("BAN_theta_delta.xlsx", "main")
                clear_sheet_contents("Sen_Fii_Pro.xlsx", "Fii & PRO vs Retailers")
                clear_sheet_contents("NF_Fii_Pro.xlsx", "Fii & PRO vs Retailers")
                clear_sheet_contents("BNF_Fii_Pro.xlsx", "Fii & PRO vs Retailers")
                clear_sheet_contents("FIN_Fii_Pro.xlsx", "Fii & PRO vs Retailers")
                clear_sheet_contents("Mid_Fii_Pro.xlsx", "Fii & PRO vs Retailers")
                clear_sheet_contents("BAN_Fii_Pro.xlsx", "Fii & PRO vs Retailers")

                print("old data in files have been cleared")
                #ogfile.flush(); os.fsync(#ogfile.fileno())       
            else:
                print("Since algo started again in the same day, NOT erasing the old data")
                #ogfile.flush(); os.fsync(#ogfile.fileno())
                
            global df,df_full_all,socket_opened,df_tick_dump,df_dummy,df_dummy2,df_dummy3
            global Weekly_expiry,Next_week_expiry,Monthly_expiry,instruments_list,instruments_list_Full
            Weekly_expiry='';Next_week_expiry='';Monthly_expiry = '';instruments_list = [];instruments_list_Full=[]
        
            curr_tm_chk1 = time.strftime("%Y%m%d_%H%M%S", time.localtime())
            
            global df_instruments,df_tick_master,batcher,master_file_nm,rotation;df_tick_master = pd.DataFrame();batcher=0
            master_file_nm = 'All_Data_Master_'+str(curr_tm_chk1)+'_'
            rotation = 1
            url = "https://api.kite.trade/instruments"
            dfw = pd.read_csv(url)
            dfw.to_csv('instruments.csv', index=False,header=True)
            dfw = pd.read_csv('instruments.csv',dtype={"lot_size":"int"})
            df_instruments = dfw[( (((dfw.name == "BANKNIFTY") | (dfw.name == "NIFTY") | (dfw.name == "FINNIFTY") | (dfw.name == "BANKEX") | (dfw.name == "SENSEX")| (dfw.name == "MIDCPNIFTY")) ))] #& (dfw.segment == 'NFO-OPT')
            df_instruments.to_csv('instruments.csv', index=False,header=True)
            


            global df_instruments_new
            df_instruments_new = dfw[(((dfw.name == "NIFTY") | (dfw.name == "BANKNIFTY") | (dfw.name == "FINNIFTY") | (dfw.name == "MIDCPNIFTY")) & ((dfw.segment == 'NFO-OPT') | (dfw.segment == 'NFO-FUT'))) | ((dfw.segment == 'INDICES') & (dfw.exchange == 'NSE') & (((dfw.tradingsymbol == "NIFTY 50") | (dfw.tradingsymbol == "NIFTY BANK") | (dfw.tradingsymbol == "NIFTY FIN SERVICE") | (dfw.tradingsymbol == "NIFTY MID SELECT")))) | (((dfw.name == "SENSEX") | (dfw.name == "BANKEX")) & ((dfw.segment == 'BFO-OPT') | (dfw.segment == 'BFO-FUT'))) | ((dfw.segment == 'INDICES') & (dfw.exchange == 'BSE') & ((dfw.tradingsymbol == "SENSEX") | (dfw.tradingsymbol == "BANKEX"))) ]
            

            print('Extracting Login details.')
            #ogfile.flush(); os.fsync(#ogfile.fileno())
        

            curr_tm_new = time.strftime("%Y%m%d", time.localtime())
        
            
        
        
        if True:
            #socket_opened = False
            df = pd.DataFrame();print('\n ############STARTED Refresh!!. ALL the best!!#############')
            df_tick_dump = pd.DataFrame();
            #exit()

            Data_instruments = pd.read_csv('instruments.csv')
            Data_instruments['expry'] = Data_instruments["expiry"].astype(str) #'2021-08-05' 'BANKNIFTY'
            #Data_instruments['strike'] = pd.to_numeric(Data_instruments['strike'])

            Manual_expiry = np.str_('2021-01-20')
            Expiry_Day_use_Expiry_to_USE_W = 'NO' 
            Expiry_Day_use_Expiry_to_USE_M = 'NO'
            FIFTY_strike_To_Consider = 'YES'
            ##is last week of the expiry
            now = datetime.datetime.now() 
            current_DAY = now.strftime("%A").upper() #("%H:%M:%S")
            today = now.strftime('%Y-%m-%d')
            print('Today is:',current_DAY,today,'Lets Rock!!!');Expiry_to_use = 'WEEKLY';
            exp_logic={}
        if True:
            global Three_expiries_BN,Two_expiries_BN,One_expiry_BN,Three_expiries_NY,Two_expiries_NY,One_expiry_NY
            Expiry_dt,MonthlyExpiry = Expiry_selection('WEEKLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'BANKNIFTY',[])
            Expiry_dt_nw,MonthlyExpiry = Expiry_selection('NEXT_WEEK',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'BANKNIFTY',[])
            Expiry_dt_nm,MonthlyExpiry = Expiry_selection('MONTHLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'BANKNIFTY',[Expiry_dt,Expiry_dt_nw])
            Three_expiries_BN = [Expiry_dt,Expiry_dt_nw,Expiry_dt_nm]
            Two_expiries_BN = [Expiry_dt,Expiry_dt_nw]
            Three_expiries_BN=Two_expiries_BN 
            One_expiry_BN = [Expiry_dt]
            print('CW NW CM Expiries list Banknifty:',One_expiry_BN,Two_expiries_BN,Three_expiries_BN)
            exp_logic['BANKNIFTY']=Expiry_dt
            print(Three_expiries_BN)
            bn_monthly=[MonthlyExpiry]
            print("bn monthly expiry is",bn_monthly)
        
            #ogfile.flush(); os.fsync(#ogfile.fileno())
            
            Expiry_dt,MonthlyExpiry = Expiry_selection('WEEKLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'NIFTY',[])
            Expiry_dt_nw,MonthlyExpiry = Expiry_selection('NEXT_WEEK',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'NIFTY',[])
            Expiry_dt_nm,MonthlyExpiry = Expiry_selection('MONTHLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'NIFTY',[Expiry_dt,Expiry_dt_nw])
            Three_expiries_NY = [Expiry_dt,Expiry_dt_nw,Expiry_dt_nm]
            Two_expiries_NY = [Expiry_dt,Expiry_dt_nw]
            Three_expiries_NY=Two_expiries_NY
            One_expiry_NY = [Expiry_dt]
            n_monthly=[MonthlyExpiry]
            print('CW NW CM Expiries list Nifty:',One_expiry_NY,Two_expiries_NY,Three_expiries_NY)
            exp_logic['NIFTY']=Expiry_dt

            #ogfile.flush(); os.fsync(#ogfile.fileno())
            Expiry_dt,MonthlyExpiry = Expiry_selection('WEEKLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'FINNIFTY',[])
            Expiry_dt_nw,MonthlyExpiry = Expiry_selection('NEXT_WEEK',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'FINNIFTY',[])
            Expiry_dt_nm,MonthlyExpiry = Expiry_selection('MONTHLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'FINNIFTY',[Expiry_dt,Expiry_dt_nw])
            # Three_expiries_FIN = [Expiry_dt,Expiry_dt_nw,Expiry_dt_nm]
            Two_expiries_FIN = [Expiry_dt,Expiry_dt_nw]
            One_expiry_FIN = [Expiry_dt]
            Three_expiries_FIN= [Expiry_dt]
            exp_logic['FINNIFTY']=Expiry_dt
            fin_monthly=[MonthlyExpiry]


            print('CW NW CM Expiries list FINNifty:',One_expiry_FIN,Two_expiries_FIN,Three_expiries_FIN)
            #ogfile.flush(); os.fsync(#ogfile.fileno())
            Expiry_dt,MonthlyExpiry = Expiry_selection('WEEKLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'SENSEX',[])
            Expiry_dt_nw,MonthlyExpiry = Expiry_selection('NEXT_WEEK',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'SENSEX',[])
            Expiry_dt_nm,MonthlyExpiry = Expiry_selection('MONTHLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'SENSEX',[Expiry_dt,Expiry_dt_nw])
            
            Two_expiries_SEN = [Expiry_dt]
            Three_expiries_SEN = Two_expiries_SEN
            One_expiry_SEN = [Expiry_dt]
            exp_logic['SENSEX']=Expiry_dt
            sen_monthly=[MonthlyExpiry]
            print("sensex monthly expiry is",MonthlyExpiry)
            

            print('CW NW CM Expiries list SENSEX:',One_expiry_SEN,Two_expiries_SEN,Three_expiries_SEN)
            #ogfile.flush(); os.fsync(#ogfile.fileno())
            Expiry_dt,MonthlyExpiry = Expiry_selection('WEEKLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'MIDCPNIFTY',[])
            Expiry_dt_nw,MonthlyExpiry = Expiry_selection('NEXT_WEEK',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'MIDCPNIFTY',[])
            Expiry_dt_nm,MonthlyExpiry = Expiry_selection('MONTHLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'MIDCPNIFTY',[Expiry_dt,Expiry_dt_nw])
            
            Two_expiries_MID = [Expiry_dt]
            Three_expiries_MID = Two_expiries_MID
            One_expiry_MID = [Expiry_dt]
            print('CW NW CM Expiries list MIDCPNIFTY:',One_expiry_MID,Two_expiries_MID,Three_expiries_MID)
            exp_logic['MIDCPNIFTY']=Expiry_dt
            mid_monthly=[MonthlyExpiry]
            
            #ogfile.flush(); os.fsync(#ogfile.fileno())
            Expiry_dt,MonthlyExpiry = Expiry_selection('WEEKLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'BANKEX',[])
            Expiry_dt_nw,MonthlyExpiry = Expiry_selection('NEXT_WEEK',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'BANKEX',[])
            Expiry_dt_nm,MonthlyExpiry = Expiry_selection('MONTHLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'BANKEX',[Expiry_dt,Expiry_dt_nw])
            
            Two_expiries_BANK = [Expiry_dt,Expiry_dt_nw]
            One_expiry_BANK = [Expiry_dt]
            Three_expiries_BANK = [Expiry_dt] 
        


            print('CW NW CM Expiries list BANKEX:',One_expiry_BANK,Two_expiries_BANK,Three_expiries_BANK)
            # Expiry_dt,MonthlyExpiry = Expiry_selection('WEEKLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'HDFCBANK',[])
            # Three_expiries_HDFCBANK = [Expiry_dt]

            # print("HDFCBANK expiry date is",Three_expiries_HDFCBANK)
            # Expiry_dt,MonthlyExpiry = Expiry_selection('WEEKLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'ICICIBANK',[])
            # Three_expiries_ICICIBANK = [Expiry_dt]

            # print("ICICI expiry date is",Three_expiries_ICICIBANK)
            # Expiry_dt,MonthlyExpiry = Expiry_selection('WEEKLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'SBIN',[])
            # Three_expiries_SBIN = [Expiry_dt]

            # print("SBI expiry date is",Three_expiries_SBIN)
            # Expiry_dt,MonthlyExpiry = Expiry_selection('WEEKLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'KOTAKBANK',[])
            # Three_expiries_KOTAKBANK = [Expiry_dt]

            # print("KOTAKBANK expiry date is",Three_expiries_KOTAKBANK)
            # Expiry_dt,MonthlyExpiry = Expiry_selection('WEEKLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'AXISBANK',[])
            # Three_expiries_AXISBANK = [Expiry_dt]

            # print("AXISBANK expiry date is",Three_expiries_AXISBANK)
            # Expiry_dt,MonthlyExpiry = Expiry_selection('WEEKLY',Manual_expiry,Expiry_Day_use_Expiry_to_USE_W,Expiry_Day_use_Expiry_to_USE_M,'INDUSINDBK',[])
            # Three_expiries_INDUSINDBANK = [Expiry_dt]

            # print("INDUSIND expiry date is",Three_expiries_INDUSINDBANK)
        
            
        def find_expiry_today(dates, today):
            expiring_today = []
            for key, date in dates.items():
                if date == today:
                    expiring_today.append(key)
            return expiring_today

        expiry_today = find_expiry_today(exp_logic, today)
        print("Todays expiring index list is",expiry_today)
        actual_expiry_index=expiry_today.copy()
        if 'BANKNIFTY' not in expiry_today:
            expiry_today.append('BANKNIFTY')

        if 'NIFTY' not in expiry_today:
            expiry_today.append('NIFTY')
        
        one_time_print=0
        entered=0  

        bank_cw, bank_nw, nif_cw, nif_nw, mid_cw, fin_cw, sen_cw = [pd.DataFrame() for _ in range(7)];ban_cw=pd.DataFrame()

        print('Going to wait till to take the data!')
        while entered == 0:
            if True:
                global nifty_above_limit,nifty_below_limit,B_nifty_above_limit,B_nifty_below_limit,df_z_Token_Symbol
                ltp_nifty = 0;ltp_Banknifty = 0; 
                ltp_nifty,ltp_Banknifty,ltp_finnifty,ltp_sensex,ltp_bankex,ltp_midcpnif = get_N_BN_spot()
                nifty_above_limit = ltp_nifty + 50
                nifty_below_limit = ltp_nifty - 50
                B_nifty_above_limit = ltp_Banknifty + 100
                B_nifty_below_limit = ltp_Banknifty - 100
                entered = 1
                First_Candle_Start_tm = ''
                if True:
                                ##Nifty
                                Price_Difference = 50
                                val2 = math.fmod(ltp_nifty, Price_Difference); #print('val2:',val,val2)
                                if val2 >=(Price_Difference/2) :
                                    ATM_PRICE_N = ltp_nifty - val2 +Price_Difference
                                else:
                                    ATM_PRICE_N = ltp_nifty - val2
                                ##Bank Nifty
                                Price_Difference = 100
                                val2 = math.fmod(ltp_Banknifty, Price_Difference); #print('val2:',val,val2)
                                if val2 >=(Price_Difference/2) :
                                    ATM_PRICE_BN = ltp_Banknifty - val2 +Price_Difference
                                else:
                                    ATM_PRICE_BN = ltp_Banknifty - val2
                            
                                Price_Difference = 50
                                val2 = math.fmod(ltp_finnifty, Price_Difference); #print('val2:',val,val2)
                                if val2 >=(Price_Difference/2) :
                                    ATM_PRICE_FIN = ltp_finnifty - val2 +Price_Difference
                                else:
                                    ATM_PRICE_FIN = ltp_finnifty - val2

                                Price_Difference = 100
                                val2 = math.fmod(ltp_sensex, Price_Difference); #print('val2:',val,val2)
                                if val2 >=(Price_Difference/2) :
                                    ATM_PRICE_SEN = ltp_sensex - val2 +Price_Difference
                                else:
                                    ATM_PRICE_SEN = ltp_sensex - val2
                                

                                Price_Difference = 100
                                val2 = math.fmod(ltp_bankex, Price_Difference); #print('val2:',val,val2)
                                if val2 >=(Price_Difference/2) :
                                    ATM_PRICE_BAN = ltp_bankex - val2 +Price_Difference
                                else:
                                    ATM_PRICE_BAN = ltp_bankex - val2

                                Price_Difference = 100
                                val2 = math.fmod(ltp_midcpnif, Price_Difference); #print('val2:',val,val2)
                                if val2 >=(Price_Difference/2) :
                                    ATM_PRICE_MID = ltp_midcpnif - val2 +Price_Difference
                                else:
                                    ATM_PRICE_MID = ltp_midcpnif - val2
                                
                                print('BN & N &FIN &SEN &BAN & MIDCP spot and ATMs:',ltp_Banknifty,ltp_nifty,ltp_finnifty,ltp_sensex,ltp_bankex,ltp_midcpnif,ATM_PRICE_BN,ATM_PRICE_N,ATM_PRICE_FIN,ATM_PRICE_SEN,ATM_PRICE_BAN,ATM_PRICE_MID)
                                #ogfile.flush(); os.fsync(#ogfile.fileno())
                ##Getting instruments list:
                #stocks = df_nf200.Symbol.to_list()
                #stocks_list,instruments_list,future_contract_list = get_the_watchlist_master()
                stocks_df = df_instruments[(
                            ((df_instruments.name == "BANKNIFTY") & (df_instruments.expiry.isin(Three_expiries_BN)) |(df_instruments.expiry.isin(bn_monthly))) |
                            ((df_instruments.name == "NIFTY") & (df_instruments.expiry.isin(Three_expiries_NY)) | (df_instruments.expiry.isin(n_monthly))) |
                            ((df_instruments.name == "FINNIFTY") & (df_instruments.expiry.isin(Three_expiries_FIN)) | (df_instruments.expiry.isin(fin_monthly))) |
                            ((df_instruments.name == "SENSEX") & (df_instruments.expiry.isin(Three_expiries_SEN)) | (df_instruments.expiry.isin(sen_monthly))) |
                            ((df_instruments.name == "MIDCPNIFTY") & (df_instruments.expiry.isin(Three_expiries_MID)) | (df_instruments.expiry.isin(mid_monthly))) |
                            ((df_instruments.name == "BANKEX") & (df_instruments.expiry.isin(Three_expiries_BANK)))
                            
                            ) 
                            & ((df_instruments.segment == 'NFO-OPT')|(df_instruments.segment == 'BFO-OPT')|(df_instruments.segment == 'NFO-FUT')|(df_instruments.segment == 'BFO-FUT'))] #& (df_instruments.strike <= 40000) & (df_instruments.strike >= 39800)];#print('stocks_df::',stocks_df,stocks_df['instrument_token'].tolist());exit() #& ((df_instruments.expiry == Expiry_dt) | (df_instruments.expiry == MonthlyExpiry) ) (df_instruments.name == "NIFTY") |
                
                
                df_z_Token_Symbol = stocks_df[['instrument_token','tradingsymbol','expiry','strike','instrument_type','name','segment']]
                #print(stocks_df);exit()
                #df_z_Token_Symbol = stocks_df[['instrument_token','tradingsymbol']]
                df_z_Token_Symbol.rename(columns = {'instrument_token':'Stock','instrument_type':'CE_PE'}, inplace = True)
                #print(df_z_Token_Symbol.tradingsymbol.values,len(df_z_Token_Symbol.tradingsymbol.values.tolist()));exit()
                # Assuming df_instruments is your DataFrame and you have defined variables like Three_expiries_BN, ATM_PRICE_BN, etc.

            # Correctly format the conditions by ensuring each condition is enclosed in parentheses.
                stocks_df_only22 = df_instruments[
                    ((((df_instruments['name'] == "BANKNIFTY") & df_instruments['expiry'].isin(Three_expiries_BN) & 
                    ((df_instruments['strike'] >= (ATM_PRICE_BN - 2000)) & (df_instruments['instrument_type'] == "CE")) |
                    ((df_instruments['strike'] <= (ATM_PRICE_BN + 2000)) & (df_instruments['instrument_type'] == "PE")))) |
                    
                    ((df_instruments['name'] == "NIFTY") & df_instruments['expiry'].isin(Three_expiries_NY) & 
                    ((df_instruments['strike'] >= (ATM_PRICE_N - 1000)) & (df_instruments['instrument_type'] == "CE")) |
                    ((df_instruments['strike'] <= (ATM_PRICE_N + 1000)) & (df_instruments['instrument_type'] == "PE"))) |
                    
                    ((df_instruments['name'] == "FINNIFTY") & df_instruments['expiry'].isin(Three_expiries_FIN) & 
                    ((df_instruments['strike'] >= (ATM_PRICE_FIN - 1000)) & (df_instruments['instrument_type'] == "CE")) |
                    ((df_instruments['strike'] <= (ATM_PRICE_FIN + 1000)) & (df_instruments['instrument_type'] == "PE"))) |
                    
                    ((df_instruments['name'] == "SENSEX") & df_instruments['expiry'].isin(Three_expiries_SEN) & 
                    ((df_instruments['strike'] >= (ATM_PRICE_SEN - 2000)) & (df_instruments['instrument_type'] == "CE")) |
                    ((df_instruments['strike'] <= (ATM_PRICE_SEN + 2000)) & (df_instruments['instrument_type'] == "PE"))) |
                    
                    ((df_instruments['name'] == "MIDCPNIFTY") & df_instruments['expiry'].isin(Three_expiries_MID) & 
                    ((df_instruments['strike'] >= (ATM_PRICE_MID - 300)) & (df_instruments['instrument_type'] == "CE")) |
                    ((df_instruments['strike'] <= (ATM_PRICE_MID + 300)) & (df_instruments['instrument_type'] == "PE"))) |
                    
                    ((df_instruments['name'] == "BANKEX") & df_instruments['expiry'].isin(Three_expiries_BANK) & 
                    ((df_instruments['strike'] >= (ATM_PRICE_BAN - 2500)) & (df_instruments['instrument_type'] == "CE")) |
                    ((df_instruments['strike'] <= (ATM_PRICE_BAN + 2500)) & (df_instruments['instrument_type'] == "PE")))|

                    ((df_instruments['name'] == "BANKNIFTY") & df_instruments['expiry'].isin(bn_monthly))|
                    ((df_instruments['name'] == "NIFTY") & df_instruments['expiry'].isin(n_monthly))|
                    ((df_instruments['name'] == "FINNIFTY") & df_instruments['expiry'].isin(fin_monthly))|
                    ((df_instruments['name'] == "SENSEX") & df_instruments['expiry'].isin(sen_monthly))|
                    ((df_instruments['name'] == "MIDCPNIFTY") & df_instruments['expiry'].isin(mid_monthly))|
                    ((df_instruments['name'] == "BANKEX") & df_instruments['expiry'].isin(mid_monthly))


                    ) & 
                    # This last condition is applied to all selected rows above
                    ((df_instruments['segment'] == 'NFO-OPT') | (df_instruments['segment'] == 'BFO-OPT')|(df_instruments['segment'] == 'BFO-FUT')|(df_instruments['segment'] == 'NFO-FUT'))]
                

            # Make sure all the variables (like Three_expiries_BN, ATM_PRICE_BN

                
                instruments_list = [];instruments_list = stocks_df_only22['instrument_token'].tolist()
                tradingsymbol_list = stocks_df_only22['tradingsymbol'].tolist()    
                print('\nsystem Time:',datetime.datetime.now())
                # print('Live Price checking for:',stocks_df_only22['tradingsymbol'].tolist(),'\n');
                del stocks_df_only22;del stocks_df
                instruments_list = list(set(instruments_list))
                #print('Subscribed instruments_list:',instruments_list,len(instruments_list));
                #exit()
                instruments_list = [int(x) for x in instruments_list]
                instruments_list_Full = instruments_list.copy()
                df_full_all = pd.DataFrame();df_combined3 = pd.DataFrame();
                curr_dt_rank = time.strftime("%Y%m%d", time.localtime())
                Ranked_dff_name_prev = f"previous_data_{curr_dt_rank}.csv"

                ## Getting the previous last candle greeks data.
            
            else:
                if one_time_print==0:
                    one_time_print=1
        
        #ogfile.flush(); os.fsync(#ogfile.fileno())
        
        entered2 = 0
        while entered2 == 0:
            if mkt_open_time - timedelta(seconds=2) < datetime.datetime.now() < mkt_close_time + timedelta(minutes=1):
                    kws.on_ticks = on_ticks
                    kws.on_connect = on_connect
                    kws.on_reconnect = on_reconnect
                    kws.on_error = on_error
                    kws.connect(threaded=True)
                    entered2 = 1
        if True:
                curr_dt = time.strftime("%Y-%m-%d", time.localtime())
                curr_dt_rank = time.strftime("%Y%m%d", time.localtime())
                processed_tm_minute = [];Noted_len = 0
                exp_index_side_df_data=[]
                ##It is to store the First candle or 9:15 EACH strikes greeks data.
                Ranked_dff_csv = pd.DataFrame(columns=['Date','Stock','name','strike','CE_PE','LTP','Net Change','%Chg','Vol','iv','delta','gamma','rho','theta','vega','sum_vega','sum_iv','Batch','expiry','rank','Curr_RANK1_min'])
                Ranked_dff_name = 'Ranked_Index_Data_'+ str(curr_dt_rank) + '.csv'
                ##It is to store all the each minute 1 Minute consolidated expiry,index,ce_pe Greek wise data.
                dff_Ranked_Greeks_nm = 'Ranked_Greeks_Data_'+ str(curr_dt_rank) + '.csv'
                ##It is to store all the each minute 1 Minute Hist Greek wise data.
                dff_Ranked_Hist_Greeks_nm = 'Ranked_Hist_Greeks_Data_'+ str(curr_dt_rank) + '.csv'
                df_Ranked_Greeks_First_Col = 'Ranked_Greeks_Data_First_Col_'+ str(curr_dt_rank) + '.csv'
                #dff_Ranked_Hist_Greeks_nm = 'Ranked_Hist_Greeks_Data_20230504.csv'
                if not path.exists(Ranked_dff_name) :
                    Ranked_dff_csv.to_csv(Ranked_dff_name,index=False) ###LIVEEEEEEEEEEEEE
                
                else:
                    print(Ranked_dff_name,'already available')
                    Ranked_dff_csv = pd.read_csv(Ranked_dff_name)
            #print(Ranked_dff_csv)
        global df_hist_all_Data,new_df_com,df_combined22
        market_open = True;df_combined2_column_count = 0;df_hist_all_Data = pd.DataFrame()
        First_Candle_Start_tm_dt = datetime.datetime.now();new_df_com = pd.DataFrame();df_combined22 = pd.DataFrame()
        if path.exists(dff_Ranked_Hist_Greeks_nm) and path.getsize(dff_Ranked_Hist_Greeks_nm) > 0:
                print(dff_Ranked_Hist_Greeks_nm,'already available')
                try:
                        new_df_com = pd.read_csv(dff_Ranked_Hist_Greeks_nm)
                except pd.errors.EmptyDataError:
                        print(dff_Ranked_Hist_Greeks_nm, "is empty")
                        for filename in Path(os.getcwd()).glob(dff_Ranked_Hist_Greeks_nm):
                            filename.unlink()
                        new_df_com = pd.DataFrame()
        
        j=0
        i=0
        l=0
        first_time= True
        f_time=True
        starting_cells={}
        starting_cells_d={}
        f_time_d=True
        starting_cells_pro={}
        p_time=True
        result_onemin_df=pd.DataFrame(); result_threemin_df=pd.DataFrame() ; result_fivemin_df=pd.DataFrame() ; result_tenmin_df=pd.DataFrame() ;result_fiveteenmin_df=pd.DataFrame() ;result_thirtymin_df=pd.DataFrame() ;result_live_df=pd.DataFrame()
        f_name= "excel"
        if not os.path.exists(f_name):
            os.makedirs(f_name)
            print(f"The folder '{f_name}' has been created.")
            #ogfile.flush(); os.fsync(#ogfile.fileno())

        else:
                print(f"The folder '{f_name}' already exists.")
                #ogfile.flush(); os.fsync(#ogfile.fileno())
        
        i=0
        prev_vol_df=pd.DataFrame()
        prev_tick=pd.DataFrame()
        atv_df=pd.DataFrame()
        indices = ["banknifty", "nifty", "finnifty", "midcpnifty", "sensex","bankex"]
        conditions = ["increase", "decrease"]
        columns = ['name', 'expiry', 'Time', 'protrader_data', 'protrader_senti', 'retailtrader_data', 'retailtrader_senti', 'ATV_CE', 'CE_change', 'ATV_PE', 'PE_change']
        empty_df = pd.DataFrame(columns=columns)

        for index in indices:
            for condition in conditions:
                if index in ["banknifty", "nifty"]:  
                    globals()[f'initial_{index}_current_expiry_{condition}'] = pd.DataFrame()
                    globals()[f'initial_{index}_next_expiry_{condition}'] = pd.DataFrame()
                else:  
                    globals()[f'initial_{index}_current_expiry_{condition}'] = pd.DataFrame()



        def wait_until_market_opens():
            now = datetime.datetime.now()
            market_open_time = now.replace(hour=9, minute=15, second=10, microsecond=0)

            if now < market_open_time:
                print("Market is not open")
                
                sleep_time = (market_open_time - now).total_seconds()
                time.sleep(sleep_time)
            else:
                print("market is open")


        wait_until_market_opens()

        #   banks=['HDFCBANK','ICICIBANK','SBIN','AXISBANK','KOTAKBANK','INDUSINDBK']
        #   for symbols in banks:
        #     volume_df= get_historic_data(symbols, 'day', 10,'NSE')
        #     selected_row = volume_df.iloc[-2:-1]
        #     prev_vol_df = pd.concat([prev_vol_df, selected_row], ignore_index=True)


        def fetch_market_data():

            symbols = ['NIFTY BANK', 'NIFTY 50', 'NIFTY FIN SERVICE', 'NIFTY MID SELECT', 'SENSEX', 'INDIA VIX']
            market_data = pd.DataFrame(columns=['Name', 'Open', 'LTP', 'Diff'])
            for symbol in symbols:
                exchange = 'BSE' if symbol == 'SENSEX' else 'NSE'
                data = get_historic_data(symbol, 'day', 10, exchange)
                if data is not None and not data.empty:
                    last_open = data['open'].iloc[-1]
                    new_row = pd.DataFrame({'Name': [symbol], 'Open': [last_open]})
                    market_data = pd.concat([market_data, new_row], ignore_index=True)

            def get_ltp():
                try:
                    symbols_quote = ["NSE:NIFTY BANK", "NSE:NIFTY 50", "NSE:NIFTY FIN SERVICE", "NSE:NIFTY MID SELECT", "BSE:SENSEX", "NSE:INDIA VIX"]
                    data = kite.quote(symbols_quote)
                    ltps = {symbol.split(':')[1]: data[symbol]['last_price'] for symbol in symbols_quote}
                    print("LTPs collected successfully.")
                    return ltps
                except Exception as e:
                    print("Error fetching LTPs:", e)
                    return {}

            ltp_data = get_ltp()
            market_data['LTP'] = market_data['Name'].apply(lambda x: ltp_data.get(x, None))
            market_data['Diff'] = market_data['LTP'] - market_data['Open']

            return market_data

        def aggregate_data(hdf):
            agg_dict = {
                'Date': 'first',
                'name': 'first',
                'strike': 'first',
                'CE_PE': 'first',
                'Net Change': 'first',
                '%Chg': 'first',
                'Vol': 'first',
                'oi': 'first',
                'Totl Bid Qty': 'first',
                'Totl Ask Qty': 'first',
                'iv': 'first',
                'expiry': 'first',
                'LTP': 'mean',   
                'theta': 'mean', 
                'vega': 'mean', 
                'delta': 'mean',  
                'gamma': 'first', 
                'rho': 'first',    
                'sum_vega': 'first',  
                'sum_iv': 'first',    
                'Batch': 'first'      
            }

            aggregated_df = hdf.groupby('Stock').agg(agg_dict).reset_index()
            columns_order = [
                'Date', 'Stock', 'name', 'strike', 'CE_PE', 'LTP', 'Net Change', '%Chg', 'Vol',
                'oi', 'Totl Bid Qty', 'Totl Ask Qty', 'iv', 'delta', 'gamma', 'rho', 'theta',
                'vega', 'expiry', 'sum_vega', 'sum_iv', 'Batch'
            ]

            return aggregated_df[columns_order]
        
        
        def identify_starting_cells(file_names, sheet_info):
            app = xw.App(visible=False) if xw.apps.count == 0 else xw.apps.active
            starting_cells_info = {}

            for file_name, sheets in sheet_info.items():
                full_path = os.path.join(os.getcwd(), file_name)
                try:
                    wb = app.books.open(full_path)
                    for sheet_name, columns in sheets.items():
                        sheet_key = f"{file_name} - {sheet_name}"
                        starting_cells = {}
                        # Load data without headers to determine last non-empty row in specified columns
                        hdf = pd.read_excel(full_path, sheet_name=sheet_name, header=None, engine='openpyxl')
                        
                        for mode, col_index in columns.items():
                            try:
                                last_row = hdf.iloc[:, col_index - 1].dropna().index[-1] if not hdf.iloc[:, col_index - 1].dropna().empty else -1
                            except IndexError:
                                last_row = -1

                            column_letter = get_column_letter(col_index)
                            if last_row == -1:
                                starting_cells[mode] = f'{column_letter}1'
                            else:
                                starting_cells[mode] = f'{column_letter}{last_row + 2}'

                        starting_cells_info[sheet_key] = starting_cells

                    wb.close()
                except Exception as e:
                    print(f"Failed to process {file_name} due to: {e}")

            if len(app.books) == 0:
                app.quit()
            
            return starting_cells_info
        

        sheet_info = {
            'NF_Pro_Trading.xlsx': {'VEGA & Theta_Chng_CW': {'current': 1, 'next': 17}, 'VEGA & Theta_Chng_NW': {'current': 1, 'next': 17}},
            'BNF_Pro_Trading.xlsx': {'VEGA & Theta_Chng_CW': {'current': 1, 'next': 17}, 'VEGA & Theta_Chng_NW': {'current': 1, 'next': 17}},
            'FIN_Pro_Trading.xlsx': {'VEGA & Theta_Chng_CW': {'current': 1, 'next': 17}},
            'Midcp_Pro_Trading.xlsx': {'VEGA & Theta_Chng_CW': {'current': 1, 'next': 17}},
            'Sen_Pro_Trading.xlsx': {'VEGA & Theta_Chng_CW': {'current': 1, 'next': 17}},
            'BAN_Pro_Trading.xlsx': {'VEGA & Theta_Chng_CW': {'current': 1, 'next': 17}},
            "Sen_theta_delta.xlsx": {'main': {'current': 1, 'next': 21}},
            "BAN_theta_delta.xlsx": {'main': {'current': 1, 'next': 21}},
            "NF_theta_delta.xlsx": {'main': {'current': 1, 'next': 21}},
            "Midcp_theta_delta.xlsx": {'main': {'current': 1, 'next': 21}},
            "FIN_theta_delta.xlsx": {'main': {'current': 1, 'next': 21}},
            "BNF_theta_delta.xlsx": {'main': {'current': 1, 'next': 21}}}




  
    # Example call to the function
        starting_cells_info = identify_starting_cells(['NF_Pro_Trading.xlsx', 'BNF_Pro_Trading.xlsx', 'FIN_Pro_Trading.xlsx', 'Midcp_Pro_Trading.xlsx', 'Sen_Pro_Trading.xlsx','BAN_Pro_Trading.xlsx'], sheet_info)
        print("starting cells are",starting_cells_info)

        bn_ce_f=pd.DataFrame();n_ce_f=pd.DataFrame();bn_ne_f=pd.DataFrame();ban_ne_f=pd.DataFrame();n_ne_f=pd.DataFrame();f_ne_f=pd.DataFrame();m_ne_f=pd.DataFrame();s_ne_f=pd.DataFrame();b_ne_f=pd.DataFrame()
        bnf_new_df = pd.DataFrame()
        nf_new_df = pd.DataFrame()
        fin_new_df = pd.DataFrame()
        mid_new_df = pd.DataFrame()
        sen_new_df = pd.DataFrame()
        ban_new_df=pd.DataFrame()
        while market_open:
            try:
        
                if mkt_open_time - timedelta(seconds=2) < datetime.datetime.now() < mkt_close_time + timedelta(minutes=1):
                    if True:
                        tm = time.time()
                        min_remain = math.fmod(tm, 60)
                        
                            #   prev_vol_df=pd.concat([prev_vol_df,volume_df], ignore_index=True, axis=0)

                        

                        if True: #not (min_remain < 50 and min_remain > 10):
                                if len(df_tick_dump.index) >0 : ## FOR TESTING keep it
                                    df_tick_dump.to_csv("sample.csv")
                                    now = datetime.datetime.now()
                                    start_time = now.replace(hour=8, minute=0, second=0, microsecond=0)
                                    end_time = now.replace(hour=9, minute=15, second=5, microsecond=0)

                                    if start_time <= now <= end_time:

                                            sleep_seconds = (end_time - now).total_seconds()
                                            print(f"Going to wait till 09:15:19 for collecting ltp and greeks data")
                                            time.sleep(sleep_seconds)
                                    else:
                                        print("Current time has crossed 09:15:19. No sleep needed.") 
                                    
                                    df_v = process_df_dump(df_tick_dump.copy())  ## FOR TESTING keep it
                                    df_v.to_csv("sv.csv")
                                    

                                    if i==0:

                                        if new_day:
                                            start_time = now.replace(hour=8, minute=0, second=0, microsecond=0)
                                            end_time = now.replace(hour=9, minute=15, second=45, microsecond=0)
                                            s_t=now.replace(hour=9, minute=15, second=45, microsecond=0)
                                            print("going to take average data till 09:15:45")
                                            now = datetime.datetime.now()
                                            if now<s_t:

                                                while start_time <= now <= end_time:
                                                    now = datetime.datetime.now()                            
                                                    first_df=df_v.copy()
                                                    df_v = process_df_dump(df_tick_dump.copy())

                                                first_df.to_csv("first df.csv")
                                                print("averaging over")
                                                first_df=aggregate_data(first_df)
                                                first_df.to_csv("constant_data.csv")
                                            else:
                                                first_df=df_v.copy()
                                                first_df.to_csv("constant_data.csv")

                                        else:
                                            first_df=pd.read_csv("constant_data.csv")
                                        
                                        
                                    if len(df_v.index)>0:
                                        banks_df=df_v.copy()
                                        if 'iv' in df_v and 'delta' in df_v and 'theta' in df_v and 'vega' in df_v and  'Date' in df_v :
                                                Min_Start = datetime.datetime.today()
                                                Min_Start = Min_Start.replace(second=0, microsecond=0) #- datetime.timedelta(minutes=2344) ## FOR TESTING #hour=0, minute=0, #- datetime.timedelta(days=2) #2266 23:00
                                                Min_End = Min_Start + datetime.timedelta(seconds=60)
                                                print('Started New Refresh!! Min_Start,Min_End:',Min_Start,Min_End,'Current Time:',datetime.datetime.now()); #exit() 
                                                #.flush(); os.fsync(#.fileno())
                                                tm = Min_Start.strftime('%H%M')
                                                print("time is",tm)
                                                #   if i==0:                           
                                                #     first_time=True
                                                #     print("first time is true")
                                                #   else:
                                                #     time_checker.update_time(tm)

                                                    
                                                #.flush(); os.fsync(#.fileno())
                                                timeframe_checks = check_timeframes(tm)                                                    
                                                new_df_columns = ['Index', 'expiry', 'call_put', 'Greek_type']
                                                new_df = pd.DataFrame(columns=new_df_columns)
                                                
                                                if True:
                                                                    ltp_nifty,ltp_Banknifty,ltp_finnifty,ltp_sensex,ltp_bankex,ltp_midcpnif = get_N_BN_spot()
                                                                    ##Nifty
                                                                    Price_Difference = 50
                                                                    val2 = math.fmod(ltp_nifty, Price_Difference); #print('val2:',val,val2)
                                                                    if val2 >=(Price_Difference/2) :
                                                                        ATM_PRICE_N = ltp_nifty - val2 +Price_Difference
                                                                    else:
                                                                        ATM_PRICE_N = ltp_nifty - val2
                                                                    ##Bank Nifty
                                                                    Price_Difference = 100
                                                                    val2 = math.fmod(ltp_Banknifty, Price_Difference); #print('val2:',val,val2)
                                                                    if val2 >=(Price_Difference/2) :
                                                                        ATM_PRICE_BN = ltp_Banknifty - val2 +Price_Difference
                                                                    else:
                                                                        ATM_PRICE_BN = ltp_Banknifty - val2
                                                                    #print('BN & N spot and ATMs:',ltp_Banknifty,ltp_nifty,ATM_PRICE_BN,ATM_PRICE_N)
                                                                    Price_Difference = 50
                                                                    val2 = math.fmod(ltp_finnifty, Price_Difference); #print('val2:',val,val2)
                                                                    if val2 >=(Price_Difference/2) :
                                                                        ATM_PRICE_FIN = ltp_finnifty - val2 +Price_Difference
                                                                    else:
                                                                        ATM_PRICE_FIN = ltp_finnifty - val2

                                                                    Price_Difference = 100
                                                                    val2 = math.fmod(ltp_sensex, Price_Difference); #print('val2:',val,val2)
                                                                    if val2 >=(Price_Difference/2) :
                                                                        ATM_PRICE_SEN = ltp_sensex - val2 +Price_Difference
                                                                    else:
                                                                        ATM_PRICE_SEN = ltp_sensex - val2
                                                                    
                                                                    Price_Difference = 100
                                                                    val2 = math.fmod(ltp_bankex, Price_Difference); #print('val2:',val,val2)
                                                                    if val2 >=(Price_Difference/2) :
                                                                        ATM_PRICE_BAN = ltp_bankex - val2 +Price_Difference
                                                                    else:
                                                                        ATM_PRICE_BAN = ltp_bankex - val2

                                                                    Price_Difference = 100
                                                                    val2 = math.fmod(ltp_midcpnif, Price_Difference); #print('val2:',val,val2)
                                                                    if val2 >=(Price_Difference/2) :
                                                                        ATM_PRICE_MID = ltp_midcpnif - val2 +Price_Difference
                                                                    else:
                                                                        ATM_PRICE_MID = ltp_midcpnif - val2
                                                                    
                                                                    print('BN & N &FIN &SEN &BAN & MIDCP spot and ATMs:',ltp_Banknifty,ltp_nifty,ltp_finnifty,ltp_sensex,ltp_bankex,ltp_midcpnif,ATM_PRICE_BN,ATM_PRICE_N,ATM_PRICE_FIN,ATM_PRICE_SEN,ATM_PRICE_BAN,ATM_PRICE_MID)
                                                                    #.flush(); os.fsync(#.fileno())

                                                n=0
                                                result_df=pd.DataFrame()
                                                pro_df=pd.DataFrame()
                                                pro_result_df=pd.DataFrame()
                                                new_pro_df=pd.DataFrame()
                                                new_result_df=pd.DataFrame()
                                                for it,exp_list in enumerate([Three_expiries_BN,Three_expiries_NY,Three_expiries_FIN,Three_expiries_MID,Three_expiries_SEN,Three_expiries_BANK]):
                                                    for exp in exp_list:
                                                        index = ['BANKNIFTY', 'NIFTY','FINNIFTY','MIDCPNIFTY','SENSEX','BANKEX'][it]
                                                        for kg,CE_PE in enumerate(['CE','PE']):
                                                            for method in ["normal","notnormal"]:

                                                                attempt = 0  
                                                                df_v_new = pd.DataFrame()
                                                                max_retries=5  
                                                                while df_v_new.empty and attempt < max_retries:
                                                                    # print(f"trying for {index},{CE_PE},{exp}")
                                                                    #.flush(); os.fsync(#.fileno())
                                                                    attempt+=1
                                                                    
                                                                    df_v_new = df_v[(df_v.iv != 0) & (df_v.delta != 0) & (df_v.gamma != 0) & (df_v.rho != 0) & (df_v.theta != 0) & (df_v.vega != 0)] ##working valid value
                                                                    df_v_new['Date'] = pd.to_datetime(df_v_new['Date'], format='%d-%m-%Y %H:%M')                                                             
                                                                    df_v_new = df_v_new[(df_v_new.Date >= Min_Start) & (df_v_new.Date < Min_End)] ##working                              
                                                                    df_v_lat=df_v_new.copy()
                                                                    #print(len(df_v_new.index));exit()
                                                                    #df_v_new = df_v_new[(df_v_new.Date >= '2023-03-23 09:15:00') & (df_v_new.Date < '2023-03-23 09:16:00')] ##working
                                                                    atm_bn = ATM_PRICE_BN
                                                                    atm_n = ATM_PRICE_N
                                                                    atm_fin=ATM_PRICE_FIN
                                                                    atm_midcp=ATM_PRICE_MID
                                                                    atm_sen=ATM_PRICE_SEN
                                                                    atm_ban=ATM_PRICE_BAN                             
                                                                    first_value_expiry = df_v_new['expiry'].iloc[0]                             
                                                                    exp=str(exp)
                                                                    df_v_new['expiry'] = df_v_new['expiry'].astype(str)  
                                                                    bn_ce=False;bn_pe=False;n_ce=False;n_pe=False;f_ce=False;f_pe=False;mid_ce=False;mid_pe=False;sen_ce=False;sen_pe=False                       
                                                                    if CE_PE == 'CE' and index == 'BANKNIFTY':
                                                                        bn_ce=True
                                                                        if index in actual_expiry_index and method=='notnormal':
                                                                            strikes_list=calculate_strikes(atm_bn,100,'CE',"new")
                                                                        else:
                                                                            strikes_list=calculate_strikes(atm_bn,100,'CE',method)

                                                                        df_v_new = df_v_new[(df_v_new['expiry'] == exp) & (df_v_new['name'] == index) & (df_v_new['CE_PE'] == CE_PE) & df_v_new['strike'].isin(strikes_list)]
                                                                        df_v_lat = df_v_lat[(df_v_lat['expiry'] == exp) & (df_v_lat['name'] == index) & (df_v_lat['CE_PE'] == CE_PE)]
                                                                        
                                                                    elif CE_PE == 'PE' and index == 'BANKNIFTY':
                                                                        bn_pe=True
                                                                        
                                                                        if index in actual_expiry_index and method=='notnormal':
                                                                            strikes_list=calculate_strikes(atm_bn,100,'PE',"new")
                                                                        else:
                                                                            strikes_list=calculate_strikes(atm_bn,100,'PE',method)
                                                                        df_v_new = df_v_new[(df_v_new['expiry'] == exp) & (df_v_new['name'] == index) & (df_v_new['CE_PE'] == CE_PE) & df_v_new['strike'].isin(strikes_list)]
                                                                        df_v_lat = df_v_lat[(df_v_lat['expiry'] == exp) & (df_v_lat['name'] == index) & (df_v_lat['CE_PE'] == CE_PE)]

                                                                    elif CE_PE == 'CE' and index == 'NIFTY':
                                                                        n_ce=True
                                                                        if index in actual_expiry_index and method=='notnormal':
                                                                            strikes_list=calculate_strikes(atm_n,50,'CE',"new")
                                                                        else:
                                                                            strikes_list=calculate_strikes(atm_n,50,'CE',method)
                                                                        
                                                                        df_v_new = df_v_new[(df_v_new['expiry'] == exp) & (df_v_new['name'] == index) & (df_v_new['CE_PE'] == CE_PE) & df_v_new['strike'].isin(strikes_list)]
                                                                        df_v_lat = df_v_lat[(df_v_lat['expiry'] == exp) & (df_v_lat['name'] == index) & (df_v_lat['CE_PE'] == CE_PE)]

                                                                    elif CE_PE == 'PE' and index == 'NIFTY':
                                                                        n_pe=True
                                                                        
                                                                        if index in actual_expiry_index and method=='notnormal':
                                                                            strikes_list=calculate_strikes(atm_n,50,'PE',"new")
                                                                        else:
                                                                            strikes_list=calculate_strikes(atm_n,50,'PE',method)
                                                                        df_v_new = df_v_new[(df_v_new['expiry'] == exp) & (df_v_new['name'] == index) & (df_v_new['CE_PE'] == CE_PE) & df_v_new['strike'].isin(strikes_list)]
                                                                        df_v_lat = df_v_lat[(df_v_lat['expiry'] == exp) & (df_v_lat['name'] == index) & (df_v_lat['CE_PE'] == CE_PE)]
                                                                        
                                                                    elif CE_PE == 'CE' and index == 'FINNIFTY':
                                                                        f_ce=True
                                                                        
                                                                        if index in actual_expiry_index and method=='notnormal':
                                                                            strikes_list=calculate_strikes(atm_fin,50,'CE',"new")
                                                                        else:
                                                                            strikes_list=calculate_strikes(atm_fin,50,'CE',method)

                                                                        df_v_new = df_v_new[(df_v_new['expiry'] == exp) & (df_v_new['name'] == index) & (df_v_new['CE_PE'] == CE_PE) & df_v_new['strike'].isin(strikes_list)]
                                                                        df_v_lat = df_v_lat[(df_v_lat['expiry'] == exp) & (df_v_lat['name'] == index) & (df_v_lat['CE_PE'] == CE_PE)]
                                                                        
                                                                    elif CE_PE == 'PE' and index == 'FINNIFTY':
                                                                        f_pe=True
                                                                        if index in actual_expiry_index and method=='notnormal':
                                                                            strikes_list=calculate_strikes(atm_fin,50,'PE',"new")
                                                                        else:
                                                                            strikes_list=calculate_strikes(atm_fin,50,'PE',method)
                                                                        
                                                                        df_v_new = df_v_new[(df_v_new['expiry'] == exp) & (df_v_new['name'] == index) & (df_v_new['CE_PE'] == CE_PE) & df_v_new['strike'].isin(strikes_list)]
                                                                        df_v_lat = df_v_lat[(df_v_lat['expiry'] == exp) & (df_v_lat['name'] == index) & (df_v_lat['CE_PE'] == CE_PE)]

                                                                    elif CE_PE == 'CE' and index == 'MIDCPNIFTY':
                                                                        mid_ce=True
                                                                        if index in actual_expiry_index and method=='notnormal':
                                                                            strikes_list=calculate_strikes(atm_midcp,25,'CE',"new")
                                                                        else:
                                                                            strikes_list=calculate_strikes(atm_midcp,25,'CE',method)
                                                                        
                                                                        df_v_new = df_v_new[(df_v_new['expiry'] == exp) & (df_v_new['name'] == index) & (df_v_new['CE_PE'] == CE_PE) & df_v_new['strike'].isin(strikes_list)]
                                                                        df_v_lat = df_v_lat[(df_v_lat['expiry'] == exp) & (df_v_lat['name'] == index) & (df_v_lat['CE_PE'] == CE_PE)]

                                                                    elif CE_PE == 'PE' and index == 'MIDCPNIFTY':
                                                                        mid_pe=True
                                                                        if index in actual_expiry_index and method=='notnormal':
                                                                            strikes_list=calculate_strikes(atm_midcp,25,'PE',"new")
                                                                        else:
                                                                            strikes_list=calculate_strikes(atm_midcp,25,'PE',method)
                                                                        
                                                                        df_v_new = df_v_new[(df_v_new['expiry'] == exp) & (df_v_new['name'] == index) & (df_v_new['CE_PE'] == CE_PE) & df_v_new['strike'].isin(strikes_list)]
                                                                        df_v_lat = df_v_lat[(df_v_lat['expiry'] == exp) & (df_v_lat['name'] == index) & (df_v_lat['CE_PE'] == CE_PE)]

                                                                    elif CE_PE == 'CE' and index == 'SENSEX':
                                                                        sen_ce=True
                                                                        if index in actual_expiry_index and method=='notnormal':
                                                                            strikes_list=calculate_strikes(atm_sen,100,'CE',"new")
                                                                        else:
                                                                            strikes_list=calculate_strikes(atm_sen,100,'CE',method)
                                                                        
                                                                        df_v_new = df_v_new[(df_v_new['expiry'] == exp) & (df_v_new['name'] == index) & (df_v_new['CE_PE'] == CE_PE) & df_v_new['strike'].isin(strikes_list)]
                                                                        df_v_lat = df_v_lat[(df_v_lat['expiry'] == exp) & (df_v_lat['name'] == index) & (df_v_lat['CE_PE'] == CE_PE)]

                                                                    elif CE_PE == 'PE' and index == 'SENSEX':
                                                                        sen_pe=True
                                                                        if index in actual_expiry_index and method=='notnormal':
                                                                            strikes_list=calculate_strikes(atm_sen,100,'PE',"new")
                                                                        else:
                                                                            strikes_list=calculate_strikes(atm_sen,100,'PE',method)
                                                                        
                                                                        df_v_new = df_v_new[(df_v_new['expiry'] == exp) & (df_v_new['name'] == index) & (df_v_new['CE_PE'] == CE_PE) & df_v_new['strike'].isin(strikes_list)]
                                                                        df_v_lat = df_v_lat[(df_v_lat['expiry'] == exp) & (df_v_lat['name'] == index) & (df_v_lat['CE_PE'] == CE_PE)]

                                                                    elif CE_PE == 'CE' and index == 'BANKEX':
                                                                        ban_ce=True
                                                                        if index in actual_expiry_index and method=='notnormal':
                                                                            strikes_list=calculate_strikes(atm_ban,100,'CE',"new")

                                                                        else:
                                                                            strikes_list=calculate_strikes(atm_ban,100,'CE',method)

                                                                            df_v_new = df_v_new[(df_v_new['expiry'] == exp) & (df_v_new['name'] == index) & (df_v_new['CE_PE'] == CE_PE) & df_v_new['strike'].isin(strikes_list)]
                                                                            df_v_lat = df_v_lat[(df_v_lat['expiry'] == exp) & (df_v_lat['name'] == index) & (df_v_lat['CE_PE'] == CE_PE)]

                                                                    elif CE_PE == 'PE' and index == 'BANKEX':
                                                                        ban_pe=True
                                                                        if index in actual_expiry_index and method=='notnormal':
                                                                            strikes_list=calculate_strikes(atm_ban,100,'PE',"new")

                                                                        else:
                                                                            strikes_list=calculate_strikes(atm_ban,100,'PE',method)
                                                                            
                                                                            df_v_new = df_v_new[(df_v_new['expiry'] == exp) & (df_v_new['name'] == index) & (df_v_new['CE_PE'] == CE_PE) & df_v_new['strike'].isin(strikes_list)]
                                                                            df_v_lat = df_v_lat[(df_v_lat['expiry'] == exp) & (df_v_lat['name'] == index) & (df_v_lat['CE_PE'] == CE_PE)]

                                                                    if len(df_v_new.index) >0 :                
                                                                            df_v_new['rank'] = df_v_new.sort_values(by=['expiry', 'name', 'CE_PE', 'strike', 'Date'], ascending=False).groupby(['expiry', 'name', 'CE_PE', 'strike']).cumcount(ascending=True) + 1
                                                                            df_v_lat['rank'] = df_v_lat.sort_values(by=['expiry', 'name', 'CE_PE', 'strike', 'Date'], ascending=False).groupby(['expiry', 'name', 'CE_PE', 'strike']).cumcount(ascending=True) + 1
                                                                            df_v_new = df_v_new[(df_v_new['rank'] == 1) ]
                                                                            df_v_lat = df_v_lat[(df_v_lat['rank'] == 1) ]
                                                                            df_v_new.sort_values(by=['expiry','Stock'], ascending=True,inplace=True)
                                                                            df_v_lat.sort_values(by=['expiry','Stock'], ascending=True,inplace=True)
                                                                            df_v_new['Curr_RANK1_min'] = Min_Start
                                                                            df_v_lat['Curr_RANK1_min'] = Min_Start
                                                                            if method=='normal': 
                                                                                result_df = pd.concat([result_df, df_v_new], ignore_index=True)
                                                                                pro_result_df= pd.concat([pro_result_df, df_v_lat], ignore_index=True)
                                                                            else:
                                                                                new_result_df= pd.concat([new_result_df, df_v_new], ignore_index=True)
                                                                        
                                                index = ['BANKNIFTY', 'NIFTY','FINNIFTY','MIDCPNIFTY','SENSEX','BANKEX']
                                                future_df=pd.DataFrame()
                                                for j in index:
                                                    df_n=df_v.copy()
                                                    df_n['rank'] = df_n.sort_values(by=['expiry', 'name', 'CE_PE', 'strike', 'Date'], ascending=False).groupby(['expiry', 'name', 'CE_PE', 'strike']).cumcount(ascending=True) + 1
                                                    df_n = df_n[(df_n['rank'] == 1) ]
                                                    df_n = df_n[(df_n['name'] == j) & (df_n['CE_PE'] == 'FUT')]
                                                    future_df = pd.concat([future_df, df_n], ignore_index=True)



                                                atm_values = {'BANKNIFTY': atm_bn, 'NIFTY': atm_n, 'FINNIFTY': atm_fin, 'MIDCPNIFTY': atm_midcp, 'SENSEX': atm_sen,'BANKEX':atm_ban}
                                                print("atm values are",atm_values)
                                                print("completed all")   
                                                result_df.to_csv("ggg.csv")
                                                # Remove duplicates and keep the first occurrence
                                                pro_result_df = pro_result_df.drop_duplicates(keep='first')

                                                pro_result_df.to_csv("proooooo.csv")
                                                
                                                new_result_df.to_csv("ggg2.csv")                 
                                                dfs_dict = process_df_to_dfs(result_df)
                                                print("first step over")
                                                try:
                                                    dfs_dict= add_greeks_to_dfs(dfs_dict, first_df)
                                                except Exception as e:
                                                    print("error in neww,",e)
                                                print("2nd step over")
                                                try:
                                                    dfs_dict= add_change_columns(dfs_dict)  
                                                except Exception as e:
                                                    print("error in new addition,",e) 
                                                print("3rd step over")                       
                                                (banknifty_ce_currentexp, banknifty_ce_nextexp, banknifty_pe_currentexp, banknifty_pe_nextexp, nifty_ce_currentexp, nifty_ce_nextexp, nifty_pe_currentexp, nifty_pe_nextexp, finnifty_ce_currentexp, finnifty_pe_currentexp, midcpnifty_ce_currentexp, midcpnifty_pe_currentexp, sensex_ce_currentexp, sensex_pe_currentexp, bankex_ce_currentexp, bankex_pe_currentexp) = (dfs_dict['banknifty_ce_currentexp'], dfs_dict['banknifty_ce_nextexp'], dfs_dict['banknifty_pe_currentexp'], dfs_dict['banknifty_pe_nextexp'], dfs_dict['nifty_ce_currentexp'], dfs_dict['nifty_ce_nextexp'], dfs_dict['nifty_pe_currentexp'], dfs_dict['nifty_pe_nextexp'], dfs_dict['finnifty_ce_currentexp'], dfs_dict['finnifty_pe_currentexp'], dfs_dict['midcpnifty_ce_currentexp'], dfs_dict['midcpnifty_pe_currentexp'], dfs_dict['sensex_ce_currentexp'], dfs_dict['sensex_pe_currentexp'], dfs_dict['bankex_ce_currentexp'], dfs_dict['bankex_pe_currentexp'])
                                                banknifty_ce_currentexp.to_csv("newcecurrent.csv")
                                                banknifty_pe_currentexp.to_csv("newpecurrent.csv")
                                                banknifty_ce_nextexp.to_csv("newcenext.csv")
                                                banknifty_pe_nextexp.to_csv("newpenext.csv")
                                                print("3rd step over")

                                                try:
                                                    merged_opt_chain = merge_all_option_chains(dfs_dict)
                                                except Exception as e:
                                                    print("error in merging")


                                                print("entering into process opt chain")                       
                                                merged_opt_chain=process_option_chains(merged_opt_chain)
                                                print("exit")
                                                keys = [
                                                        'banknifty_currentexp_opt_chain', 'banknifty_nextexp_opt_chain',
                                                        'nifty_currentexp_opt_chain', 'nifty_nextexp_opt_chain',
                                                        'finnifty_currentexp_opt_chain', 
                                                        'midcpnifty_currentexp_opt_chain', 
                                                        'sensex_currentexp_opt_chain','bankex_currentexp_opt_chain'
                                                    ]
                                                    # Assuming merged_dfs is your dictionary of merged DataFrames for option chains
                                                (banknifty_currentexp_opt_chain, banknifty_nextexp_opt_chain,
                                                    nifty_currentexp_opt_chain, nifty_nextexp_opt_chain,
                                                    finnifty_currentexp_opt_chain,
                                                    midcpnifty_currentexp_opt_chain, 
                                                    sensex_currentexp_opt_chain,bankex_currentexp_opt_chain) = (merged_opt_chain.get(key) for key in keys)
                                                banknifty_currentexp_opt_chain.to_csv("bnf_optchain.csv")
                                                #   banknifty_currentexp_opt_chain = trim_dataframe(banknifty_currentexp_opt_chain)
                                                #   banknifty_nextexp_opt_chain = trim_dataframe(banknifty_nextexp_opt_chain)
                                                #   nifty_currentexp_opt_chain = trim_dataframe(nifty_currentexp_opt_chain)
                                                #   nifty_nextexp_opt_chain = trim_dataframe(nifty_nextexp_opt_chain)
                                                #   finnifty_currentexp_opt_chain = trim_dataframe(finnifty_currentexp_opt_chain)
                                                #   midcpnifty_currentexp_opt_chain = trim_dataframe(midcpnifty_currentexp_opt_chain)
                                                #   sensex_currentexp_opt_chain = trim_dataframe(sensex_currentexp_opt_chain)
                                                #   bankex_currentexp_opt_chain = trim_dataframe(bankex_currentexp_opt_chain)
                                                #   update_option_chain(banknifty_currentexp_opt_chain,'BANKNIFTY', 'BNF_ProTrading_CW',expiry_today)
                                                #   #update_option_chain(banknifty_nextexp_opt_chain,'BANKNIFTY', 'BNF_ProTrading_NW',expiry_today)
                                                #   update_option_chain(nifty_currentexp_opt_chain,'NIFTY', 'NF_ProTrading_CW',expiry_today)
                                                #   #update_option_chain(nifty_nextexp_opt_chain,'NIFTY', 'NF_ProTrading_NW',expiry_today)
                                                #   update_option_chain(finnifty_currentexp_opt_chain,'FINNIFTY', 'Fin_ProTrading_CW',expiry_today)
                                                #   update_option_chain(midcpnifty_currentexp_opt_chain,'MIDCPNIFTY', 'Mid_ProTrading_CW',expiry_today)
                                                #   update_option_chain(sensex_currentexp_opt_chain,'SENSEX', 'Sen_ProTrading_CW',expiry_today)
                                                #   update_option_chain(bankex_currentexp_opt_chain,'BANKEX', 'BAN_ProTrading_CW',expiry_today)
                                                #   banknifty_currentexp_opt_chain.to_csv("bnk.csv")
                                                if i==0:
                                                    new_tm=tm
                                                if new_tm==tm:                            
                                                    first_time=True
                                                    print("first time is true")
                                                elif new_tm!=tm:
                                                    first_time=False
                                                    print("first time is false")
                                                    new_tm=tm
                                                
                                                if first_time:
                                                    
                                                    last_row_values = banknifty_currentexp_opt_chain.iloc[-1].dropna()                          
                                                    bn_ce = pd.DataFrame([last_row_values], columns=last_row_values.index) 
                                                    bn_ce_f = pd.concat([bn_ce_f , bn_ce], ignore_index=True)
                                                    bn_ce_f.to_csv("bn_ce_fresultant.csv")
                                                    print("bn_ce_fresultant.csv")
                                                    last_row_values = banknifty_nextexp_opt_chain.iloc[-1].dropna()                          
                                                    bn_ne = pd.DataFrame([last_row_values], columns=last_row_values.index) 
                                                    bn_ne_f = pd.concat([bn_ne_f , bn_ne], ignore_index=True)
                                                    bn_ne_f.to_csv("bn_ne_fresultant.csv")
                                                    print("bn_ne_fresultant.csv")
                                                    last_row_values = nifty_currentexp_opt_chain.iloc[-1].dropna()                          
                                                    n_ce = pd.DataFrame([last_row_values], columns=last_row_values.index) 
                                                    n_ce_f = pd.concat([n_ce_f , n_ce], ignore_index=True)
                                                    n_ce_f.to_csv("n_ce_fresultant.csv")
                                                    print("n_ce_fresultant.csv")
                                                    last_row_values = nifty_nextexp_opt_chain.iloc[-1].dropna()                          
                                                    n_ne = pd.DataFrame([last_row_values], columns=last_row_values.index) 
                                                    n_ne_f = pd.concat([n_ne_f , n_ne], ignore_index=True)
                                                    n_ne_f.to_csv("n_ne_fresultant.csv")
                                                    print("n_ne_fresultant.csv")
                                                    last_row_values = finnifty_currentexp_opt_chain.iloc[-1].dropna()                          
                                                    f_ne = pd.DataFrame([last_row_values], columns=last_row_values.index) 
                                                    f_ne_f = pd.concat([f_ne_f , f_ne], ignore_index=True)
                                                    f_ne_f.to_csv("f_ne_fresultant.csv")
                                                    print("f_ne_fresultant.csv")
                                                    last_row_values = midcpnifty_currentexp_opt_chain.iloc[-1].dropna()                          
                                                    m_ne = pd.DataFrame([last_row_values], columns=last_row_values.index) 
                                                    m_ne_f = pd.concat([m_ne_f , m_ne], ignore_index=True)
                                                    m_ne_f.to_csv("m_ne_fresultant.csv")
                                                    print("m_ne_fresultant.csv")
                                                    last_row_values = sensex_currentexp_opt_chain.iloc[-1].dropna()                          
                                                    s_ne = pd.DataFrame([last_row_values], columns=last_row_values.index) 
                                                    s_ne_f = pd.concat([s_ne_f , s_ne], ignore_index=True)
                                                    s_ne_f.to_csv("s_ne_fresultant.csv")
                                                    print("s_ne_fresultant.csv")
                                                    last_row_values = bankex_currentexp_opt_chain.iloc[-1].dropna()                          
                                                    ban_ne = pd.DataFrame([last_row_values], columns=last_row_values.index) 
                                                    ban_ne_f = pd.concat([ban_ne_f , ban_ne], ignore_index=True)
                                                    ban_ne_f.to_csv("ban_ne_fresultant.csv")
                                                else:
                                                    print("going to calculate mean")
                                                    bn_ce_f.to_csv('bn_ce_f.csv', index=False)
                                                    bn_ne_f.to_csv('bn_ne_f.csv', index=False)
                                                    n_ce_f.to_csv('n_ce_f.csv', index=False)
                                                    n_ne_f.to_csv('n_ne_f.csv', index=False)
                                                    f_ne_f.to_csv('f_ne_f.csv', index=False)
                                                    m_ne_f.to_csv('m_ne_f.csv', index=False)
                                                    s_ne_f.to_csv('s_ne_f.csv', index=False)
                                                    ban_ne_f.to_csv('ban_ne_f.csv', index=False)

                                                    print("All DataFrames have been saved to CSV files.")
                                                    try:
                                                        bn_ce_f_i = pd.DataFrame([bn_ce_f.mean()]); bn_ne_f_i = pd.DataFrame([bn_ne_f.mean()]); n_ce_f_i = pd.DataFrame([n_ce_f.mean()]); n_ne_f_i = pd.DataFrame([n_ne_f.mean()]); f_ne_f_i = pd.DataFrame([f_ne_f.mean()]); m_ne_f_i = pd.DataFrame([m_ne_f.mean()]); s_ne_f_i = pd.DataFrame([s_ne_f.mean()]); ban_ne_f_i = pd.DataFrame([ban_ne_f.mean()])
                                                    except Exception as e:
                                                        print("error is",e)
                                                        exit()
                                                    print("going to reorder")
                                                    columns_to_reorder = ['vega_change_ce', 'vega_change_pe', 'theta_change_ce', 'theta_change_pe', 'delta_change_ce', 'delta_change_pe']
                                                    bn_ce_f_i = bn_ce_f_i[columns_to_reorder]; bn_ne_f_i = bn_ne_f_i[columns_to_reorder]; n_ce_f_i = n_ce_f_i[columns_to_reorder]; n_ne_f_i = n_ne_f_i[columns_to_reorder]; f_ne_f_i = f_ne_f_i[columns_to_reorder]; m_ne_f_i = m_ne_f_i [columns_to_reorder]; s_ne_f_i = s_ne_f_i[columns_to_reorder]; ban_ne_f_i = ban_ne_f_i[columns_to_reorder]
                                                    print("reorder done")
                                                    print("1 columns",bn_ce_f_i.columns)
                                                    print("1 columns",ban_ne_f_i.columns)
                                                    bn_ce_f_i['theta_sentiment_ce'] = bn_ce_f_i['theta_change_ce'].apply(determine_theta_sentiment); bn_ce_f_i['theta_sentiment_pe'] = bn_ce_f_i['theta_change_pe'].apply(determine_theta_sentiment); bn_ne_f_i['theta_sentiment_ce'] = bn_ne_f_i['theta_change_ce'].apply(determine_theta_sentiment); bn_ne_f_i['theta_sentiment_pe'] = bn_ne_f_i['theta_change_pe'].apply(determine_theta_sentiment); n_ce_f_i['theta_sentiment_ce'] = n_ce_f_i['theta_change_ce'].apply(determine_theta_sentiment); n_ce_f_i['theta_sentiment_pe'] = n_ce_f_i['theta_change_pe'].apply(determine_theta_sentiment); n_ne_f_i['theta_sentiment_ce'] = n_ne_f_i['theta_change_ce'].apply(determine_theta_sentiment); n_ne_f_i['theta_sentiment_pe'] = n_ne_f_i['theta_change_pe'].apply(determine_theta_sentiment); f_ne_f_i['theta_sentiment_ce'] = f_ne_f_i['theta_change_ce'].apply(determine_theta_sentiment); f_ne_f_i['theta_sentiment_pe'] = f_ne_f_i['theta_change_pe'].apply(determine_theta_sentiment); m_ne_f_i['theta_sentiment_ce'] = m_ne_f_i['theta_change_ce'].apply(determine_theta_sentiment); m_ne_f_i['theta_sentiment_pe'] = m_ne_f_i['theta_change_pe'].apply(determine_theta_sentiment); s_ne_f_i['theta_sentiment_ce'] = s_ne_f_i['theta_change_ce'].apply(determine_theta_sentiment); s_ne_f_i['theta_sentiment_pe'] = s_ne_f_i['theta_change_pe'].apply(determine_theta_sentiment); ban_ne_f_i['theta_sentiment_pe'] = ban_ne_f_i['theta_change_pe'].apply(determine_theta_sentiment); print("theta sentiment done");
                                                    ban_ne_f_i['theta_sentiment_ce'] = ban_ne_f_i['theta_change_ce'].apply(determine_theta_sentiment)
                                                    print("2 columns",bn_ce_f_i.columns)
                                                    print("2 columns",ban_ne_f_i.columns)
                                                
                                                    bn_ce_f_i['delta_sentiment_ce'] = bn_ce_f_i['delta_change_ce'].apply(determine_delta_sentiment_ce); bn_ce_f_i['delta_sentiment_pe'] = bn_ce_f_i['delta_change_pe'].apply(determine_delta_sentiment_pe); bn_ne_f_i['delta_sentiment_ce'] = bn_ne_f_i['delta_change_ce'].apply(determine_delta_sentiment_ce); bn_ne_f_i['delta_sentiment_pe'] = bn_ne_f_i['delta_change_pe'].apply(determine_delta_sentiment_pe); n_ce_f_i['delta_sentiment_ce'] = n_ce_f_i['delta_change_ce'].apply(determine_delta_sentiment_ce); n_ce_f_i['delta_sentiment_pe'] = n_ce_f_i['delta_change_pe'].apply(determine_delta_sentiment_pe); n_ne_f_i['delta_sentiment_ce'] = n_ne_f_i['delta_change_ce'].apply(determine_delta_sentiment_ce); n_ne_f_i['delta_sentiment_pe'] = n_ne_f_i['delta_change_pe'].apply(determine_delta_sentiment_pe); f_ne_f_i['delta_sentiment_ce'] = f_ne_f_i['delta_change_ce'].apply(determine_delta_sentiment_ce); f_ne_f_i['delta_sentiment_pe'] = f_ne_f_i['delta_change_pe'].apply(determine_delta_sentiment_pe); m_ne_f_i['delta_sentiment_ce'] = m_ne_f_i['delta_change_ce'].apply(determine_delta_sentiment_ce); m_ne_f_i['delta_sentiment_pe'] = m_ne_f_i['delta_change_pe'].apply(determine_delta_sentiment_pe); s_ne_f_i['delta_sentiment_ce'] = s_ne_f_i['delta_change_ce'].apply(determine_delta_sentiment_ce); s_ne_f_i['delta_sentiment_pe'] = s_ne_f_i['delta_change_pe'].apply(determine_delta_sentiment_pe); ban_ne_f_i['delta_sentiment_pe'] = ban_ne_f_i['delta_change_pe'].apply(determine_delta_sentiment_pe); print("delta sentiment done");
                                                    ban_ne_f_i['delta_sentiment_ce'] = ban_ne_f_i['delta_change_ce'].apply(determine_theta_sentiment)
                                                    
                                                    bn_ce_f_i['vega_sentiment_ce'] = bn_ce_f_i['vega_change_ce'].apply(determine_vega_sentiment); bn_ce_f_i['vega_sentiment_pe'] = bn_ce_f_i['vega_change_pe'].apply(determine_vega_sentiment); bn_ne_f_i['vega_sentiment_ce'] = bn_ne_f_i['vega_change_ce'].apply(determine_vega_sentiment); bn_ne_f_i['vega_sentiment_pe'] = bn_ne_f_i['vega_change_pe'].apply(determine_vega_sentiment); n_ce_f_i['vega_sentiment_ce'] = n_ce_f_i['vega_change_ce'].apply(determine_vega_sentiment); n_ce_f_i['vega_sentiment_pe'] = n_ce_f_i['vega_change_pe'].apply(determine_vega_sentiment); n_ne_f_i['vega_sentiment_ce'] = n_ne_f_i['vega_change_ce'].apply(determine_vega_sentiment); n_ne_f_i['vega_sentiment_pe'] = n_ne_f_i['vega_change_pe'].apply(determine_vega_sentiment); f_ne_f_i['vega_sentiment_ce'] = f_ne_f_i['vega_change_ce'].apply(determine_vega_sentiment); f_ne_f_i['vega_sentiment_pe'] = f_ne_f_i['vega_change_pe'].apply(determine_vega_sentiment); m_ne_f_i['vega_sentiment_ce'] = m_ne_f_i['vega_change_ce'].apply(determine_vega_sentiment); m_ne_f_i['vega_sentiment_pe'] = m_ne_f_i['vega_change_pe'].apply(determine_vega_sentiment); s_ne_f_i['vega_sentiment_ce'] = s_ne_f_i['vega_change_ce'].apply(determine_vega_sentiment); s_ne_f_i['vega_sentiment_pe'] = s_ne_f_i['vega_change_pe'].apply(determine_vega_sentiment); ban_ne_f_i['vega_sentiment_pe'] = ban_ne_f_i['vega_change_pe'].apply(determine_vega_sentiment); print("vega sentiment done");
                                                    ban_ne_f_i['vega_sentiment_ce'] = ban_ne_f_i['vega_change_ce'].apply(determine_theta_sentiment)
                                                    
                                                    bn_ce_f_i['Time'] = tm; bn_ne_f_i['Time'] = tm; n_ce_f_i['Time'] = tm; n_ne_f_i['Time'] = tm; f_ne_f_i['Time'] = tm; m_ne_f_i['Time'] = tm; s_ne_f_i['Time'] = tm; ban_ne_f_i['Time'] = tm; print("entering into add decay");
                                                    bn_ce_f_i = add_decay_and_percentage_changes(bn_ce_f_i); bn_ne_f_i = add_decay_and_percentage_changes(bn_ne_f_i); n_ce_f_i = add_decay_and_percentage_changes(n_ce_f_i); n_ne_f_i = add_decay_and_percentage_changes(n_ne_f_i); f_ne_f_i = add_decay_and_percentage_changes(f_ne_f_i); m_ne_f_i = add_decay_and_percentage_changes(m_ne_f_i); s_ne_f_i = add_decay_and_percentage_changes(s_ne_f_i); ban_ne_f_i = add_decay_and_percentage_changes(ban_ne_f_i); print("add decay over");
                                                    print(ban_ne_f_i.columns)
                                                
                                                    
                                                    columns_to_reorder = ['Time','vega_change_ce','vega_sentiment_ce', 'vega_change_pe','vega_sentiment_pe','vega_decay','vega_%_change', 'theta_change_ce','theta_sentiment_ce', 'theta_change_pe','theta_sentiment_pe','theta_decay' ,'theta_%_change','delta_change_ce','delta_sentiment_ce', 'delta_change_pe','delta_sentiment_pe','delta_decay','delta_%_change']
                                                    # bn_ce_f = bn_ce_f[columns_to_reorder]; bn_ne_f = bn_ce_f[columns_to_reorder]; n_ce_f = bn_ce_f[columns_to_reorder]; n_ne_f = bn_ce_f[columns_to_reorder]; f_ne_f = bn_ce_f[columns_to_reorder]; m_ne_f = bn_ce_f[columns_to_reorder]; s_ne_f = bn_ce_f[columns_to_reorder]
                                                    bn_ce_f_i = bn_ce_f_i[columns_to_reorder]; bn_ne_f_i = bn_ne_f_i[columns_to_reorder]; n_ce_f_i = n_ce_f_i[columns_to_reorder]; n_ne_f_i = n_ne_f_i[columns_to_reorder]; f_ne_f_i = f_ne_f_i[columns_to_reorder]; m_ne_f_i = m_ne_f_i [columns_to_reorder]; s_ne_f_i = s_ne_f_i[columns_to_reorder];ban_ne_f_i = ban_ne_f_i[columns_to_reorder]
                                                    
                                                    bn_ce_f_i.to_csv("ltest.csv")
                                                    print("gonna concat")
                                                    bank_cw.to_csv(" bank_cw.csv")
                                                    bank_cw = pd.concat([bank_cw,bn_ce_f_i], ignore_index=True)
                                                    bank_cw.to_csv(" bank_cw2.csv")
                                                
                                                    bank_nw = pd.concat([bank_nw,bn_ne_f_i], ignore_index=True)
                                                    nif_cw = pd.concat([nif_cw,n_ce_f_i], ignore_index=True)
                                                    nif_nw = pd.concat([nif_nw,n_ne_f_i], ignore_index=True)
                                                    mid_cw = pd.concat([mid_cw,m_ne_f_i], ignore_index=True)
                                                    fin_cw = pd.concat([fin_cw,f_ne_f_i], ignore_index=True)
                                                    sen_cw = pd.concat([sen_cw,s_ne_f_i], ignore_index=True)
                                                    ban_cw = pd.concat([ban_cw,ban_ne_f_i], ignore_index=True)
                                                    print("gonna update")
                                                    bank_cw,starting_cells_d,f_time_d=filter_and_export(starting_cells_d,f_time_d,bank_cw,"BANKNIFTY", "current","main",expiry_today)
                                                    #bank_nw,starting_cells_d,f_time_d=filter_and_export(starting_cells_d,f_time_d,bank_nw,"BANKNIFTY", "next","main",expiry_today)
                                                    nif_cw,starting_cells_d,f_time_d=filter_and_export(starting_cells_d,f_time_d,nif_cw,"NIFTY", "current","main",expiry_today)
                                                    #nif_nw,starting_cells_d,f_time_d=filter_and_export(starting_cells_d,f_time_d,nif_nw,"NIFTY", "next","main",expiry_today)
                                                    mid_cw,starting_cells_d,f_time_d=filter_and_export(starting_cells_d,f_time_d,mid_cw,"MIDCPNIFTY", "current","main",expiry_today)
                                                    fin_cw,starting_cells_d,f_time_d=filter_and_export(starting_cells_d,f_time_d,fin_cw,"FINNIFTY","current","main",expiry_today)
                                                    sen_cw,starting_cells_d,f_time_d=filter_and_export(starting_cells_d,f_time_d,sen_cw,"SENSEX", "current","main",expiry_today)
                                                    f_time_d=False
                                                
                                                    bn_ce_f=pd.DataFrame();n_ce_f=pd.DataFrame();bn_ne_f=pd.DataFrame();n_ne_f=pd.DataFrame();f_ne_f=pd.DataFrame();m_ne_f=pd.DataFrame();s_ne_f=pd.DataFrame();ban_ne_f=pd.DataFrame()
                                                    bn_ce_f.to_csv('bn_ce_f3.csv', index=False)
                                                
                                                    
                                                
                                                #   save_option_chains_to_csv(merged_opt_chain )
                                                print("option chain formation over") 
                                                print("value of i is,",i)
                                                #   try:
                                                #     symbols_quote = ["NSE:HDFCBANK", "NSE:ICICIBANK", "NSE:SBIN", "NSE:AXISBANK", "NSE:KOTAKBANK", "NSE:INDUSINDBK"]
                                                #     data = kite.quote(symbols_quote)
                                                #     volumes = {}
                                                #     fo
                                                # r symbol in symbols_quote:  # Ensure you're iterating over symbols_quote
                                                #         volumes[symbol] = data[symbol]['volume']
                                                #     print("Volumes collected successfully.")
                                                #   except Exception as e:
                                                #         print("Error in volume", e)

                                                #   noted_ce_oi_corrected = {}
                                                #   noted_pe_oi_corrected = {}
                                                #   symbols=["HDFCBANK","ICICIBANK","SBIN","AXISBANK","KOTAKBANK","INDUSINDBK"]
                                                
                                                #   for symbol in symbols:
                                                        
                                                #         ce_oi_total_corrected = first_df[(first_df['name'] == symbol) & (first_df['CE_PE'] == 'CE')]['oi'].sum()
                                                #         pe_oi_total_corrected = first_df[(first_df['name'] == symbol) & (first_df['CE_PE'] == 'PE')]['oi'].sum()
                                                #         noted_ce_oi_corrected[symbol] = ce_oi_total_corrected
                                                #         noted_pe_oi_corrected[symbol] = pe_oi_total_corrected
                                                #   print("oi corrected over")

                                                #   bank_dfs = create_bank_dfs(volumes, prev_vol_df, noted_ce_oi_corrected, noted_pe_oi_corrected, result_df,tm)
                                                #   print("bankdfs over")
                                                #   hdfcbank_df, icicibank_df, sbin_df, axisbank_df, kotakbank_df, indusindbk_df = (bank_dfs[symbol] for symbol in ["HDFCBANK", "ICICIBANK", "SBIN", "AXISBANK", "KOTAKBANK", "INDUSINDBK"])
                                                #   print("individual bankdf extracted")
                                                # #   hdfcbank_df.to_csv('hdfcbank_df.csv', index=False)
                                                # #   icicibank_df.to_csv('icicibank_df.csv', index=False)
                                                # #   sbin_df.to_csv('sbin_df.csv', index=False)
                                                # #   axisbank_df.to_csv('axisbank_df.csv', index=False)
                                                # #   kotakbank_df.to_csv('kotakbank_df.csv', index=False)
                                                # #   indusindbk_df.to_csv('indusindbk_df.csv', index=False)
                                                #   update_option_chain(hdfcbank_df,"HDFC", "BNF_ProTrading_CW")
                                                #   update_option_chain(icicibank_df,"ICICI", "BNF_ProTrading_CW")
                                                #   update_option_chain(sbin_df,"SBI", "BNF_ProTrading_CW")
                                                #   update_option_chain(kotakbank_df,"KOTAK", "BNF_ProTrading_CW")
                                                #   update_option_chain(axisbank_df,"AXIS", "BNF_ProTrading_CW")
                                                #   update_option_chain(indusindbk_df,"INDUSIND", "BNF_ProTrading_CW")
                                                
                                                #   if i==0:
                                                #     prev_tick=result_df.copy()
                                                #     prev_tick.to_csv("previoustick.csv")

                                                #   else:
                                                #     current_df=result_df.copy()
                                                #     previous_df=prev_tick.copy()
                                                #     current_df = standardize_date_format(current_df, 'expiry')
                                                #     previous_df = standardize_date_format(previous_df, 'expiry')
                                                #     print("hiii")
                                                #     print(current_df.columns)
                                                #     print(previous_df.columns)
                                                #     results_combined =process_conditions_separately(current_df, previous_df,tm,atm_values,expiry_today)
                                                #     print("hiiii2")
                                                #     for key, hdf in results_combined.items():
                                                #         file_path = f"{key}.csv"
                                                #         hdf.to_csv(file_path, index=False)
                                                        

                                                #     key_to_initial = {
                                                #         'banknifty_current_increase': 'initial_banknifty_current_expiry_increase',
                                                #         'banknifty_current_decrease': 'initial_banknifty_current_expiry_decrease',
                                                #         'banknifty_next_increase': 'initial_banknifty_next_expiry_increase',
                                                #         'banknifty_next_decrease': 'initial_banknifty_next_expiry_decrease',
                                                #         'nifty_current_increase': 'initial_nifty_current_expiry_increase',
                                                #         'nifty_current_decrease': 'initial_nifty_current_expiry_decrease',
                                                #         'nifty_next_increase': 'initial_nifty_next_expiry_increase',
                                                #         'nifty_next_decrease': 'initial_nifty_next_expiry_decrease',
                                                #         'finnifty_current_increase': 'initial_finnifty_current_expiry_increase',
                                                #         'finnifty_current_decrease': 'initial_finnifty_current_expiry_decrease',
                                                #         'midcpnifty_current_increase': 'initial_midcpnifty_current_expiry_increase',
                                                #         'midcpnifty_current_decrease': 'initial_midcpnifty_current_expiry_decrease',
                                                #         'sensex_current_increase': 'initial_sensex_current_expiry_increase',
                                                #         'sensex_current_decrease': 'initial_sensex_current_expiry_decrease'
                                                #     }

                                                #     for key, hdf in results_combined.items():
                                                #         initial_df_name = key_to_initial[key]

                                                #         if initial_df_name not in globals():
                                                #             globals()[initial_df_name] = pd.DataFrame()

                                                #         hdf = update_strike_counts(hdf)
                                                #         globals()[initial_df_name] = pd.concat([globals()[initial_df_name], hdf], ignore_index=True)
                                                #         globals()[initial_df_name] = update_strike_counts(globals()[initial_df_name])
                                                #         file_path = f"new_{initial_df_name}.csv"
                                                #         globals()[initial_df_name].to_csv(file_path, index=False)
                                                        


                                                #         file_name = 'BNF_Pro_Trading.xlsx'
                                                #         sheet_name = 'VEGA & Theta_Chng_CW'
                                                #         mode = 'current'  
                                                #         key = f"{file_name} - {sheet_name}"
                                                #         if key in starting_cells_info:
                                                #             mode_info = starting_cells_info[key]
                                                #             starting_cell = mode_info.get(mode, 'Not Defined')  # Get the cell for the 'current' mode

                                                        
                                                #         filter_and_export_to_excel(starting_cell,f_time,initial_banknifty_current_expiry_increase, "BANKNIFTY", "current","VEGA & Theta_Chng_CW",expiry_today)
                                                        
                                                #         file_name = 'BNF_Pro_Trading.xlsx'
                                                #         sheet_name = 'VEGA & Theta_Chng_CW'
                                                #         mode = 'next'  
                                                #         key = f"{file_name} - {sheet_name}"
                                                #         if key in starting_cells_info:
                                                #             mode_info = starting_cells_info[key]
                                                #             starting_cell = mode_info.get(mode, 'Not Defined')
                                                #         filter_and_export_to_excel(starting_cell,f_time,initial_banknifty_current_expiry_decrease, "BANKNIFTY", "next","VEGA & Theta_Chng_CW",expiry_today)
                                                        
                                                #         file_name = 'BNF_Pro_Trading.xlsx'
                                                #         sheet_name = 'VEGA & Theta_Chng_NW'
                                                #         mode = 'current'  
                                                #         key = f"{file_name} - {sheet_name}"
                                                #         if key in starting_cells_info:
                                                #             mode_info = starting_cells_info[key]
                                                #             starting_cell = mode_info.get(mode, 'Not Defined')
                                                        
                                                #         filter_and_export_to_excel(starting_cell,f_time,initial_banknifty_next_expiry_increase, "BANKNIFTY", "current","VEGA & Theta_Chng_NW",expiry_today)
                                                #         file_name = 'BNF_Pro_Trading.xlsx'
                                                #         sheet_name = 'VEGA & Theta_Chng_NW'
                                                #         mode = 'next'  
                                                #         key = f"{file_name} - {sheet_name}"
                                                #         if key in starting_cells_info:
                                                #             mode_info = starting_cells_info[key]
                                                #             starting_cell = mode_info.get(mode, 'Not Defined')
                                                #         filter_and_export_to_excel(starting_cell,f_time,initial_banknifty_next_expiry_decrease, "BANKNIFTY", "next","VEGA & Theta_Chng_NW",expiry_today)
                                                        
                                                #         file_name = 'NF_Pro_Trading.xlsx'
                                                #         sheet_name = 'VEGA & Theta_Chng_CW'
                                                #         mode = 'current'  
                                                #         key = f"{file_name} - {sheet_name}"
                                                #         if key in starting_cells_info:
                                                #             mode_info = starting_cells_info[key]
                                                #             starting_cell = mode_info.get(mode, 'Not Defined')
                                                #         filter_and_export_to_excel(starting_cell,f_time,initial_nifty_current_expiry_increase, "NIFTY", "current","VEGA & Theta_Chng_CW",expiry_today)
                                                        
                                                #         file_name = 'NF_Pro_Trading.xlsx'
                                                #         sheet_name = 'VEGA & Theta_Chng_CW'
                                                #         mode = 'next'  
                                                #         key = f"{file_name} - {sheet_name}"
                                                #         if key in starting_cells_info:
                                                #             mode_info = starting_cells_info[key]
                                                #             starting_cell = mode_info.get(mode, 'Not Defined')
                                                #         filter_and_export_to_excel(starting_cell,f_time,initial_nifty_current_expiry_decrease, "NIFTY", "next","VEGA & Theta_Chng_CW",expiry_today)
                                                        
                                                #         file_name = 'NF_Pro_Trading.xlsx'
                                                #         sheet_name = 'VEGA & Theta_Chng_NW'
                                                #         mode = 'current'  
                                                #         key = f"{file_name} - {sheet_name}"
                                                #         if key in starting_cells_info:
                                                #             mode_info = starting_cells_info[key]
                                                #             starting_cell = mode_info.get(mode, 'Not Defined')
                                                #         filter_and_export_to_excel(starting_cell,f_time,initial_nifty_next_expiry_increase, "NIFTY", "current","VEGA & Theta_Chng_NW",expiry_today)
                                                #         file_name = 'NF_Pro_Trading.xlsx'
                                                #         sheet_name = 'VEGA & Theta_Chng_NW'
                                                #         mode = 'next'  
                                                #         key = f"{file_name} - {sheet_name}"
                                                #         if key in starting_cells_info:
                                                #             mode_info = starting_cells_info[key]
                                                #             starting_cell = mode_info.get(mode, 'Not Defined')
                                                        
                                                #         filter_and_export_to_excel(starting_cell,f_time,initial_nifty_next_expiry_decrease, "NIFTY", "next","VEGA & Theta_Chng_NW",expiry_today)
                                                #         file_name = 'FIN_Pro_Trading.xlsx'
                                                #         sheet_name = 'VEGA & Theta_Chng_CW'
                                                #         mode = 'current'  
                                                #         key = f"{file_name} - {sheet_name}"
                                                #         if key in starting_cells_info:
                                                #             mode_info = starting_cells_info[key]
                                                #             starting_cell = mode_info.get(mode, 'Not Defined')
                                                #         filter_and_export_to_excel(starting_cell,f_time,initial_finnifty_current_expiry_increase, "FINNIFTY", "current","VEGA & Theta_Chng_CW",expiry_today)
                                                        
                                                #         file_name = 'FIN_Pro_Trading.xlsx'
                                                #         sheet_name = 'VEGA & Theta_Chng_CW'
                                                #         mode = 'next'  
                                                #         key = f"{file_name} - {sheet_name}"
                                                #         if key in starting_cells_info:
                                                #             mode_info = starting_cells_info[key]
                                                #             starting_cell = mode_info.get(mode, 'Not Defined')
                                                #         filter_and_export_to_excel(starting_cell,f_time,initial_finnifty_current_expiry_decrease, "FINNIFTY", "next","VEGA & Theta_Chng_CW",expiry_today)
                                                        
                                                #         file_name = 'Midcp_Pro_Trading.xlsx'
                                                #         sheet_name = 'VEGA & Theta_Chng_CW'
                                                #         mode = 'current'  
                                                #         key = f"{file_name} - {sheet_name}"
                                                #         if key in starting_cells_info:
                                                #             mode_info = starting_cells_info[key]
                                                #             starting_cell = mode_info.get(mode, 'Not Defined')
                                                #         filter_and_export_to_excel(starting_cell,f_time,initial_midcpnifty_current_expiry_increase, "MIDCPNIFTY", "current","VEGA & Theta_Chng_CW",expiry_today)
                                                        
                                                #         file_name = 'Midcp_Pro_Trading.xlsx'
                                                #         sheet_name = 'VEGA & Theta_Chng_CW'
                                                #         mode = 'next'  
                                                #         key = f"{file_name} - {sheet_name}"
                                                #         if key in starting_cells_info:
                                                #             mode_info = starting_cells_info[key]
                                                #             starting_cell = mode_info.get(mode, 'Not Defined')
                                                #         filter_and_export_to_excel(starting_cell,f_time,initial_midcpnifty_current_expiry_decrease, "MIDCPNIFTY", "next","VEGA & Theta_Chng_CW",expiry_today)
                                                        
                                                #         file_name = 'Sen_Pro_Trading.xlsx'
                                                #         sheet_name = 'VEGA & Theta_Chng_CW'
                                                #         mode = 'current'  
                                                #         key = f"{file_name} - {sheet_name}"
                                                #         if key in starting_cells_info:
                                                #             mode_info = starting_cells_info[key]
                                                #             starting_cell = mode_info.get(mode, 'Not Defined')
                                                #         filter_and_export_to_excel(starting_cell,f_time,initial_sensex_current_expiry_increase, "SENSEX", "current","VEGA & Theta_Chng_CW",expiry_today)
                                                        
                                                #         file_name = 'Sen_Pro_Trading.xlsx'
                                                #         sheet_name = 'VEGA & Theta_Chng_CW'
                                                #         mode = 'next'  
                                                #         key = f"{file_name} - {sheet_name}"
                                                #         if key in starting_cells_info:
                                                #             mode_info = starting_cells_info[key]
                                                #             starting_cell = mode_info.get(mode, 'Not Defined')
                                                #         filter_and_export_to_excel(starting_cell,f_time,initial_sensex_current_expiry_decrease, "SENSEX", "next","VEGA & Theta_Chng_CW",expiry_today)
                                                        



                                                #     prev_tick=result_df.copy()
                                                print("df overrr")
                                                pro_df=pro_result_df.copy()
                                                fut_df=pro_result_df.copy()
                                                pro_result_df.to_csv("pro_result_df.csv")
                                                pro_df = aggregate_based_on_atm(pro_df,atm_values)
                                                pro_df.to_csv("prodf1.csv")
                                                names_to_filter = ['BANKNIFTY', 'NIFTY', 'FINNIFTY', 'MIDCPNIFTY', 'SENSEX',"BANKEX"]
                                                pro_df = pro_df [pro_df['Name'].isin(names_to_filter)]
                                                
                                                pro_df['Net_ITM'] = pro_df['Buy_ITM'] - pro_df['Sell_ITM']
                                                pro_df['Net_OTM'] = pro_df['Buy_OTM'] - pro_df['Sell_OTM']
                                                pro_df['final_buy']=pro_df['Net_ITM']+pro_df['Net_OTM']
                                                pro_df['final_sell']= pro_df['Sell_ITM']+pro_df['Sell_OTM']
                                                pro_df.to_csv("prodf2.csv")
                                                print("final sell over")
                                                pivot_df = pro_df.pivot_table(index=['Name', 'Expiry'], columns='CE_PE', values=['final_buy', 'final_sell'], aggfunc='first')
                                                pivot_df['result_buy'] = pivot_df[('final_buy', 'CE')] / pivot_df[('final_buy', 'PE')]
                                                pivot_df['result_sell'] = pivot_df[('final_sell', 'PE')] / pivot_df[('final_sell', 'CE')]
                                                result_pro_df = pivot_df.reset_index()
                                                result_pro_df.columns = ['Name', 'Expiry', 'Buy_CE', 'Buy_PE', 'Sell_CE', 'Sell_PE', 'result_buy', 'result_sell']
                                                result_pro_df = result_pro_df[['Name', 'Expiry', 'result_buy', 'result_sell']]
                                                result_pro_df['result_buy'] = (result_pro_df['result_buy'] - 1) * 100
                                                result_pro_df['result_sell'] = (result_pro_df['result_sell'] - 1) * 100
                                                result_pro_df.replace([float('inf'), float('-inf'), pd.NA], 0, inplace=True)   
                                                result_pro_df.to_csv("prodf.csv")
                                                print("prodf formation over")

                                                if first_time:
                                                    print("entering into if first time")
                                                    trial_df=result_pro_df.copy()
                                                    trial2_df=new_result_df.copy()
                                                    trial_df.rename(columns={'Name':'name','Expiry':'expiry','result_buy': 'protrader_data', 'result_sell': 'retailtrader_data'}, inplace=True)
                                                    trial_df['protrader_senti'] = trial_df['protrader_data'].apply(determine_sentiment)
                                                    trial_df['retailtrader_senti'] = trial_df['retailtrader_data'].apply(determine_sentiment)
                                                    # new_tm = subtract_one_minute(tm)
                                                    print("time inside pro first time is",tm)

                                                    # trial_df['Time'] = tm
                                                    # column_order = ['name', 'expiry', 'Time', 'protrader_data', 'protrader_senti', 'retailtrader_data', 'retailtrader_senti']
                                                    column_order = ['name', 'expiry', 'protrader_data', 'protrader_senti', 'retailtrader_data', 'retailtrader_senti']
                                                    trial_df= trial_df[column_order]                            
                                                    trial2_df = trial2_df.groupby(['name', 'CE_PE', 'expiry']).agg({
                                                        'Totl Bid Qty': 'sum',
                                                        'Totl Ask Qty': 'sum'
                                                    }).reset_index()
                                                    print(future_df.columns)
                                                    future_df['Fut_Atv']=future_df['Totl Bid Qty']-future_df['Totl Ask Qty']
                                                    future_df=future_df[['name','Fut_Atv']]
                                                    trial2_df['ATV'] = trial2_df['Totl Bid Qty'] - trial2_df['Totl Ask Qty']
                                                    trial2_df = trial2_df[['name', 'CE_PE', 'expiry', 'ATV']]
                                                    pivot_df = trial2_df.pivot_table(index=['name', 'expiry'], columns='CE_PE', values='ATV', aggfunc='first').reset_index()
                                                    pivot_df.columns = ['name', 'expiry', 'ATV_CE', 'ATV_PE']
                                                    merged_df = trial_df.merge(pivot_df, on=['name', 'expiry'], how='left')
                                                    merged_df['CE_PE']=merged_df['ATV_CE']-merged_df['ATV_PE']
                                                    merged_df= merged_df.merge(future_df, on=['name'], how='left')
                                                    merged_df.dropna(subset=['ATV_CE'], inplace=True)
                                                    atv_df= pd.concat([atv_df, merged_df], ignore_index=True)
                                                    atv_df.to_csv("atv_df.csv")
                                                    print("saving atv df")
                                                    
                                                else:
                                                    if not atv_df.empty:
                                                        print("time is",datetime.datetime.now())                              
                                                        grouped_df = atv_df.groupby(['name', 'expiry']).agg({
                                                                'ATV_CE': 'mean',
                                                                'ATV_PE': 'mean',
                                                                'protrader_data': 'mean',
                                                                'retailtrader_data': 'mean',
                                                                # 'Time': 'last',
                                                                'CE_PE': 'mean', 
                                                                'Fut_Atv':'mean'

                                                            }).reset_index()
                                                        grouped_df['Time']=tm
                                                        grouped_df['protrader_senti'] = grouped_df['protrader_data'].apply(determine_sentiment)
                                                        grouped_df['retailtrader_senti'] = grouped_df['retailtrader_data'].apply(determine_sentiment)
                                                        column_order = ['name', 'expiry', 'Time', 'protrader_data', 'protrader_senti', 'retailtrader_data', 'retailtrader_senti','ATV_CE','ATV_PE','CE_PE','Fut_Atv']
                                                        grouped_df= grouped_df[column_order] 
                                                        grouped_df.to_csv("grouped_df.csv")
                                                        print("entering into update and calculate changes")
                                                        empty_df=update_and_calculate_changes(grouped_df,empty_df)
                                                        empty_df['overall']=empty_df['protrader_data']-empty_df['retailtrader_data']
                                                        empty_df['o_senti']=empty_df['overall'].apply(determine_sentiment)
                                                        print("over")
                                                        new_order = [
                                                                    'name', 'expiry', 'Time', 'protrader_data', 'protrader_senti',
                                                                    'retailtrader_data', 'retailtrader_senti','overall','o_senti','ATV_CE', 
                                                                    'ATV_PE','CE_PE','Fut_Atv']
                                                                
                                                        empty_df= empty_df[new_order]
                                                        empty_df.to_csv("empty_df.csv")
                                                        new_data=empty_df.copy()
                                                        new_data['expiry'] = pd.to_datetime(new_data['expiry'], format='%Y-%m-%d')
                                                        new_data['Time'] = new_data['Time'].astype(str)
                                                        dfs = {}
                                                        indices = ['BANKNIFTY', 'NIFTY', 'FINNIFTY', 'MIDCPNIFTY', 'SENSEX']
                                                        df_names = ['bnf_df', 'nf_df', 'fin_df', 'mid_df', 'sen_df']

                                                        for index, df_name in zip(indices, df_names):
                                                            index_df = new_data[new_data['name'] == index]
                                                            index_df = index_df.sort_values(by=['expiry', 'Time'], ascending=[True, False])
                                                            earliest_expiry = index_df['expiry'].iloc[0]
                                                            index_df = index_df[index_df['expiry'] == earliest_expiry].sort_values(by='Time', ascending=False).head(1)
                                                            index_df = index_df.drop(columns=['name', 'expiry'])
                                                            index_df = index_df.reset_index(drop=True)
                                                            globals()[df_name] = index_df
                                                    
                                                        bnf_new_df = pd.concat([bnf_new_df, bnf_df], ignore_index=True)
                                                        nf_new_df = pd.concat([nf_new_df, nf_df], ignore_index=True)
                                                        fin_new_df = pd.concat([fin_new_df, fin_df], ignore_index=True)
                                                        sen_new_df = pd.concat([sen_new_df, sen_df], ignore_index=True)
                                                        mid_new_df = pd.concat([mid_new_df, mid_df], ignore_index=True)
                                                        
                                                        bnf_new_df,starting_cells_pro,p_time = pro_export(starting_cells_pro,p_time, bnf_new_df, "BANKNIFTY", "current", 'Fii & PRO vs Retailers',expiry_today)
                                                        #empty_df,starting_cells_pro,p_time = pro_export(starting_cells_pro,p_time, nf_new_df, "BANKNIFTY", "next", 'Fii & PRO vs Retailers',expiry_today)
                                                        nf_new_df,starting_cells_pro,p_time = pro_export(starting_cells_pro,p_time, nf_new_df, "NIFTY", "current", 'Fii & PRO vs Retailers',expiry_today)
                                                        #empty_df,starting_cells_pro,p_time = pro_export(starting_cells_pro,p_time, empty_df, "NIFTY", "next", 'Fii & PRO vs Retailers',expiry_today)
                                                        fin_new_df,starting_cells_pro,p_time = pro_export(starting_cells_pro,p_time, fin_new_df, "FINNIFTY", "current", 'Fii & PRO vs Retailers',expiry_today)
                                                        mid_new_df,starting_cells_pro,p_time = pro_export(starting_cells_pro,p_time, mid_new_df, "MIDCPNIFTY", "current", 'Fii & PRO vs Retailers',expiry_today)
                                                        sen_new_df,starting_cells_pro,p_time = pro_export(starting_cells_pro,p_time, sen_new_df, "SENSEX", "current", 'Fii & PRO vs Retailers',expiry_today)
                                                        p_time=False
                                                        print("saving empty df")
                                                        atv_df=pd.DataFrame()
                                                #   pro_export_minute(merged_df,  "BANKNIFTY","current", 'BNF_ProTrading_CW',expiry_today)
                                                #   pro_export_minute(merged_df,  "BANKNIFTY","next", 'BNF_ProTrading_NW',expiry_today)
                                                #   pro_export_minute(merged_df,  "NIFTY","current", 'NF_ProTrading_CW',expiry_today)
                                                #   pro_export_minute(merged_df,  "NIFTY","current", 'NF_ProTrading_NW',expiry_today)
                                                #   pro_export_minute(merged_df,  "FINNIFTY","current", 'Fin_ProTrading_CW',expiry_today)
                                                #   pro_export_minute(merged_df,  "MIDCPNIFTY","current", 'Mid_ProTrading_CW',expiry_today)
                                                #   pro_export_minute(merged_df,  "SENSEX","current", 'Sen_ProTrading_CW',expiry_today)
                                                time.sleep(0.2)
                                                #   market_data_df = fetch_market_data()
                                                #   spot_update(market_data_df, 'NIFTY BANK')
                                                #   spot_update(market_data_df, 'NIFTY 50')
                                                #   spot_update(market_data_df, 'NIFTY FIN SERVICE')
                                                #   spot_update(market_data_df, 'NIFTY MID SELECT')
                                                #   spot_update(market_data_df, 'SENSEX')
                                                #   spot_update(market_data_df, 'INDIA VIX')
                                                now = datetime.datetime.now()
                                                
                                                exit_time = datetime.datetime.now()
                                                comparison_time = exit_time.replace(hour=15, minute=29, second=0, microsecond=0)

                                                if exit_time > comparison_time:
                                                    print("exit time is greater than comparison time")
                                                    print("time now is",exit_time)
                                                    source_files = ['BNF_theta_delta.xlsx', 'NF_theta_delta.xlsx','FIN_theta_delta.xlsx','Sen_theta_delta.xlsx','Midcp_theta_delta.xlsx','BAN_theta_delta.xlsx']
                                                    today_date = datetime.datetime.now().strftime('%Y%m%d')
                                                    new_folder_name = f"Theta_Delta_{today_date}"
                                                    if not os.path.exists(new_folder_name):
                                                        os.makedirs(new_folder_name)
                                                        print(f"Folder '{new_folder_name}' created.")
                                                    else:
                                                        print(f"Folder '{new_folder_name}' already exists.")
                                                    
                                                    for source_file in source_files:
                                                        if os.path.exists(source_file):
                                                            destination_file = os.path.join(new_folder_name, source_file)
                                                            shutil.copy(source_file, destination_file)
                                                            print(f"File '{source_file}' copied to '{new_folder_name}'.")
                                                        else:
                                                            print(f"Source file '{source_file}' does not exist.")

                                                    new_folder_name = f"Fii_PRO_{today_date}"
                                                    source_files = ['BNF_Fii_Pro.xlsx', 'Mid_Fii_Pro.xlsx','Sen_Fii_Pro.xlsx','NF_Fii_Pro.xlsx','FIN_Fii_Pro.xlsx','BAN_Fii_Pro.xlsx']

                                                    if not os.path.exists(new_folder_name):
                                                        os.makedirs(new_folder_name)
                                                        print(f"Folder '{new_folder_name}' created.")
                                                    else:
                                                        print(f"Folder '{new_folder_name}' already exists.")
                                                    
                                                    for source_file in source_files:
                                                        if os.path.exists(source_file):
                                                            destination_file = os.path.join(new_folder_name, source_file)
                                                            shutil.copy(source_file, destination_file)
                                                            print(f"File '{source_file}' copied to '{new_folder_name}'.")
                                                        else:
                                                            print(f"Source file '{source_file}' does not exist.")

                                                    exit()          
                                                i+=1                          
                                                print("going for next round")

                        
                                                    
                                            
            except Exception as e:
                print("exception occured main:" + str(e),datetime.datetime.now())
                #ogfile.flush(); os.fsync(#ogfile.fileno())
                
                

                
                pass
                            
                                                    